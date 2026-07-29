"""Use cases for Production indicators (производственные показатели).

Snapshot-based (like forensic). Raw natura/money stored; growth% and
execution% are computed HONESTLY here (audit lessons):
  • «флаг≠факт» — производные не берём из файла, пересчитываем из сырья.
  • 3-state исполнения: pct / «нет факта» / «нет плана».
  • reject plan≤0 (не даём отрицательному плану инвертировать знак).
  • overpar >110% — отдельная зона (фронт красит), не «успех».
  • суммы портфеля — по строке-итогу компании (не по продуктам → не задваиваем «в т.ч.»).
"""
from __future__ import annotations

import io
import logging
from typing import Any, Optional

from fastapi import HTTPException
from fastapi import status as http_status

from app.core.i18n import current_locale, tr
from app.uow.ports import UnitOfWorkABC

log = logging.getLogger(__name__)

SECTOR_COLOR = {
    "mining": "#9B8EC4", "oilgas": "#1D9E75", "energy": "#EF9F27",
    "transport": "#378ADD", "other": "#888780",
}
SECTOR_ORDER = {"mining": 0, "oilgas": 1, "energy": 2, "transport": 3, "other": 4}

# Excel sheet-name → company code aliases (короткие имена листов «Свода» не
# совпадают с полными name_ru в Company). Используется внутренним seed-импортом.
_ALIASES: dict[str, str] = {
    "НГМК": "ngmk", "Навоийуран": "nur", "Узметкомбинат": "umk", "Узбекуголь": "uug",
    "АГМК": "agmk", "Узтрансгаз": "utg", "Худудгазтаъминот": "hgt", "UzGasTrade": "ugt",
    "НЭС": "nes", "РЭС": "res", "ТЭС": "tes", "Узбекгидроэнерго": "uge",
    "Узбекистон темир йуллари": "uty", "Uzbekistan Airways": "uhy", "Uzbekistan Airports": "uap",
    "Тошшахартрансхизмат": "tst", "Узбектелеком": "utc", "Узкимёсаноат": "uks",
    "Навоийазот": "naz", "Узавтосаноат": "uas",
}


# ─── honest numeric helpers ───────────────────────────────────────

def _num(v: Any) -> Optional[float]:
    """Число или None. Терпит 'х'/'x'/'-'/'', пробелы и запятую-десятичную."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if s in ("", "х", "x", "-", "—", "n/a"):
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _is_number(v: Any) -> bool:
    return _num(v) is not None


def _growth(base: Any, cur: Any) -> Optional[float]:
    """Темп роста % = cur/base·100. Честно: None при base≤0/нет (не выдумываем)."""
    b, c = _num(base), _num(cur)
    if b is None or c is None or b <= 0:
        return None
    return round(c / b * 100, 1)


def _exec(plan: Any, expect: Any) -> tuple[Optional[float], str]:
    """Исполнение % = expect/plan·100 (3-state).
    (None,'noplan') — плана нет/≤0; (None,'nofact') — факта нет;
    (0.0,'pct') — факт=0 при плане>0 (провал, не «нет данных»)."""
    p, e = _num(plan), _num(expect)
    if p is None or p <= 0:
        return None, "noplan"
    if e is None:
        return None, "nofact"
    return round(e / p * 100, 1), "pct"


def _sector_group(code: Optional[str], name: Optional[str]) -> str:
    c = (code or "").lower().strip()
    if c in SECTOR_COLOR:
        return c
    n = (name or "").lower()
    if "нефт" in n or "газ" in n or "oil" in c or "gas" in c:
        return "oilgas"
    if "горн" in n or "metall" in n or "mining" in c or "мет" in n:
        return "mining"
    if "энерг" in n or "electr" in n or "energ" in c or "электр" in n:
        return "energy"
    if "транс" in n or "связ" in n or "коммун" in n or "trans" in c or "авиа" in n:
        return "transport"
    return "other"


def _norm(s: Any) -> str:
    """Нормализация имени для сопоставления лист↔компания."""
    t = str(s or "").lower().strip()
    t = t.translate(str.maketrans("ўёқҳғ", "уекхг"))
    for ch in "«»\"'`“”":
        t = t.replace(ch, "")
    return " ".join(t.split())


def _enrich_line(l: dict) -> dict:
    o = dict(l)
    o["baseN"], o["baseM"] = _num(l.get("baseN")), _num(l.get("baseM"))
    o["planN"], o["planM"] = _num(l.get("planN")), _num(l.get("planM"))
    o["expN"], o["expM"] = _num(l.get("expN")), _num(l.get("expM"))
    o["factN"], o["factM"] = _num(l.get("factN")), _num(l.get("factM"))
    # Результат периода = ФАКТ, если введён, иначе ОЖИДАЕМОЕ (прогноз). Темп и
    # исполнение считаем от результата → факт даёт реальное исполнение (факт/план),
    # без факта остаётся прогнозное (ожид/план). execKind сообщает фронту тип.
    has_fact = o["factM"] is not None or o["factN"] is not None
    resM = o["factM"] if o["factM"] is not None else o["expM"]
    resN = o["factN"] if o["factN"] is not None else o["expN"]
    o["growthM"] = _growth(o["baseM"], resM)
    o["growthN"] = _growth(o["baseN"], resN)
    # Исполнение по деньгам; если денежного плана нет — по натуре (компании
    # без денежного объёма: газоснабжение/пассажироперевозки).
    ep, es = _exec(o["planM"], resM)
    basis = "money"
    if es == "noplan" and o["planN"] is not None:
        ep, es = _exec(o["planN"], resN)
        basis = "natura"
    o["execPct"], o["execState"], o["execBasis"] = ep, es, basis
    o["execKind"] = "fact" if has_fact else "forecast"
    o["growthPct"] = o["growthM"] if o["growthM"] is not None else o["growthN"]
    return o


def _company_payload(code: str, raw: dict, m: dict) -> dict:
    """Единый per-company payload (используется overview и company_detail).
    Итог берём со строки-итога (total), продукты — в lines[]."""
    name = m.get("name_short") or m.get("name_ru") or raw.get("n") or code
    sector = _sector_group(m.get("sector_code"), m.get("sector_name"))
    lines = [_enrich_line(x) for x in (raw.get("lines") or []) if isinstance(x, dict)]
    total = next((x for x in lines if x.get("total")), lines[0] if lines else None)
    t = total or {}
    planM, expM, baseM = t.get("planM"), t.get("expM"), t.get("baseM")
    has = bool(total) and (
        _is_number(planM) or _is_number(expM)
        or _is_number(t.get("planN")) or _is_number(t.get("expN"))
    )
    return {
        "k": code, "n": name, "s": sector,
        "sector_color": SECTOR_COLOR.get(sector, SECTOR_COLOR["other"]),
        "unit": t.get("unit"),
        "baseM": baseM, "planM": planM, "expM": expM, "factM": t.get("factM"),
        "baseN": t.get("baseN"), "planN": t.get("planN"), "expN": t.get("expN"), "factN": t.get("factN"),
        "execPct": t.get("execPct"), "execState": t.get("execState", "noplan"),
        "execBasis": t.get("execBasis", "money"), "execKind": t.get("execKind", "forecast"),
        "growthPct": t.get("growthPct"),
        "lines": lines,
        "has_data": has,
    }


class ProductionService:
    def __init__(self, uow: UnitOfWorkABC) -> None:
        self.uow = uow

    # ─── overview ─────────────────────────────────────────────────
    async def overview(
        self, *, year: int, period: str,
        allowed_codes: Optional[set[str]] = None,
    ) -> dict[str, Any]:
        async with self.uow:
            snap = await self.uow.production.load_snapshot()
            meta = await self.uow.production.companies_meta()

        meta_by_code = {m["code"].lower(): m for m in meta}
        companies: list[dict] = []
        sum_plan = sum_exp = sum_fact = sum_result = 0.0
        with_data = over = under = ontarget = overpar = 0

        for raw in snap:
            if not isinstance(raw, dict):
                continue
            if raw.get("year") != year or (raw.get("period") or "h1") != period:
                continue
            code = (raw.get("k") or "").lower()
            if allowed_codes is not None and code not in allowed_codes:
                continue
            # meta_by_code (companies_meta) содержит только is_active=true — код,
            # которого там нет, = деактивированная (или несуществующая) компания.
            if code not in meta_by_code:
                continue
            c = _company_payload(code, raw, meta_by_code.get(code, {}))
            if c["has_data"]:
                with_data += 1
            if c["planM"] is not None:       # денежная сумма портфеля — только деньги
                sum_plan += c["planM"]
            if c["expM"] is not None:
                sum_exp += c["expM"]
            if c["factM"] is not None:
                sum_fact += c["factM"]
            resM = c["factM"] if c["factM"] is not None else c["expM"]
            if resM is not None:
                sum_result += resM
            exec_pct = c["execPct"]
            if exec_pct is not None:
                if exec_pct > 110:
                    overpar += 1
                    over += 1
                elif exec_pct >= 90:
                    ontarget += 1
                else:
                    under += 1
            companies.append(c)

        companies.sort(key=lambda c: (SECTOR_ORDER.get(c["s"], 99), c["n"]))
        # Портфельное исполнение — от результата (факт, где есть, иначе ожид).
        port_exec, _ = _exec(sum_plan, sum_result)
        kpis = {
            "present": len(companies),
            "with_data": with_data,
            "plan_total": round(sum_plan, 1),
            "expect_total": round(sum_exp, 1),
            "fact_total": round(sum_fact, 1),
            "exec_pct": port_exec,
            "over": over, "under": under, "ontarget": ontarget, "overpar": overpar,
        }
        return {"companies": companies, "kpis": kpis, "year": year, "period": period}

    # ─── one company (вкладка БП в карточке компании) ─────────────
    async def company_detail(
        self, code: str, *, year: int, period: str,
    ) -> dict[str, Any]:
        """Производство одной компании. Возвращает enriched-компанию (или
        пустую с has_data=False, если данных за период нет) + доступные именно
        у этой компании (year, period) — для локального селектора периода."""
        code_l = code.lower()
        async with self.uow:
            snap = await self.uow.production.load_snapshot()
            meta = await self.uow.production.companies_meta()
        m = next((x for x in meta if (x.get("code") or "").lower() == code_l), {})

        mine = [
            r for r in snap
            if isinstance(r, dict) and (r.get("k") or "").lower() == code_l and r.get("year")
        ]
        combos = sorted(
            {(r.get("year"), r.get("period") or "h1") for r in mine},
            key=lambda t: (-(t[0] or 0), t[1]),
        )
        years = sorted({y for (y, _) in combos}, reverse=True)

        raw = next(
            (r for r in mine
             if r.get("year") == year and (r.get("period") or "h1") == period),
            None,
        )
        if raw is not None:
            company = _company_payload(code_l, raw, m)
        else:
            name = m.get("name_short") or m.get("name_ru") or code_l
            sector = _sector_group(m.get("sector_code"), m.get("sector_name"))
            company = {
                "k": code_l, "n": name, "s": sector,
                "sector_color": SECTOR_COLOR.get(sector, SECTOR_COLOR["other"]),
                "unit": None, "baseM": None, "planM": None, "expM": None,
                "baseN": None, "planN": None, "expN": None,
                "execPct": None, "execState": "noplan", "execBasis": "money",
                "growthPct": None, "lines": [], "has_data": False,
            }
        return {
            "company": company, "year": year, "period": period,
            "years": years,
            "combos": [{"year": y, "period": p} for (y, p) in combos],
        }

    # ─── available periods (для селектора) ────────────────────────
    async def available(self) -> dict[str, Any]:
        async with self.uow:
            snap = await self.uow.production.load_snapshot()
        combos = sorted({
            (r.get("year"), r.get("period") or "h1")
            for r in snap if isinstance(r, dict) and r.get("year")
        }, key=lambda t: (-(t[0] or 0), t[1]))
        years = sorted({y for (y, _) in combos}, reverse=True)
        return {"years": years, "combos": [{"year": y, "period": p} for (y, p) in combos]}

    # ─── upsert one company (editor) ──────────────────────────────
    async def upsert_company(self, code: str, payload) -> dict[str, Any]:
        year = payload.year
        period = payload.period or "h1"
        lines = [
            {k: getattr(l, k) for k in
             ("name", "unit", "total", "parent", "baseN", "baseM",
              "planN", "planM", "expN", "expM", "factN", "factM")}
            for l in payload.lines
        ]
        async with self.uow:
            snap = await self.uow.production.load_snapshot()
            snap = [
                e for e in snap
                if not (isinstance(e, dict) and (e.get("k") or "").lower() == code.lower()
                        and e.get("year") == year and (e.get("period") or "h1") == period)
            ]
            snap.append({"k": code.lower(), "year": year, "period": period, "lines": lines})
            await self.uow.production.save_snapshot(snap)
        return {"ok": True, "code": code.lower(), "year": year, "period": period, "lines": len(lines)}

    async def get_company_label(self, code: str) -> str:
        async with self.uow:
            meta = await self.uow.production.companies_meta()
        for m in meta:
            if (m.get("code") or "").lower() == code.lower():
                return m.get("name_short") or m.get("name_ru") or code
        return code

    # ─── scope helper ─────────────────────────────────────────────
    async def resolve_codes_for_scope(self, scope_ids) -> set[str]:
        async with self.uow:
            return await self.uow.production.codes_for_company_ids(scope_ids)

    # ─── Excel import (свод бизнес-плана) ─────────────────────────
    async def import_xlsx(self, raw_bytes: bytes, *, year: int = 2026,
                          period: str = "h1") -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "openpyxl not installed")
        if len(raw_bytes) > 25 * 1024 * 1024:
            raise HTTPException(413, "Файл слишком большой (макс. 25 МБ).")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(
                400,
                tr("Не удалось разобрать xlsx: {error}", current_locale(), error=str(e)),
            )

        async with self.uow:
            meta = await self.uow.production.companies_meta()
            snap = await self.uow.production.load_snapshot()

        # name → code index (name_short/ru/uz/en/code + aliases)
        idx: dict[str, str] = {}
        for m in meta:
            for f in ("name_short", "name_ru", "name_uz", "name_en", "code"):
                v = m.get(f)
                if v:
                    idx.setdefault(_norm(v), m["code"].lower())
        for alias, cd in _ALIASES.items():
            idx.setdefault(_norm(alias), cd)

        new_entries: list[dict] = []
        unmatched: list[str] = []
        empty = 0
        for sheet in wb.sheetnames:
            code = idx.get(_norm(sheet))
            if not code:
                unmatched.append(sheet)
                continue
            lines = self._parse_sheet(wb[sheet])
            new_entries.append({"k": code, "year": year, "period": period, "lines": lines})
            if not any(_is_number(l.get("planM")) or _is_number(l.get("expM")) for l in lines):
                empty += 1

        codes_new = {e["k"] for e in new_entries}
        keep = [
            e for e in snap
            if not (isinstance(e, dict) and e.get("year") == year
                    and (e.get("period") or "h1") == period
                    and (e.get("k") or "").lower() in codes_new)
        ]
        keep.extend(new_entries)
        async with self.uow:
            await self.uow.production.save_snapshot(keep)

        return {
            "ok": True, "year": year, "period": period,
            "matched": len(new_entries),
            "with_data": len(new_entries) - empty,
            "empty": empty,
            "unmatched": unmatched,
            "lines_total": sum(len(e["lines"]) for e in new_entries),
        }

    @staticmethod
    def _parse_sheet(ws) -> list[dict]:
        """Разбор одного листа-компании: 3-строчная шапка, данные с 4-й строки.
        Колонки: A№ B имя C ед | D-F 2025факт(нат/деньги/темп) | G-I 2026план |
        J-L 2026ожид | M исполнение. Темп/исполнение из файла ИГНОРИРУЕМ."""
        lines: list[dict] = []
        max_r = ws.max_row or 0
        last_top: Optional[int] = None
        first = True
        for r in range(4, max_r + 1):
            b = ws.cell(r, 2).value
            name = str(b).strip() if b is not None else ""
            if not name:
                continue
            unit = ws.cell(r, 3).value
            line = {
                "name": name,
                "unit": (str(unit).strip() if unit not in (None, "") else None),
                "total": first,
                "parent": None,
                "baseN": _num(ws.cell(r, 4).value), "baseM": _num(ws.cell(r, 5).value),
                "planN": _num(ws.cell(r, 7).value), "planM": _num(ws.cell(r, 8).value),
                "expN": _num(ws.cell(r, 10).value), "expM": _num(ws.cell(r, 11).value),
            }
            idx = len(lines)
            if first:
                first = False
                lines.append(line)
                continue
            low = name.lower()
            is_child = (low.startswith("в том числе") or low.startswith("в т.ч")
                        or low.startswith("в тч") or low.startswith("из них"))
            if is_child and last_top is not None:
                line["parent"] = last_top
            else:
                last_top = idx
            lines.append(line)
        return lines
