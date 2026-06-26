"""Procurement xlsx import service — xarid 22-sheet format.

Parse → compute median per product_code → bulk-insert with benchmark fields.
Big import operation; logic preserved 1:1 from legacy.
"""
from __future__ import annotations

import io
import json
import re
import statistics
from datetime import date
from typing import Any

from fastapi import HTTPException
from fastapi import status as http_status
from pydantic import BaseModel

from app.uow.ports import UnitOfWorkABC


class PaImportSummary(BaseModel):
    """Result of bulk xlsx import — counts in/out для UI summary."""
    inserted: int
    skipped_no_company: int
    skipped_no_data: int
    sheets_processed: int
    product_codes: int
    benchmark_rows: int

# Sheet-name → company `code` (lowercase abbr; matches governance seed).
# Жёстко зашит здесь — единая точка истины для xarid 22-sheet формата.
_PA_SHEET_TO_CODE: dict[str, str] = {
    "NGMK": "ngmk", "NAVOIYURAN": "nur", "AGMK": "agmk", "UMK": "umk", "UUG": "uug",
    "UNG":  "ung",  "UTG":        "utg", "HGT":  "hgt",  "UGT": "ugt", "NES": "nes",
    "TES":  "tes",  "RES":        "res", "UGE":  "uge",  "UTY": "uty", "UHY": "uhy",
    "UAP":  "uap",  "UTC":        "utc", "TSHT": "tst",  "UPT": "upt", "UAS": "uas",
    "NAZ":  "naz",  "UKS":        "uks",
}

_PURCHASE_TYPE_NORM = {
    "E-SHOP":                    "e_shop",
    "E_STORE":                   "e_shop",
    "AUCTION":                   "auction",
    "BEST_OFFER":                "best_offer",
    "OTHER_COMPETITIVE_METHODS": "competitive",
}


def _norm_purchase_type(s):
    if not s:
        return None
    return _PURCHASE_TYPE_NORM.get(str(s).upper().strip(), str(s).strip()[:32])


def _parse_category_id(text_val):
    if text_val is None:
        return None
    try:
        return int(str(text_val).strip())
    except (TypeError, ValueError):
        return None


class ProcurementImportService:
    def __init__(self, uow: UnitOfWorkABC) -> None:
        self.uow = uow

    async def import_xlsx(self, raw_bytes: bytes) -> PaImportSummary:
        """Parse xlsx (22 sheets), compute median per productCode,
        bulk-insert into procurement_closures."""
        if not raw_bytes:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST, "Empty file",
            )

        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                "openpyxl not installed",
            )

        if len(raw_bytes) > 25 * 1024 * 1024:   # L-19: лимит размера загружаемого xlsx
            raise HTTPException(
                http_status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                "Файл слишком большой (макс. 25 МБ).",
            )
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                f"Failed to parse xlsx: {e}",
            )

        # Load company maps inside one short read tx
        async with self.uow:
            if not self.uow.procurement.closures_available:
                raise HTTPException(
                    http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                    "ProcurementClosure model unavailable",
                )
            by_code = await self.uow.procurement.get_companies_code_map()
            sector_by_co = await self.uow.procurement.get_sector_by_company_map()

        # 1st pass: parse sheets → collect dict-rows
        parsed, sheets_processed, skipped_no_company, skipped_no_data = \
            self._parse_sheets(wb, by_code, sector_by_co)

        # 2nd pass: median per product_code → benchmark fields
        by_pcode = self._group_prices_by_pcode(parsed)
        medians = {p: float(statistics.median(prices)) for p, prices in by_pcode.items()}
        benchmark_rows = self._apply_benchmark(parsed, medians)

        # 3rd pass: bulk insert (single tx)
        rows_for_insert = [self._row_for_insert(r) for r in parsed]
        async with self.uow:
            inserted = await self.uow.procurement.bulk_insert(rows_for_insert)

        return PaImportSummary(
            inserted=inserted,
            skipped_no_company=skipped_no_company,
            skipped_no_data=skipped_no_data,
            sheets_processed=sheets_processed,
            product_codes=len(by_pcode),
            benchmark_rows=benchmark_rows,
        )

    # ─── parse helpers ─────────────────────────────────────────────

    def _parse_sheets(self, wb, by_code, sector_by_co):
        parsed: list[dict[str, Any]] = []
        sheets_processed = 0
        skipped_no_company = 0
        skipped_no_data = 0

        for sheet_name in wb.sheetnames:
            code = _PA_SHEET_TO_CODE.get(sheet_name.upper().strip())
            if not code:
                code = sheet_name.lower().strip()
            company_id = by_code.get(code)
            if company_id is None:
                skipped_no_company += 1
                continue
            sheets_processed += 1

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue
            headers = [str(c or "").strip() for c in all_rows[0]]
            col_idx = {h: i for i, h in enumerate(headers)}

            def col(name: str) -> int:
                return col_idx.get(name, -1)

            cols = {
                "lot":     col("lotId"),
                "vendor":  col("vendor"),
                "vinn":    col("vendorInn"),
                "start":   col("startSumma"),
                "camt":    col("contractAmount"),
                "saved":   col("savedAmount"),
                "spct":    col("savedPercent"),
                "cdate":   col("contractDate"),
                "sdate":   col("startDate"),
                "ptype":   col("purchaseType"),
                "plat":    col("platformName"),
                "cat":     col("Category"),
                "pcode":   col("productCode"),
                "pname":   col("productName"),
                "ptype2":  col("productType"),
                "unit":    col("unit"),
                "amt":     col("amount"),
                "price":   col("Unit price"),
                "region":  col("regionName"),
            }

            for row in all_rows[1:]:
                if not row:
                    continue
                rec = self._parse_row(row, cols, company_id, code, sector_by_co)
                if rec is None:
                    skipped_no_data += 1
                    continue
                parsed.append(rec)
        return parsed, sheets_processed, skipped_no_company, skipped_no_data

    def _parse_row(self, row, cols, company_id, code, sector_by_co):
        def cell(i: int):
            return row[i] if 0 <= i < len(row) else None

        try:
            up_raw = cell(cols["price"])
            vol_raw = cell(cols["amt"])
            up = float(up_raw) if up_raw not in (None, "") else None
            vol = float(vol_raw) if vol_raw not in (None, "") else None
        except (TypeError, ValueError):
            return None
        # Салвадж цены за единицу: если «Unit price» не указана, но есть
        # количество и сумма контракта — считаем цену = сумма / количество
        # (одно-товарные e-shop/каталог лоты, напр. UMK). Иначе строка отбрасывается.
        if (up is None or up <= 0) and vol and vol > 0:
            tot = None
            for ci in ("camt", "start"):
                try:
                    raw = cell(cols[ci])
                    t = float(raw) if raw not in (None, "") else None
                except (TypeError, ValueError):
                    t = None
                if t and t > 0:
                    tot = t
                    break
            if tot:
                up = tot / vol
        if up is None or vol is None or up <= 0 or vol <= 0:
            return None

        def s(i: int):
            v = cell(i)
            return None if v in (None, "") else str(v).strip()

        def iso_date(i: int):
            v = cell(i)
            if not v:
                return None
            m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(v).strip())
            if not m:
                return None
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None

        def num(i: int):
            try:
                v = cell(i)
                return float(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None

        return {
            "company_id":    company_id,
            "code":          code,
            "year":          2026,
            "closure_date":  iso_date(cols["cdate"]) or iso_date(cols["sdate"]),
            "category_id":   _parse_category_id(s(cols["cat"])),
            "product_code":  s(cols["pcode"]),
            "product_name":  (s(cols["pname"]) or "")[:1024] or None,
            "unit_price":    up,
            "volume":        vol,
            "total_amount":  up * vol,
            "contract_amount": num(cols["camt"]),
            "start_summa":   num(cols["start"]),
            "saved_amount":  num(cols["saved"]),
            "saved_percent": num(cols["spct"]),
            "supplier_name": (s(cols["vendor"]) or "")[:512] or None,
            "supplier_inn":  s(cols["vinn"]),
            "lot_id":        s(cols["lot"]),
            "platform":      (s(cols["plat"]) or "")[:64] or None,
            "purchase_type": _norm_purchase_type(s(cols["ptype"])),
            "region":        (s(cols["region"]) or "")[:128] or None,
            "unit":          s(cols["unit"]),
            "product_type":  s(cols["ptype2"]),
            "sector":        sector_by_co.get(company_id),
        }

    def _group_prices_by_pcode(self, parsed) -> dict[str, list[float]]:
        out: dict[str, list[float]] = {}
        for r in parsed:
            if r["product_code"]:
                out.setdefault(r["product_code"], []).append(r["unit_price"])
        return out

    def _apply_benchmark(self, parsed, medians) -> int:
        benchmark_rows = 0
        for r in parsed:
            med = medians.get(r["product_code"]) if r["product_code"] else None
            if med is not None and med > 0:
                r["market_avg"] = med
                r["deviation_pct"] = (r["unit_price"] - med) / med * 100
                benchmark_rows += 1
            else:
                r["market_avg"] = None
                r["deviation_pct"] = None
        return benchmark_rows

    def _row_for_insert(self, r) -> dict[str, Any]:
        return {
            "company_id":    r["company_id"],
            "year":          r["year"],
            "closure_date":  r["closure_date"],
            "category_id":   r["category_id"],
            "product_code":  r["product_code"],
            "product_name":  r["product_name"],
            "unit_price":    r["unit_price"],
            "market_avg":    r["market_avg"],
            "deviation_pct": r["deviation_pct"],
            "unit":          r["unit"],
            "volume":        r["volume"],
            "total_amount":  r["total_amount"],
            "saved_amount":  r["saved_amount"],
            "supplier_name": r["supplier_name"],
            "supplier_inn":  r["supplier_inn"],
            "lot_id":        r["lot_id"],
            "platform":      r["platform"],
            "purchase_type": r["purchase_type"],
            "region":        r["region"],
            "sector":        r["sector"],
            "extra": json.dumps({
                "source": "manual-upload",
                "start_summa":     r["start_summa"],
                "contract_amount": r["contract_amount"],
                "saved_percent":   r["saved_percent"],
                "product_type":    r["product_type"],
            }, ensure_ascii=False),
        }
