"""Excel parser for detailed audited financial reports.

Two supported layouts:

Layout A — "Single section per sheet" (typical PwC/KPMG single-statement file):
  Each sheet contains ONE financial statement (SOFP, PnL, or CF). Header row
  has year columns; rows below are data lines. Section type inferred from
  user input (POST param).

Layout B — "High_Level_Financials" (UzAssets internal multi-statement file):
  Each sheet = one company. Sheet contains multiple sections (SOFP, PNL,
  Cash flow) stacked vertically, each with its own header row. Section
  marker text ('SOFP' / 'PNL' / 'P&L' / 'Cash flow') appears in a cell
  above each header row. Multiple currency blocks side-by-side
  (bln UZS / mln USD / x-rates) — we parse ONLY the bln UZS block.

The parser auto-detects layout B when it finds section markers on the sheet.
Otherwise falls back to layout A.

Both layouts produce ParsedSection objects: one per (sheet, statement type).
The API endpoint maps each ParsedSection to one or more FinancialReport
records (one per year).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

YEAR_MIN, YEAR_MAX = 2000, 2100

# Section markers found in cells: text → canonical report_type
SECTION_MARKERS: dict[str, str] = {
    "sofp": "BS",
    "balance sheet": "BS",
    "statement of financial position": "BS",
    "баланс": "BS",
    "pnl": "PL",
    "p&l": "PL",
    "p & l": "PL",
    "profit and loss": "PL",
    "profit & loss": "PL",
    "income statement": "PL",
    "отчет о прибылях": "PL",
    "cash flow": "CF",
    "cashflow": "CF",
    "statement of cash flow": "CF",
    "movement of cash": "CF",
    "одтс": "CF",
}

UNIT_MARKERS_UZS = ["bln uzs", "млрд сум", "млрд. сум", "млрд сумм", "blnuzs"]


@dataclass
class ParsedRow:
    kind: str                                      # "section_header" | "line" | "subtotal"
    label: str
    code: str
    indent_level: int = 0
    parent_code: Optional[str] = None
    section_label: Optional[str] = None
    is_subtotal: bool = False
    values: dict[int, float] = field(default_factory=dict)
    # Canonical mapping (filled by parse step):
    canonical_code: Optional[str] = None           # If matched, the canonical_code
    is_unmapped: bool = False                      # True if could not be mapped to canonical


@dataclass
class ParsedSection:
    """One financial statement parsed from one sheet."""
    sheet_name: str
    report_type: str                               # BS | PL | CF
    years: list[int]
    rows: list[ParsedRow]
    warnings: list[str] = field(default_factory=list)
    company_hint: Optional[str] = None             # sheet name lowercased


@dataclass
class ParsedSheet:
    """All sections found on one sheet."""
    sheet_name: str
    sections: list[ParsedSection]
    warnings: list[str] = field(default_factory=list)


# ── number / year parsing ───────────────────────────────────────────────


def _to_year(cell) -> Optional[int]:
    if cell is None:
        return None
    if isinstance(cell, int | float) and float(cell).is_integer():
        v = int(cell)
        if YEAR_MIN <= v <= YEAR_MAX:
            return v
    s = str(cell).strip()
    m = re.search(r"\b(20\d{2}|19\d{2})\b", s)
    if m:
        v = int(m.group(1))
        if YEAR_MIN <= v <= YEAR_MAX:
            return v
    return None


def _to_number(cell) -> Optional[float]:
    if cell is None or cell == "":
        return None
    if isinstance(cell, int | float):
        return float(cell)
    s = str(cell).strip()
    if not s or s in ("—", "-", "–", "n/a", "na", "nil", "N/A"):
        return None
    is_neg = False
    if s.startswith("(") and s.endswith(")"):
        is_neg = True
        s = s[1:-1]
    s = s.replace("\u00a0", " ").replace(" ", "").replace(",", "")
    try:
        v = float(s)
        return -v if is_neg else v
    except ValueError:
        return None


def _slugify_code(text: str, used: set[str]) -> str:
    s = re.sub(r"[^a-zA-Zа-яА-ЯёЁ0-9 ]", "", text or "").strip().lower()
    s = re.sub(r"\s+", "_", s)[:30]
    s = s or "line"
    base = s
    n = 1
    while s in used:
        n += 1
        s = f"{base[:28]}_{n}"
    used.add(s)
    return s


# ── label / row classification ──────────────────────────────────────────


SECTION_HEADER_PATTERNS = [
    "АКТИВЫ", "ПАССИВЫ", "СОБСТВЕННЫЙ КАПИТАЛ", "ОБЯЗАТЕЛЬСТВА", "КАПИТАЛ",
    "ВНЕОБОРОТНЫЕ", "ОБОРОТНЫЕ", "ДОЛГОСРОЧНЫЕ", "КРАТКОСРОЧНЫЕ",
    "ДОХОДЫ", "РАСХОДЫ",
    "ASSETS", "LIABILITIES", "EQUITY", "REVENUES", "EXPENSES",
    "NON-CURRENT", "CURRENT",
    "OPERATING ACTIVITIES", "INVESTING ACTIVITIES", "FINANCING ACTIVITIES",
    "ADJUSTMENTS",
]

SUBTOTAL_KEYWORDS = ["TOTAL", "ИТОГО", "ВСЕГО", "GROSS PROFIT", "OPERATING PROFIT",
                     "PROFIT FOR", "PROFIT BEFORE", "OPERATING CASH FLOW",
                     "INVESTING CASH FLOW", "FINANCING CASH FLOW",
                     "NET INCREASE", "NET DECREASE",
                     "CASH AND CASH EQUIVALENTS AT", "CCE AT THE",
                     "TOTAL COMPREHENSIVE"]


def _is_section_header(label: str) -> bool:
    if not label:
        return False
    s = label.strip()
    upper_ratio = sum(1 for c in s if c.isalpha() and c.isupper()) / max(
        sum(1 for c in s if c.isalpha()), 1
    )
    if upper_ratio > 0.7 and len(s) >= 3:
        return True
    return any(p in s.upper() for p in SECTION_HEADER_PATTERNS) and not any(
        ch.isdigit() for ch in s
    )


def _is_subtotal(label: str) -> bool:
    up = label.upper()
    return any(k in up for k in SUBTOTAL_KEYWORDS)


# ── section detection ───────────────────────────────────────────────────


def _detect_sections(rows: list) -> list[tuple[int, str]]:
    """Find rows that contain a section marker (SOFP/PNL/CF text in any cell).

    Returns list of (row_index, report_type), in order of appearance.
    """
    out = []
    for i, row in enumerate(rows):
        if not row:
            continue
        for cell in row:
            if not isinstance(cell, str):
                continue
            cs = cell.strip().lower()
            if cs in SECTION_MARKERS:
                out.append((i, SECTION_MARKERS[cs]))
                break
    return out


def _find_header_for_marker(
    rows: list,
    marker_row: int,
    prev_header_idx: int = -1,
) -> Optional[tuple[int, list[tuple[int, int]]]]:
    """Find the year header associated with a section marker.

    Try in order:
      1. Look BELOW marker (next 10 rows) — most common layout (NGMK, NUR…)
      2. If nothing below, reuse the previous section's header — TST layout
         where one shared header serves all sections (single header in R4,
         PL/SOFP/CF markers below).
    """
    # Try below first
    below = _find_header_after(rows, marker_row + 1, max_lookahead=10)
    if below:
        return below

    # Fall back: reuse previous section's header (shared header layout)
    if prev_header_idx >= 0:
        row = rows[prev_header_idx]
        if row:
            yrs = [(j, _to_year(c)) for j, c in enumerate(row) if _to_year(c) is not None]
            if len(yrs) >= 2:
                return prev_header_idx, yrs

    # Last resort: walk backwards from marker_row to row 0 looking for any header
    for i in range(marker_row - 1, -1, -1):
        row = rows[i]
        if not row:
            continue
        yrs = [(j, _to_year(c)) for j, c in enumerate(row) if _to_year(c) is not None]
        if len(yrs) >= 2:
            return i, yrs

    return None


def _find_header_after(rows: list, start_idx: int, max_lookahead: int = 10) -> Optional[tuple[int, list[tuple[int, int]]]]:
    """Find the first row at start_idx..start_idx+max_lookahead with ≥ 2 year cells."""
    for i in range(start_idx, min(start_idx + max_lookahead, len(rows))):
        row = rows[i]
        if not row:
            continue
        yrs = [(j, _to_year(c)) for j, c in enumerate(row) if _to_year(c) is not None]
        if len(yrs) >= 2:
            return i, yrs
    return None


def _find_uzs_columns(rows: list, header_idx: int, year_cols: list[tuple[int, int]]) -> dict[int, int]:
    """Identify which year columns belong to the bln UZS block.

    Strategy: look for "bln UZS" markers in rows above the header. UZS block
    is the columns where unit cell text matches. Falls back to first 4 year
    columns if no unit marker found.
    """
    uzs_cols: set[int] = set()
    for r in range(max(0, header_idx - 4), header_idx):
        row = rows[r]
        if not row:
            continue
        for j, c in enumerate(row):
            if not isinstance(c, str):
                continue
            cs = c.strip().lower()
            if any(m in cs for m in UNIT_MARKERS_UZS):
                uzs_cols.add(j)

    if not uzs_cols:
        # No unit markers — assume first contiguous block of year columns is UZS.
        # Detect contiguity: take year_cols sorted by col, stop at first gap > 1.
        sorted_yc = sorted(year_cols, key=lambda p: p[0])
        result: dict[int, int] = {}
        prev_col = None
        for col, year in sorted_yc:
            if prev_col is not None and col - prev_col > 1:
                break
            if year not in result:
                result[year] = col
            prev_col = col
        return result

    # Filter year_cols to those whose column is in uzs_cols
    out: dict[int, int] = {}
    for col, year in year_cols:
        if col in uzs_cols and year not in out:
            out[year] = col
    if not out:
        # Fallback: first 4
        return {y: c for c, y in year_cols[:4]}
    return out


def _find_label_column(rows: list, start_idx: int, end_idx: int, max_year_col: int) -> int:
    """Pick the column that holds row labels. Scans data rows, counting text hits per col."""
    text_counts: dict[int, int] = {}
    for i in range(start_idx, min(end_idx, len(rows))):
        row = rows[i]
        if not row:
            continue
        for j, c in enumerate(row):
            if j >= max_year_col:
                break
            if isinstance(c, str) and c.strip() and not _to_year(c):
                text_counts[j] = text_counts.get(j, 0) + 1
    if not text_counts:
        return 1
    # Prefer rightmost text column < max_year_col (typically B is English label, A is Uzbek).
    # We want English (B), so take the column with highest count, breaking ties toward higher col.
    best = max(text_counts.items(), key=lambda kv: (kv[1], kv[0]))
    return best[0]


def _row_label(row: list, label_col: int) -> Optional[tuple[str, int]]:
    """Get label text + indent level. Falls back to col 0 if label_col is empty."""
    label_cell = row[label_col] if label_col < len(row) else None
    label = (str(label_cell).strip() if label_cell else "")
    raw = str(label_cell) if label_cell else ""

    if not label and label_col > 0:
        c0 = row[0] if 0 < len(row) else None
        label = (str(c0).strip() if c0 else "")
        raw = str(c0) if c0 else ""

    if not label:
        return None

    leading_ws = len(raw) - len(raw.lstrip())
    indent = min(leading_ws // 2, 3)
    return label, indent


# ── parse one section ───────────────────────────────────────────────────


def _parse_section(
    sheet_name: str,
    rows: list,
    header_idx: int,
    year_cols: list[tuple[int, int]],
    end_idx: int,
    report_type: str,
    data_start_idx: Optional[int] = None,
) -> ParsedSection:
    uzs_cols = _find_uzs_columns(rows, header_idx, year_cols)
    if not uzs_cols:
        return ParsedSection(sheet_name=sheet_name, report_type=report_type,
                             years=[], rows=[],
                             warnings=["No UZS columns found"])

    years = sorted(uzs_cols.keys())
    max_year_col = max(uzs_cols.values()) + 1
    start_idx = data_start_idx if data_start_idx is not None else (header_idx + 1)

    label_col = _find_label_column(rows, start_idx, end_idx, max_year_col)

    parsed = ParsedSection(sheet_name=sheet_name, report_type=report_type,
                           years=years, rows=[])

    section_stack: list[str] = []
    used_codes: set[str] = set()

    for i in range(start_idx, min(end_idx, len(rows))):
        row = rows[i]
        if not row:
            continue

        label_info = _row_label(row, label_col)
        if not label_info:
            continue
        label, indent_from_ws = label_info

        # Read values from UZS columns only
        values: dict[int, float] = {}
        for year, col in uzs_cols.items():
            if col < len(row):
                v = _to_number(row[col])
                if v is not None:
                    values[year] = v

        if not values:
            if _is_section_header(label):
                section_stack = [label]
            elif section_stack and indent_from_ws == 0 and label.endswith(":"):
                section_stack = [section_stack[0], label.rstrip(":")]
            continue

        kind = "subtotal" if _is_subtotal(label) else "line"
        code = _slugify_code(label, used_codes)
        current_section = " · ".join(section_stack) if section_stack else None

        # Try to map to canonical
        from app.services.financial_canonical import match_canonical
        canonical = match_canonical(report_type, label, current_section)
        canonical_code = canonical.code if canonical else None

        parsed.rows.append(ParsedRow(
            kind=kind, label=label, code=code,
            indent_level=indent_from_ws,
            parent_code=None,
            section_label=current_section,
            is_subtotal=(kind == "subtotal"),
            values=values,
            canonical_code=canonical_code,
            is_unmapped=(canonical_code is None),
        ))

    if not parsed.rows:
        parsed.warnings.append("No data rows after header")

    return parsed


# ── UTY-style fallback: split a sectionless sheet by row content ────────


# Keyword → report_type. Used when a sheet has NO section markers (e.g. UTY),
# to bucket each row into BS / PL / CF based on its label content.
_PL_KEYWORDS = [
    "revenue", "выручка",
    "cos", "cogs", "cost of sales", "себестоимость",
    "gross profit", "валовая прибыль",
    "operating profit", "ebit", "операционная прибыль",
    "selling expenses", "g&a", "g a", "gna",
    "fin income", "fin cost", "finance income", "finance cost",
    "forex", "foreign exchange",
    "pbt", "profit before tax", "profit before income tax",
    "income tax", "tax expense",
    "net income", "net profit", "net loss", "соф фойда", "чистая прибыль",
    "comprehensive income",
]

_CF_KEYWORDS = [
    "depreciation",
    "dividends paid", "dividends declared", "deemed dividends",
    "cash flow", "cash flows",
    "operating cash flow", "investing cash flow", "financing cash flow",
    "operating cf", "investing cf", "financing cf",
    "cce", "cash and cash equivalents at",
    "capex", "purchases of ppe", "ppe additions",
    "proceeds from", "repayment of",
    "interest paid", "interest received",
    "changes in working capital",
]


def _classify_row_by_content(label: str) -> Optional[str]:
    """Return BS/PL/CF based on label keywords, or None if undetermined."""
    if not label:
        return None
    s = label.lower()
    for kw in _PL_KEYWORDS:
        if kw in s:
            return "PL"
    for kw in _CF_KEYWORDS:
        if kw in s:
            return "CF"
    return None


def _split_sectionless_section(sec: ParsedSection) -> list[ParsedSection]:
    """If a parsed section came from a marker-less sheet (UTY) and contains
    rows from multiple statement types, split it into separate sections.

    Logic: classify each row by content keywords. If we find ≥3 PL-classified
    rows and ≥3 BS-classified rows in the same section, split.
    """
    classified: list[tuple[ParsedRow, str]] = []
    for row in sec.rows:
        rtyp = _classify_row_by_content(row.label) or "BS"  # default unmatched → BS
        classified.append((row, rtyp))

    counts = {"BS": 0, "PL": 0, "CF": 0}
    for _, t in classified:
        counts[t] += 1

    # Only split if there's a meaningful mix
    has_pl = counts["PL"] >= 3
    has_cf = counts["CF"] >= 3
    has_bs = counts["BS"] >= 3
    if not (has_pl and has_bs) and not (has_cf and has_bs) and not (has_pl and has_cf):
        return [sec]

    # Build separate sections by report_type, preserving order
    out: dict[str, ParsedSection] = {}
    for row, rtyp in classified:
        if rtyp not in out:
            out[rtyp] = ParsedSection(
                sheet_name=sec.sheet_name,
                report_type=rtyp,
                years=list(sec.years),
                rows=[],
                warnings=[*list(sec.warnings), "Auto-split from sectionless sheet by row content"],
                company_hint=sec.company_hint,
            )
        out[rtyp].rows.append(row)

    # Re-run canonical mapping with the new (correct) report_type
    from app.services.financial_canonical import match_canonical
    for rtyp, new_sec in out.items():
        for row in new_sec.rows:
            canon = match_canonical(rtyp, row.label, row.section_label)
            row.canonical_code = canon.code if canon else None
            row.is_unmapped = (row.canonical_code is None)

    return list(out.values())


# ── public API ──────────────────────────────────────────────────────────


_AUX_SHEETS = {"status of ifrs reports", "company metrics", "mapping lib",
               "company names", "x-rates", "sheet1"}


def parse_workbook(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
    company_codes: Optional[set[str]] = None,
) -> list[ParsedSheet]:
    """Parse an Excel workbook.

    Parameters:
      file_bytes: raw bytes
      sheet_name: if given, only that sheet is parsed
      company_codes: if given, sheets whose name doesn't match
        (case-insensitive) any code in this set are skipped.

    Returns one ParsedSheet per processed sheet.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return []
        sheet_iter = [wb[sheet_name]]
    else:
        sheet_iter = []
        for s in wb.worksheets:
            name_lc = s.title.lower().strip()
            if name_lc in _AUX_SHEETS or s.title.startswith("_"):
                continue
            if company_codes is not None and name_lc not in company_codes:
                continue
            sheet_iter.append(s)

    out: list[ParsedSheet] = []

    for ws in sheet_iter:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        sheet_result = ParsedSheet(sheet_name=ws.title, sections=[])
        markers = _detect_sections(rows)

        if markers:
            # Multi-section sheet
            prev_header_idx = -1
            for idx, (marker_row, report_type) in enumerate(markers):
                hdr = _find_header_for_marker(rows, marker_row, prev_header_idx)
                if not hdr:
                    sheet_result.warnings.append(
                        f"Marker {report_type} at row {marker_row + 1} but no year header found")
                    continue
                header_idx, year_cols = hdr
                prev_header_idx = header_idx
                if idx + 1 < len(markers):
                    end_idx = markers[idx + 1][0]
                else:
                    end_idx = len(rows)

                # If the header is BEFORE the marker (e.g. TST), skip the marker
                # itself when collecting data rows by starting at max(header+1, marker+1).
                data_start = max(header_idx + 1, marker_row + 1)

                section = _parse_section(
                    sheet_name=ws.title, rows=rows,
                    header_idx=header_idx, year_cols=year_cols,
                    end_idx=end_idx, report_type=report_type,
                    data_start_idx=data_start,
                )
                section.company_hint = ws.title.lower()
                if section.rows:
                    sheet_result.sections.append(section)
        else:
            # Single section
            hdr = _find_header_after(rows, 0, max_lookahead=20)
            if hdr:
                header_idx, year_cols = hdr
                section = _parse_section(
                    sheet_name=ws.title, rows=rows,
                    header_idx=header_idx, year_cols=year_cols,
                    end_idx=len(rows), report_type="BS",
                )
                section.company_hint = ws.title.lower()
                if section.rows:
                    # Sectionless sheet (e.g. UTY) — try to auto-split into BS/PL/CF
                    # by classifying each row's content
                    split_sections = _split_sectionless_section(section)
                    if len(split_sections) > 1:
                        sheet_result.warnings.append(
                            f"No section markers — auto-split into {len(split_sections)} sections by content")
                    sheet_result.sections.extend(split_sections)

        if sheet_result.sections or sheet_result.warnings:
            out.append(sheet_result)

    return out
