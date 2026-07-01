"""Use cases for Forensic procurement snapshot."""
from __future__ import annotations

import io
import logging
from typing import Any, Optional

from fastapi import HTTPException
from fastapi import status as http_status

from app.uow.ports import UnitOfWorkABC

log = logging.getLogger(__name__)


SECTOR_COLOR = {
    "mining":    "#9B8EC4",
    "oilgas":    "#1D9E75",
    "energy":    "#EF9F27",
    "transport": "#378ADD",
    "other":     "#888780",
}

SECTOR_ORDER = {"mining": 0, "oilgas": 1, "energy": 2, "transport": 3, "other": 4}


def _ensure_year_row(co: dict, year: int) -> dict:
    if not isinstance(co.get("years"), list):
        co["years"] = []
    for yr in co["years"]:
        if yr.get("y") == year:
            return yr
    new_yr: dict = {"y": year}
    co["years"].append(new_yr)
    return new_yr


def _is_number(v: Any) -> bool:
    if v is None or v == "":
        return False
    try:
        float(str(v).replace(" ", "").replace(",", "."))
        return True
    except (TypeError, ValueError):
        return False


def _has_plan_number(co: dict) -> bool:
    """Есть ли РЕАЛЬНОЕ число плана (H-1/H-2): в years[].plan, legacy yP*, или в
    самом поле plan (7 флагманов хранят число в статус-поле). Честный признак
    «план заведён», в отличие от строкового статуса без единой суммы."""
    if _is_number(co.get("plan")):
        return True
    for yr in (co.get("years") or []):
        if isinstance(yr, dict) and _is_number(yr.get("plan")):
            return True
    for f in ("yP24", "yP25", "yP26"):
        if _is_number(co.get(f)):
            return True
    return False


def _forensic_really_done(co: dict) -> bool:
    """Форензик реально завершён (H-3): статус «Завершён» И указан аудитор И годы —
    формальный флаг без аудитора/лет не считаем проведённым аудитом."""
    return (
        co.get("forensic") == "Завершён"
        and bool((co.get("auditor") or "").strip())
        and bool((co.get("aYears") or "").strip())
    )


class ForensicService:
    def __init__(self, uow: UnitOfWorkABC) -> None:
        self.uow = uow

    # ─── overview ─────────────────────────────────────────────────

    async def overview(
        self,
        *,
        allowed_codes: Optional[set[str]],
    ) -> dict[str, Any]:
        async with self.uow:
            snap = await self.uow.forensic.load_snapshot()
            # Каноничные имена из таблицы Company (переопределяют запечённые в
            # снапшоте, чтобы forensic показывал те же названия, что /admin/companies).
            name_map = await self.uow.forensic.names_by_code()

        if not snap:
            return {
                "companies": [],
                "kpis": {
                    "total_companies": 0,
                    "plan_approved": 0,
                    "forensic_done": 0,
                    "with_auditor": 0,
                },
            }

        companies: list[dict[str, Any]] = []
        plan_approved = forensic_done = with_auditor = 0
        for raw in snap:
            if not isinstance(raw, dict):
                continue
            if allowed_codes is not None:
                row_code = (raw.get("k") or "").strip().lower()
                if not row_code or row_code not in allowed_codes:
                    continue
            sector = (raw.get("s") or "other").lower()
            enriched = dict(raw)
            # Каноничное имя из Company (если код есть в БД) — иначе имя из снапшота.
            code = (raw.get("k") or "").strip().lower()
            if code and code in name_map:
                enriched["n"] = name_map[code]
            enriched["sector_color"] = SECTOR_COLOR.get(sector, SECTOR_COLOR["other"])

            # H-1/H-2: «план утверждён» = есть реальное число плана (не строковый
            # статус без суммы; и ловит 7 флагманов с числом в поле plan).
            if _has_plan_number(raw):
                plan_approved += 1
            # H-3: форензик «завершён» только с аудитором и годами.
            if _forensic_really_done(raw):
                forensic_done += 1
            if (raw.get("auditor") or "").strip():
                with_auditor += 1
            companies.append(enriched)

        companies.sort(key=lambda c: (
            SECTOR_ORDER.get((c.get("s") or "other").lower(), 99),
            c.get("n") or "",
        ))

        return {
            "companies": companies,
            "kpis": {
                "total_companies": len(companies),
                "plan_approved": plan_approved,
                "forensic_done": forensic_done,
                "with_auditor": with_auditor,
            },
        }

    # ─── update company ───────────────────────────────────────────

    async def update_company(
        self,
        code: str,
        payload,
    ) -> dict[str, Any]:
        async with self.uow:
            snap = await self.uow.forensic.load_snapshot()
            if not snap:
                raise HTTPException(
                    http_status.HTTP_404_NOT_FOUND,
                    "Procurement snapshot not initialised",
                )
            co = next(
                (c for c in snap
                 if isinstance(c, dict) and (c.get("k") or "").lower() == code.lower()),
                None,
            )
            if co is None:
                raise HTTPException(
                    http_status.HTTP_404_NOT_FOUND,
                    f"Company '{code}' not found in snapshot",
                )

            if payload.plan_status is not None:
                co["plan"] = payload.plan_status
            if payload.forensic_status is not None:
                co["forensic"] = payload.forensic_status
            if payload.auditor is not None:
                co["auditor"] = payload.auditor
            if payload.audit_years is not None:
                co["aYears"] = payload.audit_years

            if payload.year_fields:
                yr = _ensure_year_row(co, payload.year)
                patch_dict = payload.year_fields.model_dump(exclude_none=True)
                for k, v in patch_dict.items():
                    yr[k] = v

            await self.uow.forensic.save_snapshot(snap)
            return {"ok": True, "code": code, "year": payload.year, "company": co}

    async def get_company_label(self, code: str) -> str:
        """Lookup display name for moderation entity_label (route uses this)."""
        async with self.uow:
            snap = await self.uow.forensic.load_snapshot()
        for c in snap:
            if isinstance(c, dict) and (c.get("k") or "").lower() == code.lower():
                return c.get("n") or code
        return code

    # ─── clear year ───────────────────────────────────────────────

    async def clear_year(self, year: Optional[int]) -> dict[str, Any]:
        async with self.uow:
            snap = await self.uow.forensic.load_snapshot()
            if not snap:
                return {"ok": True, "cleared": 0, "note": "snapshot empty"}
            cleared = 0
            for co in snap:
                if not isinstance(co, dict):
                    continue
                if year is None:
                    if isinstance(co.get("years"), list) and co["years"]:
                        cleared += len(co["years"])
                        co["years"] = []
                    for f in ("yP24", "yF24", "nP24", "nF24",
                              "yP25", "yF25", "nP25", "nF25",
                              "yP26"):
                        if f in co:
                            del co[f]
                else:
                    if isinstance(co.get("years"), list):
                        before = len(co["years"])
                        co["years"] = [yr for yr in co["years"] if yr.get("y") != year]
                        cleared += before - len(co["years"])
                    if year == 2024:
                        for f in ("yP24", "yF24", "nP24", "nF24"):
                            if f in co: del co[f]
                    elif year == 2025:
                        for f in ("yP25", "yF25", "nP25", "nF25"):
                            if f in co: del co[f]
                    elif year == 2026:
                        for f in ("yP26",):
                            if f in co: del co[f]
            await self.uow.forensic.save_snapshot(snap)
            return {"ok": True, "cleared": cleared, "year": year}

    # ─── Excel import ─────────────────────────────────────────────

    async def import_excel(self, raw_bytes: bytes) -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "openpyxl not installed")

        if len(raw_bytes) > 25 * 1024 * 1024:   # L-19: лимит размера загружаемого xlsx
            raise HTTPException(413, "Файл слишком большой (макс. 25 МБ).")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse xlsx: {e}")

        sheet_name = next(
            (n for n in wb.sheetnames if "данные" in n.lower()),
            wb.sheetnames[0] if wb.sheetnames else None,
        )
        if not sheet_name:
            raise HTTPException(400, "Workbook has no sheets")
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {
                "ok": True, "inserted": 0, "updated": 0, "skipped": 0,
                "note": "no data rows",
            }

        headers = [str(c or "").strip() for c in rows[0]]
        if not any("код" in h.lower() for h in headers):
            raise HTTPException(400, "Missing 'Код компании' column")

        def col(*needles: str) -> int:
            for i, h in enumerate(headers):
                low = h.lower()
                if all(n in low for n in needles):
                    return i
            return -1

        col_code = col("код")
        col_year = col("год")
        col_plan_y = col("план", "год")
        col_fact_y = col("факт", "год")
        col_plan_9 = col("план", "9")
        col_fact_9 = col("факт", "9")
        quarter_cols = {
            "q1p": col("q1", "план"), "q1f": col("q1", "факт"),
            "q2p": col("q2", "план"), "q2f": col("q2", "факт"),
            "q3p": col("q3", "план"), "q3f": col("q3", "факт"),
            "q4p": col("q4", "план"), "q4f": col("q4", "факт"),
        }
        col_plan_st = col("статус", "плана")
        col_for_st  = col("статус", "форензик")
        col_aud     = col("аудитор")
        col_period  = col("период")

        def _num(v: Any) -> Optional[float]:
            if v is None or v == "":
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        async with self.uow:
            snap = await self.uow.forensic.load_snapshot()
            by_code = {(c.get("k") or "").lower(): c for c in snap if isinstance(c, dict)}

            inserted = updated = skipped = 0

            for raw_row in rows[1:]:
                if not raw_row or all(c in (None, "") for c in raw_row):
                    continue

                def cell(i: int) -> Any:
                    return raw_row[i] if 0 <= i < len(raw_row) else None

                code_v = cell(col_code)
                if code_v is None or str(code_v).strip() == "":
                    skipped += 1
                    continue
                code_lc = str(code_v).strip().lower()
                co = by_code.get(code_lc)
                if co is None:
                    skipped += 1
                    continue

                year_v = cell(col_year)
                try:
                    year = int(year_v) if year_v not in (None, "") else None
                except (TypeError, ValueError):
                    skipped += 1
                    continue
                if not year:
                    skipped += 1
                    continue

                yr = _ensure_year_row(co, year)
                had_change = False
                for key, src in [
                    ("plan", col_plan_y), ("fact", col_fact_y),
                    ("n9p", col_plan_9), ("n9f", col_fact_9),
                ]:
                    v = _num(cell(src))
                    if v is not None and yr.get(key) != v:
                        yr[key] = v
                        had_change = True
                for q_key, q_col in quarter_cols.items():
                    v = _num(cell(q_col))
                    if v is not None and yr.get(q_key) != v:
                        yr[q_key] = v
                        had_change = True

                for field, src in [
                    ("plan", col_plan_st), ("forensic", col_for_st),
                    ("auditor", col_aud), ("aYears", col_period),
                ]:
                    v = cell(src)
                    if v is not None and str(v).strip() != "":
                        new_val = str(v).strip()
                        if co.get(field) != new_val:
                            co[field] = new_val
                            had_change = True

                if had_change:
                    updated += 1

            await self.uow.forensic.save_snapshot(snap)

        return {
            "ok": True,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "rows_processed": len(rows) - 1,
            "sheet": sheet_name,
        }

    # ─── scope helper ─────────────────────────────────────────────

    async def resolve_codes_for_scope(self, scope_ids) -> set[str]:
        async with self.uow:
            return await self.uow.forensic.codes_for_company_ids(list(scope_ids))
