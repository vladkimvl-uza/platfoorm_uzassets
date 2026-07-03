"""IFRS editor use-cases (Pack 7.59/7.62/7.63/7.64).

6 endpoints encapsulated:
  GET    /companies/{code}/ifrs-editor                  schema + values per (period, scope)
  PUT    /companies/{code}/ifrs-editor                  save (idempotent per slice)
  GET    /companies/{code}/ifrs-editor/history          last N audit entries
  GET    /companies/{code}/ifrs-editor/template         download 4-sheet XLSX template
  POST   /companies/{code}/ifrs-editor/parse-excel      parse uploaded XLSX
  GET    /companies/{code}/ifrs-nsbu-diff               compare IFRS vs NSBU canonical metrics

Storage layout: financial_reports + financial_lines, with:
  - standard='IFRS', is_detailed=False, source='ifrs-editor'
  - quarter mapped from period (FY→None, Q1→1, H1→2, 9M→3)
  - is_consolidated flag distinguishes group vs parent
Per-scope customization stored in `company.extra.ifrs_editor_schema_{period}_{c|s}`.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi import status as http_status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.access import allowed_company_ids
from app.core.audit_chain import append_audit_entry
from app.core.security import has_effective_permission
from app.models.audit import AuditLog
from app.models.financial import FinancialLine, FinancialReport
from app.models.user import User
from app.repositories.financials_repository import FinancialsRepository
from app.services.financials_portfolio.service import _canon_metric

# Field sets per section. MUST match useIfrsSchema.ts in frontend.
_IFRS_PL_FIELDS = {
    "revenue", "govGrants", "cogs", "grossProfit", "opProfit", "depreciation",
    "finIncome", "finCost", "interestExp", "forex",
    "pbt", "tax", "profit", "ebitda",
}
_IFRS_OCI_FIELDS = {
    "oci_currency_translation", "oci_revaluation_ppe", "oci_actuarial",
    "oci_hedge_reserve", "oci_fvtoci", "total_comprehensive_income",
}
_IFRS_BS_FIELDS = {
    "ppe", "totalNCA", "cash", "totalCA", "totalAssets",
    "equity", "shareCapital", "retainedEarnings",
    "ltBorrowings", "stBorrowings", "totalLiabilities",
    "ltBankLoans", "ltOtherLoans", "stBankLoans", "stOtherLoans",
    "longTermDebt", "debt",
}
_IFRS_CF_FIELDS = {
    "cfo", "cfo_pbt", "cfo_depreciation", "cfo_working_capital",
    "cfo_interest_paid", "cfo_tax_paid",
    "cfi", "cfi_capex", "cfi_acquisitions",
    "cff", "cff_borrowings", "cff_repayments", "dividendsPaid",
    "netCashChange", "freeCashFlow",
}

_SECTION_TO_RTYPE = {"pnl": "PL", "oci": "OCI", "sofp": "BS", "cf": "CF"}

_IFRS_SHEET_LABELS = {
    "pnl":  "ОФР",
    "oci":  "ОПД",
    "sofp": "Баланс",
    "cf":   "ДДС",
}

# Field labels for XLSX template. (canonical, label, ifrs_code, section)
_IFRS_FIELD_LABELS: dict[str, tuple[str, str, str, str]] = {
    # P&L
    "revenue":            ("revenue",            "Revenue · Выручка",                                 "", "pnl"),
    "govGrants":          ("govGrants",          "Government transfers · Господдержка (трансферы)",  "", "pnl"),
    "cogs":               ("cogs",               "Cost of sales · Себестоимость",                    "", "pnl"),
    "grossProfit":        ("grossProfit",        "Gross profit · Валовая прибыль (авто)",            "", "pnl"),
    "opProfit":           ("opProfit",           "Operating profit · Операционная прибыль",          "", "pnl"),
    "depreciation":       ("depreciation",       "D&A · Амортизация",                                "", "pnl"),
    "finIncome":          ("finIncome",          "Finance income · Финансовые доходы",               "", "pnl"),
    "finCost":            ("finCost",            "Finance costs · Финансовые расходы",               "", "pnl"),
    "interestExp":        ("interestExp",        "  Interest expense · Процентные расходы",          "", "pnl"),
    "forex":              ("forex",              "Forex · Курсовая разница",                         "", "pnl"),
    "pbt":                ("pbt",                "Profit before tax · Прибыль до налога (авто)",     "", "pnl"),
    "tax":                ("tax",                "Income tax · Налог на прибыль",                    "", "pnl"),
    "profit":             ("profit",             "Net profit · Чистая прибыль (авто)",               "", "pnl"),
    "ebitda":             ("ebitda",             "EBITDA (авто)",                                    "", "pnl"),
    # OCI
    "oci_currency_translation":   ("oci_currency_translation",   "OCI · Currency translation · Курсовые разницы пересчёта", "", "oci"),
    "oci_revaluation_ppe":        ("oci_revaluation_ppe",        "OCI · PPE revaluation · Переоценка ОС",                   "", "oci"),
    "oci_actuarial":              ("oci_actuarial",              "OCI · Actuarial · Актуарные доходы/расходы",              "", "oci"),
    "oci_hedge_reserve":          ("oci_hedge_reserve",          "OCI · Hedge reserve · Резерв хеджирования",               "", "oci"),
    "oci_fvtoci":                 ("oci_fvtoci",                 "OCI · FVTOCI · ФП по справ. стоимости",                   "", "oci"),
    "total_comprehensive_income": ("total_comprehensive_income", "Total comprehensive income · Совокупный доход (авто)",    "", "oci"),
    # Balance Sheet
    "ppe":                ("ppe",                "PPE · Основные средства",                          "", "sofp"),
    "totalNCA":           ("totalNCA",           "Total NCA · Внеоборотные активы (итог)",           "", "sofp"),
    "cash":               ("cash",               "Cash · Денежные средства",                         "", "sofp"),
    "totalCA":            ("totalCA",            "Total CA · Оборотные активы (итог)",               "", "sofp"),
    "totalAssets":        ("totalAssets",        "TOTAL ASSETS · ИТОГО Активы (авто)",               "", "sofp"),
    "equity":             ("equity",             "Equity · Собственный капитал",                     "", "sofp"),
    "shareCapital":       ("shareCapital",       "Share capital · Уставный капитал",                 "", "sofp"),
    "retainedEarnings":   ("retainedEarnings",   "Retained earnings · Нераспределённая прибыль",     "", "sofp"),
    "ltBorrowings":       ("ltBorrowings",       "LT borrowings · Долгоср. займы",                   "", "sofp"),
    "stBorrowings":       ("stBorrowings",       "ST borrowings · Краткоср. займы",                  "", "sofp"),
    "totalLiabilities":   ("totalLiabilities",   "TOTAL LIABILITIES (авто)",                         "", "sofp"),
    "ltBankLoans":        ("ltBankLoans",        "  LT bank loans",                                  "", "sofp"),
    "ltOtherLoans":       ("ltOtherLoans",       "  LT other loans",                                 "", "sofp"),
    "stBankLoans":        ("stBankLoans",        "  ST bank loans",                                  "", "sofp"),
    "stOtherLoans":       ("stOtherLoans",       "  ST other loans",                                 "", "sofp"),
    "longTermDebt":       ("longTermDebt",       "Long-term debt (separately)",                      "", "sofp"),
    "debt":               ("debt",               "Total debt · Финансовый долг (авто)",              "", "sofp"),
    # Cash Flow Statement
    "cfo":                ("cfo",                "CFO · Поток от операционной деят. (авто)",         "", "cf"),
    "cfo_pbt":            ("cfo_pbt",            "  Profit before tax (adj)",                        "", "cf"),
    "cfo_depreciation":   ("cfo_depreciation",   "  Depreciation (adj)",                             "", "cf"),
    "cfo_working_capital":("cfo_working_capital","  Change in working capital",                      "", "cf"),
    "cfo_interest_paid":  ("cfo_interest_paid",  "  Interest paid",                                  "", "cf"),
    "cfo_tax_paid":       ("cfo_tax_paid",       "  Income tax paid",                                "", "cf"),
    "cfi":                ("cfi",                "CFI · Поток от инвест. деят. (авто)",              "", "cf"),
    "cfi_capex":          ("cfi_capex",          "  CapEx · Капитальные затраты",                    "", "cf"),
    "cfi_acquisitions":   ("cfi_acquisitions",   "  Acquisitions",                                   "", "cf"),
    "cff":                ("cff",                "CFF · Поток от финансовой деят. (авто)",           "", "cf"),
    "cff_borrowings":     ("cff_borrowings",     "  Proceeds from borrowings",                       "", "cf"),
    "cff_repayments":     ("cff_repayments",     "  Repayments of borrowings",                       "", "cf"),
    "dividendsPaid":      ("dividendsPaid",      "  Dividends paid",                                 "", "cf"),
    "netCashChange":      ("netCashChange",      "Net change in cash (авто)",                        "", "cf"),
    "freeCashFlow":       ("freeCashFlow",       "Free Cash Flow (FCF) (авто)",                      "", "cf"),
}


def _ifrs_report_type(field: str) -> Optional[str]:
    if field in _IFRS_PL_FIELDS:  return "PL"
    if field in _IFRS_OCI_FIELDS: return "OCI"
    if field in _IFRS_BS_FIELDS:  return "BS"
    if field in _IFRS_CF_FIELDS:  return "CF"
    return None


def _period_to_quarter(period: str) -> Optional[int]:
    """FY → None (annual), Q1 → 1, H1 → 2, 9M → 3."""
    m = {"FY": None, "Q1": 1, "H1": 2, "9M": 3}
    if period not in m:
        raise HTTPException(
            422, f"Invalid period '{period}', expected FY/Q1/H1/9M"
        )
    return m[period]


# ─── Payload schemas ──────────────────────────────────────────────

class IfrsCustomFieldDef(BaseModel):
    id: str
    label: str
    section: Optional[str] = None  # 'pnl' | 'oci' | 'sofp' | 'cf'
    autoFormula: Optional[str] = None
    isCustom: Optional[bool] = True
    canonical: Optional[str] = None


class IfrsEditorSavePayload(BaseModel):
    """Payload from frontend IfrsEditor.vue save action."""
    period: str = "FY"
    consolidated: bool = True
    currency: str = "UZS"
    values: dict[str, dict[str, Optional[float]]] = Field(default_factory=dict)
    customFields: list[IfrsCustomFieldDef] = Field(default_factory=list)
    renames: dict[str, str] = Field(default_factory=dict)
    formulaOverrides: dict[str, str] = Field(default_factory=dict)
    manualFlags: dict[str, dict[str, bool]] = Field(default_factory=dict)
    audit_meta: Optional[dict] = None
    notes: dict[str, str] = Field(default_factory=dict)


# ─── Service ──────────────────────────────────────────────────────

@dataclass
class FinancialsIfrsService:
    @staticmethod
    def _schema_key(period: str, consolidated: bool) -> str:
        return f"ifrs_editor_schema_{period}_{'c' if consolidated else 's'}"

    async def get_schema(
        self,
        code: str,
        db: AsyncSession,
        user: User,
        *,
        period: str,
        consolidated: bool,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        quarter = _period_to_quarter(period)
        extra = co.extra or {}
        schema_key = self._schema_key(period, consolidated)
        schema = extra.get(schema_key, {}) if extra else {}

        q = (
            select(FinancialReport, FinancialLine)
            .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
            .where(
                FinancialReport.company_id == co.id,
                FinancialReport.standard == "IFRS",
                FinancialReport.is_consolidated.is_(consolidated),
                FinancialReport.is_detailed.is_(False),
            )
        )
        if quarter is None:
            q = q.where(FinancialReport.quarter.is_(None))
        else:
            q = q.where(FinancialReport.quarter == quarter)

        rows = await db.execute(q)
        values: dict[str, dict[str, float]] = {}
        audit_meta_latest: Optional[dict] = None
        audit_year_latest = 0
        for fr, fl in rows.all():
            if fl.value is None:
                continue
            try:
                v = float(fl.value)
            except (TypeError, ValueError):
                continue
            values.setdefault(fl.line_code, {})[str(fr.year)] = v
            if fr.extra and "audit" in fr.extra and fr.year > audit_year_latest:
                audit_meta_latest = fr.extra.get("audit")
                audit_year_latest = fr.year

        return {
            "code": co.code,
            "period": period,
            "consolidated": consolidated,
            "currency": "UZS",
            "values": values,
            "customFields": schema.get("customFields", []),
            "renames": schema.get("renames", {}),
            "formulaOverrides": schema.get("formulaOverrides", {}),
            "manualFlags": schema.get("manualFlags", {}),
            "notes": schema.get("notes", {}),
            "audit_meta": audit_meta_latest,
            "updatedAt": schema.get("updatedAt"),
            "updatedBy": schema.get("updatedBy"),
        }

    async def save(
        self,
        code: str,
        payload: IfrsEditorSavePayload,
        db: AsyncSession,
        user: User,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.edit"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN,
                "Permission required: financials.edit",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        quarter = _period_to_quarter(payload.period)
        now_iso = datetime.now(UTC).isoformat()

        # 1. Persist customization (per-scope slot)
        extra = dict(co.extra or {})
        schema_key = self._schema_key(payload.period, payload.consolidated)
        extra[schema_key] = {
            "customFields": [cf.model_dump() for cf in payload.customFields],
            "renames": payload.renames,
            "formulaOverrides": payload.formulaOverrides,
            "manualFlags": payload.manualFlags,
            "notes": {
                k: v for k, v in payload.notes.items() if v and v.strip()
            },
            "updatedAt": now_iso,
            "updatedBy": user.email,
        }
        co.extra = extra

        # 2. Build canonical map + bucket by report type
        label_for_field: dict[str, str] = {}
        custom_section_by_id: dict[str, str] = {}
        canonical_for_field: dict[str, Optional[str]] = {}

        for f in (
            _IFRS_PL_FIELDS | _IFRS_OCI_FIELDS
            | _IFRS_BS_FIELDS | _IFRS_CF_FIELDS
        ):
            canonical_for_field[f] = f
        for cf in payload.customFields:
            label_for_field[cf.id] = cf.label
            if cf.section:
                custom_section_by_id[cf.id] = cf.section
            canonical_for_field[cf.id] = cf.canonical or None
        for fld, renamed in payload.renames.items():
            label_for_field[fld] = renamed

        changes_by_report: dict[
            tuple[int, str],
            list[tuple[str, Optional[float], str, Optional[str]]],
        ] = {}

        for field, year_map in payload.values.items():
            rtype = _ifrs_report_type(field)
            if rtype is None and field in custom_section_by_id:
                rtype = _SECTION_TO_RTYPE.get(custom_section_by_id[field])
            if not rtype:
                continue
            for year_str, val in year_map.items():
                try:
                    year = int(year_str)
                except (TypeError, ValueError):
                    continue
                changes_by_report.setdefault((year, rtype), []).append((
                    field, val,
                    label_for_field.get(field, field),
                    canonical_for_field.get(field),
                ))

        # 3. Upsert reports + lines
        reports_created = 0
        reports_updated = 0
        lines_upserted = 0
        lines_deleted = 0
        audit_target_year = max(
            (y for (y, _) in changes_by_report.keys()), default=None,
        )

        for (year, report_type), changes in changes_by_report.items():
            base_filter = [
                FinancialReport.company_id == co.id,
                FinancialReport.year == year,
                FinancialReport.standard == "IFRS",
                FinancialReport.report_type == report_type,
                FinancialReport.is_consolidated.is_(payload.consolidated),
                FinancialReport.is_detailed.is_(False),
            ]
            if quarter is None:
                base_filter.append(FinancialReport.quarter.is_(None))
            else:
                base_filter.append(FinancialReport.quarter == quarter)
            rep_q = await db.execute(
                select(FinancialReport).where(*base_filter)
            )
            existing_reports = list(rep_q.scalars().all())
            report = None
            for r in existing_reports:
                if r.source == "ifrs-editor":
                    report = r
                    break
            if report is None and existing_reports:
                report = existing_reports[0]

            if report is None:
                report = FinancialReport(
                    company_id=co.id,
                    year=year,
                    quarter=quarter,
                    standard="IFRS",
                    report_type=report_type,
                    currency="UZS",
                    unit_scale=1_000_000_000,
                    source="ifrs-editor",
                    is_audited=bool(
                        payload.audit_meta
                        and payload.audit_meta.get("opinion") == "clean"
                    ),
                    is_detailed=False,
                    is_consolidated=payload.consolidated,
                    notes=(
                        f"Saved via IFRS editor by {user.email} on {now_iso}"
                    ),
                    extra={"editor_version": "p7.59"},
                )
                db.add(report)
                await db.flush()
                reports_created += 1
            else:
                reports_updated += 1
                report.source = report.source or "ifrs-editor"
                report.notes = (
                    f"Last edit via IFRS editor by {user.email} on {now_iso}"
                )

            if payload.audit_meta is not None and year == audit_target_year:
                rep_extra = dict(report.extra or {})
                rep_extra["audit"] = payload.audit_meta
                report.extra = rep_extra

            ln_q = await db.execute(
                select(FinancialLine).where(
                    FinancialLine.report_id == report.id
                )
            )
            existing_lines = {
                ln.line_code: ln for ln in ln_q.scalars().all()
            }

            for field, val, label, canonical in changes:
                existing = existing_lines.get(field)
                if val is None:
                    if existing is not None:
                        await db.delete(existing)
                        lines_deleted += 1
                    continue
                decimal_val = Decimal(str(val))
                new_parent = canonical
                if existing is not None:
                    existing.value = decimal_val
                    existing.line_name = label
                    existing.parent_code = new_parent
                else:
                    db.add(FinancialLine(
                        report_id=report.id,
                        line_code=field,
                        parent_code=new_parent,
                        line_name=label,
                        value=decimal_val,
                        is_subtotal=False,
                        is_calculated=False,
                        sort_order=0,
                    ))
                lines_upserted += 1

        # 4. Audit log entry
        try:
            sample_fields = sorted(set(
                f for changes in changes_by_report.values()
                for f, _, _, _ in changes
            ))[:20]
            await append_audit_entry(
                db,
                actor_id=str(user.id) if user.id else None,
                actor_email=user.email,
                action="ifrs_editor.save",
                entity_type="company",
                entity_id=str(co.id),
                diff={
                    "period": payload.period,
                    "consolidated": payload.consolidated,
                    "reports_created": reports_created,
                    "reports_updated": reports_updated,
                    "lines_upserted": lines_upserted,
                    "lines_deleted": lines_deleted,
                    "fields": sample_fields,
                    "years": sorted({
                        y for (y, _) in changes_by_report.keys()
                    }),
                },
                payload={
                    "company_code": co.code,
                    "customFields_count": len(payload.customFields),
                    "renames_count": len(payload.renames),
                    "audit_meta_set": payload.audit_meta is not None,
                },
                notes=(
                    f"IFRS editor save · {co.code} · {payload.period} · "
                    f"{'consolidated' if payload.consolidated else 'standalone'}"
                ),
            )
        except Exception:
            pass

        await db.commit()
        return {
            "ok": True,
            "saved_at": now_iso,
            "period": payload.period,
            "consolidated": payload.consolidated,
            "reports_created": reports_created,
            "reports_updated": reports_updated,
            "lines_upserted": lines_upserted,
            "lines_deleted": lines_deleted,
        }

    async def get_history(
        self, code: str, db: AsyncSession, user: User, *, limit: int,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        q = await db.execute(
            select(AuditLog)
            .where(
                AuditLog.action == "ifrs_editor.save",
                AuditLog.entity_type == "company",
                AuditLog.entity_id == str(co.id),
            )
            .order_by(AuditLog.created_at.desc())
            .limit(limit)
        )
        entries = list(q.scalars().all())
        return {
            "code": co.code,
            "company_name": co.name_short or co.name_ru,
            "total": len(entries),
            "entries": [
                {
                    "id": str(e.id),
                    "at": e.created_at.isoformat() if e.created_at else None,
                    "actor_email": e.actor_email,
                    "diff": e.diff or {},
                    "payload": e.payload or {},
                    "notes": e.notes,
                }
                for e in entries
            ],
        }

    async def ifrs_nsbu_diff(
        self,
        code: str,
        db: AsyncSession,
        user: User,
        *,
        year: int,
        consolidated: bool,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        rows = await db.execute(
            select(
                FinancialReport.standard, FinancialLine.line_code,
                FinancialLine.parent_code, FinancialLine.value,
                FinancialLine.line_name,
            )
            .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
            .where(
                FinancialReport.company_id == co.id,
                FinancialReport.year == year,
                FinancialReport.quarter.is_(None),
                FinancialReport.is_consolidated.is_(consolidated),
                FinancialReport.is_detailed.is_(False),
                FinancialReport.standard.in_(("IFRS", "NSBU")),
            )
        )

        by_std: dict[str, dict[str, tuple[float, str]]] = {
            "IFRS": {}, "NSBU": {}
        }
        for std, line_code, parent_code, value, line_name in rows.all():
            if value is None:
                continue
            canon = _canon_metric(line_code, parent_code)
            if not canon:
                continue
            try:
                v = float(value)
            except (TypeError, ValueError):
                continue
            existing = by_std[std].get(canon)
            if existing is None or abs(v) > abs(existing[0]):
                by_std[std][canon] = (v, line_name or canon)

        all_metrics = set(by_std["NSBU"].keys()) | set(by_std["IFRS"].keys())
        diffs: list[dict] = []
        for metric in sorted(all_metrics):
            nsbu_tuple = by_std["NSBU"].get(metric)
            ifrs_tuple = by_std["IFRS"].get(metric)
            nsbu_val = nsbu_tuple[0] if nsbu_tuple else None
            ifrs_val = ifrs_tuple[0] if ifrs_tuple else None
            label = (
                (ifrs_tuple or nsbu_tuple)[1]
                if (ifrs_tuple or nsbu_tuple) else metric
            )
            if nsbu_val is not None and ifrs_val is not None:
                delta = ifrs_val - nsbu_val
                denom = abs(nsbu_val) if nsbu_val != 0 else 1
                delta_pct = (delta / denom) * 100
                abs_pct = abs(delta_pct)
                if abs_pct < 5:    sig = "low"
                elif abs_pct < 20: sig = "medium"
                else:              sig = "high"
            elif nsbu_val is None and ifrs_val is not None:
                delta = ifrs_val
                delta_pct = None
                sig = "ifrs_only"
            elif ifrs_val is None and nsbu_val is not None:
                delta = -nsbu_val
                delta_pct = None
                sig = "nsbu_only"
            else:
                continue
            diffs.append({
                "metric": metric, "label": label,
                "nsbu_value": nsbu_val, "ifrs_value": ifrs_val,
                "delta": delta, "delta_pct": delta_pct,
                "significance": sig,
            })

        sig_rank = {
            "high": 0, "medium": 1, "ifrs_only": 2,
            "nsbu_only": 3, "low": 4,
        }
        diffs.sort(key=lambda d: (sig_rank.get(d["significance"], 9), d["metric"]))

        return {
            "code": co.code,
            "company_name": co.name_short or co.name_ru,
            "year": year,
            "consolidated": consolidated,
            "metrics_total": len(diffs),
            "summary": {
                "high": sum(1 for d in diffs if d["significance"] == "high"),
                "medium": sum(1 for d in diffs if d["significance"] == "medium"),
                "low": sum(1 for d in diffs if d["significance"] == "low"),
                "ifrs_only": sum(1 for d in diffs if d["significance"] == "ifrs_only"),
                "nsbu_only": sum(1 for d in diffs if d["significance"] == "nsbu_only"),
            },
            "diffs": diffs,
        }

    async def download_template(
        self,
        code: str,
        db: AsyncSession,
        user: User,
        *,
        years: str,
        period: str,
        consolidated: bool,
    ) -> StreamingResponse:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        try:
            year_list = sorted({
                int(y.strip()) for y in years.split(",") if y.strip()
            })
        except ValueError:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                "Invalid years parameter",
            )
        if not year_list:
            year_list = [2021, 2022, 2023, 2024, 2025, 2026]

        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        header_font = Font(bold=True, size=11, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF7F77DD")
        auto_fill = PatternFill("solid", fgColor="FFFFFBF0")
        auto_font = Font(italic=True, color="FFD97706")
        border = Border(
            left=Side(style="thin", color="FFE2E8F0"),
            right=Side(style="thin", color="FFE2E8F0"),
            top=Side(style="thin", color="FFE2E8F0"),
            bottom=Side(style="thin", color="FFE2E8F0"),
        )
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        scope_label = "consolidated" if consolidated else "standalone"

        def fill_sheet(ws, section_id: str):
            section_name = _IFRS_SHEET_LABELS.get(section_id, section_id)
            ws.cell(
                row=1, column=1,
                value=(
                    f"МСФО · {section_name} · {co.code} "
                    f"{co.name_short or co.name_ru or ''} · "
                    f"{period} · {scope_label}"
                ),
            )
            ws.cell(row=1, column=1).font = Font(
                bold=True, size=13, color="FF1E2A4A"
            )
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=2 + len(year_list),
            )
            helper = (
                "Заполняйте числовые поля. Поля, помеченные «авто», "
                "пересчитываются автоматически. Числа в МЛРД UZS."
            )
            ws.cell(row=2, column=1, value=helper).font = Font(
                italic=True, size=9, color="FF94A3B8"
            )
            ws.merge_cells(
                start_row=2, start_column=1,
                end_row=2, end_column=2 + len(year_list),
            )
            ws.cell(row=4, column=1, value="Код")
            ws.cell(row=4, column=2, value="Показатель")
            for i, yr in enumerate(year_list):
                ws.cell(row=4, column=3 + i, value=yr)
            for col in range(1, 3 + len(year_list)):
                c = ws.cell(row=4, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border
            row = 5
            for field_id, (_fid, label, _code, sect) in _IFRS_FIELD_LABELS.items():
                if sect != section_id:
                    continue
                is_auto = "(авто)" in label
                ws.cell(row=row, column=1, value=field_id)
                ws.cell(row=row, column=2, value=label)
                for col in range(1, 3 + len(year_list)):
                    c = ws.cell(row=row, column=col)
                    c.border = border
                    if col >= 3:
                        c.alignment = center
                    elif col == 2:
                        c.alignment = left
                    else:
                        c.alignment = center
                    if is_auto:
                        c.fill = auto_fill
                        if col == 2:
                            c.font = auto_font
                row += 1
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 55
            for i in range(len(year_list)):
                ws.column_dimensions[chr(ord("C") + i)].width = 14
            ws.freeze_panes = "C5"

        for section_id in ("pnl", "oci", "sofp", "cf"):
            ws = wb.create_sheet(_IFRS_SHEET_LABELS[section_id])
            fill_sheet(ws, section_id)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"ifrs_template_{co.code}_{period}_{scope_label}.xlsx"
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{fname}"',
            },
        )

    async def parse_excel(
        self,
        code: str,
        file: UploadFile,
        db: AsyncSession,
        user: User,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        contents = await file.read()
        if not contents:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST, "Empty file",
            )
        from openpyxl import load_workbook
        try:
            wb = load_workbook(
                io.BytesIO(contents), data_only=True, read_only=True,
            )
        except Exception as e:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                f"Cannot parse XLSX: {e}",
            )

        values: dict[str, dict[int, float]] = {}
        parse_log: list[str] = []

        label_to_field: dict[str, str] = {}
        for _fid, (canonical, label, _code, _sect) in _IFRS_FIELD_LABELS.items():
            clean = label.replace("(авто)", "").strip().lower()
            label_to_field[clean] = canonical
            label_to_field[canonical.lower()] = canonical
            # Support split parts of "English · Русский" labels
            for part in label.replace("(авто)", "").split("·"):
                p = part.strip().lower()
                if p and p not in label_to_field:
                    label_to_field[p] = canonical

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            header_row = None
            year_cols: dict[int, int] = {}
            for r in range(1, min(16, ws.max_row + 1)):
                row_cells = list(ws.iter_rows(
                    min_row=r, max_row=r, values_only=True,
                ))
                if not row_cells:
                    continue
                row_vals = row_cells[0]
                yc_local: dict[int, int] = {}
                for ci, val in enumerate(row_vals):
                    if isinstance(val, int | float) and 1990 < int(val) < 2100:
                        yc_local[ci] = int(val)
                    elif (
                        isinstance(val, str)
                        and val.strip().isdigit()
                        and 1990 < int(val.strip()) < 2100
                    ):
                        yc_local[ci] = int(val.strip())
                if yc_local:
                    header_row = r
                    year_cols = yc_local
                    break
            if not header_row or not year_cols:
                parse_log.append(
                    f"⚠ Лист «{sheet}»: не найдена строка с годами"
                )
                continue
            parse_log.append(
                f"Лист «{sheet}»: заголовок в строке {header_row}, "
                f"годы {sorted(year_cols.values())}"
            )
            rows_parsed = 0
            for r in range(header_row + 1, ws.max_row + 1):
                row_cells = list(ws.iter_rows(
                    min_row=r, max_row=r, values_only=True,
                ))
                if not row_cells:
                    continue
                row_vals = row_cells[0]
                field_id: Optional[str] = None
                for ci in range(min(3, len(row_vals))):
                    v = row_vals[ci]
                    if not isinstance(v, str):
                        continue
                    key = (
                        v.strip().lower().replace("(авто)", "").strip()
                    )
                    if key in label_to_field:
                        field_id = label_to_field[key]
                        break
                if not field_id:
                    continue
                for col_idx, year in year_cols.items():
                    if col_idx >= len(row_vals):
                        continue
                    cell_val = row_vals[col_idx]
                    if cell_val is None or cell_val == "":
                        continue
                    try:
                        num = (
                            float(cell_val) if not isinstance(cell_val, str)
                            else float(
                                cell_val.replace(",", ".").replace(" ", "")
                            )
                        )
                    except (TypeError, ValueError):
                        continue
                    if not (-1e12 < num < 1e12):
                        continue
                    values.setdefault(field_id, {})[year] = num
                rows_parsed += 1
            parse_log.append(f"  → распознано строк: {rows_parsed}")

        out_values = {
            fld: {str(y): v for y, v in ym.items()}
            for fld, ym in values.items()
        }
        return {
            "company_code": co.code,
            "company_name": co.name_short or co.name_ru,
            "values": out_values,
            "fields_count": len(out_values),
            "cells_count": sum(len(ym) for ym in out_values.values()),
            "log": parse_log,
            "filename": file.filename,
        }
