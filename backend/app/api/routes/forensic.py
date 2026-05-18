"""Forensic & Procurement audit endpoints.

Endpoint:
  GET /api/forensic/overview      → list of 22 companies with plan/fact/forensic data

Source: system_config['firebase_dump.procurementData'] (Phase 4 already snapshotted
this 10 KB JSONB from /pf/procurementData). Each row has the monolith shape:
  {
    n: 'НГМК', k: 'ngmk', s: 'mining',
    yP24, yF24, nP24, nF24,           # 2024 plan/fact (annual + 9m)
    yP25, yF25, nP25, nF25,           # 2025
    yP26,                              # 2026 plan only
    plan: 'Утверждён' | 'Не утверждён' | ...,
    forensic: 'Завершён' | 'В процессе' | 'Тендер в 2026' | 'Не начат',
    auditor: 'KPMG' | 'PwC' | 'Deloitte' | 'E&Y' | '',
    aYears: '2024' | '2022-2024' | '',
    years?: [{y, plan, fact, n9p, n9f, q1p, q1f, ...}]   # newer schema
  }

The endpoint passes data through with minimal transformation — just enriches
each row with `sector_color` (from companies table sector → color) so the
frontend can colour-tag without an extra lookup.
"""
from __future__ import annotations

import io
import json
import logging
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status as http_status
from pydantic import BaseModel, Field
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_user, get_db
from app.core.access import allowed_company_ids, has_unrestricted_view
from app.core.security import _has_permission, has_effective_permission
from app.models.company import Company, Sector
from app.models.user import User


log = logging.getLogger(__name__)
router = APIRouter(prefix="/forensic", tags=["forensic"])


# Sector → colour mapping (mirrors monolith _pSC palette)
_SECTOR_COLOR = {
    "mining":    "#9B8EC4",
    "oilgas":    "#1D9E75",
    "energy":    "#EF9F27",
    "transport": "#378ADD",
    "other":     "#888780",
}


@router.get("/overview")
async def forensic_overview(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Returns the full list of 22 companies with plan/fact/forensic data.

    Response shape:
      {
        "companies": [
          {n, k, s, sector_color,
           yP24, yF24, nP24, nF24, yP25, yF25, nP25, nF25, yP26,
           plan, forensic, auditor, aYears, years?},
          ...
        ],
        "kpis": {
          "total_companies": 22,
          "plan_approved": <count>,
          "forensic_done": <count>,
          "with_auditor": <count>
        }
      }
    """
    if not await has_effective_permission(db, user, "procurement.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.view required")

    # 1. Read snapshot from system_config (Phase 4 puts it here)
    res = await db.execute(text(
        "SELECT value FROM system_config "
        "WHERE key = 'firebase_dump.procurementData' LIMIT 1"
    ))
    row = res.first()
    if not row or not row[0]:
        return {
            "companies": [],
            "kpis": {
                "total_companies": 0,
                "plan_approved": 0,
                "forensic_done": 0,
                "with_auditor": 0,
            },
        }

    snap = row[0]
    if isinstance(snap, str):
        try:
            snap = json.loads(snap)
        except json.JSONDecodeError:
            log.warning("forensic/overview: failed to parse JSONB snapshot")
            return {"companies": [], "kpis": {}}

    if not isinstance(snap, list):
        return {"companies": [], "kpis": {}}

    # Scope filter: snapshot хранит company.code в поле `k`, не UUID.
    # Резолвим allowed_company_ids → codes → фильтруем snapshot.
    allowed_codes: Optional[set[str]] = None
    if not has_unrestricted_view(user):
        scope_ids = await allowed_company_ids(db, user)
        if not scope_ids:
            return {
                "companies": [],
                "kpis": {
                    "total_companies": 0,
                    "plan_approved": 0,
                    "forensic_done": 0,
                    "with_auditor": 0,
                },
            }
        code_q = await db.execute(
            select(Company.code).where(Company.id.in_(scope_ids))
        )
        allowed_codes = {c.lower() for (c,) in code_q.all() if c}

    # 2. Enrich each row with sector_color from sector code
    companies: list[dict[str, Any]] = []
    plan_approved = forensic_done = with_auditor = 0
    for raw in snap:
        if not isinstance(raw, dict):
            continue
        # Scope-фильтр для не-admin юзеров. Пропускаем записи компаний,
        # которых нет в allowed_companies (snapshot хранит code в поле `k`).
        if allowed_codes is not None:
            row_code = (raw.get("k") or "").strip().lower()
            if not row_code or row_code not in allowed_codes:
                continue
        sector = (raw.get("s") or "other").lower()
        enriched = dict(raw)
        enriched["sector_color"] = _SECTOR_COLOR.get(sector, _SECTOR_COLOR["other"])

        # KPIs
        if (raw.get("plan") or "").startswith("Утверждён"):
            plan_approved += 1
        if raw.get("forensic") == "Завершён":
            forensic_done += 1
        if (raw.get("auditor") or "").strip():
            with_auditor += 1

        companies.append(enriched)

    # Sort by sector order then name (mirrors monolith _pSC palette order)
    sector_order = {"mining": 0, "oilgas": 1, "energy": 2, "transport": 3, "other": 4}
    companies.sort(key=lambda c: (
        sector_order.get((c.get("s") or "other").lower(), 99),
        c.get("n") or "",
    ))

    return {
        "companies": companies,
        "kpis": {
            "total_companies": len(companies),
            "plan_approved": plan_approved,
            "forensic_done": forensic_done,
            "with_auditor": with_auditor,
        },
    }


# =====================================================================
# Mutation helpers — read/write firebase_dump.procurementData JSONB snapshot
# =====================================================================

_SNAPSHOT_KEY = "firebase_dump.procurementData"
_YEAR_FIELDS = {"plan", "fact", "n9p", "n9f", "q1p", "q1f", "q2p", "q2f", "q3p", "q3f", "q4p", "q4f"}
_META_FIELDS = {"plan", "forensic", "auditor", "aYears"}  # at company-level


async def _load_snapshot(db: AsyncSession) -> list[dict]:
    res = await db.execute(text("SELECT value FROM system_config WHERE key = :k LIMIT 1"), {"k": _SNAPSHOT_KEY})
    row = res.first()
    if not row or not row[0]:
        return []
    snap = row[0]
    if isinstance(snap, str):
        try:
            snap = json.loads(snap)
        except json.JSONDecodeError:
            return []
    return snap if isinstance(snap, list) else []


async def _save_snapshot(db: AsyncSession, snap: list[dict]) -> None:
    payload = json.dumps(snap, ensure_ascii=False)
    # Upsert into system_config — overwrites existing snapshot
    await db.execute(text("""
        INSERT INTO system_config (id, key, value, description, is_secret, created_at, updated_at)
        VALUES (gen_random_uuid(), :k, CAST(:v AS jsonb), :d, FALSE, NOW(), NOW())
        ON CONFLICT (key) DO UPDATE
        SET value = EXCLUDED.value, updated_at = NOW()
    """), {"k": _SNAPSHOT_KEY, "v": payload, "d": "Procurement plan/fact data (PROCUREMENT_DATA) mutable via /forensic endpoints"})
    await db.commit()


def _ensure_year_row(co: dict, year: int) -> dict:
    """Ensure co['years'] exists and contains an entry for `year`. Returns the entry."""
    if not isinstance(co.get("years"), list):
        co["years"] = []
    for yr in co["years"]:
        if yr.get("y") == year:
            return yr
    new_yr: dict = {"y": year}
    co["years"].append(new_yr)
    return new_yr


# =====================================================================
# PUT /forensic/companies/{code} — update single company × year
# =====================================================================

class YearPatch(BaseModel):
    """Editable per-(company, year) fields."""
    plan: Optional[float] = None
    fact: Optional[float] = None
    n9p:  Optional[float] = None
    n9f:  Optional[float] = None
    q1p:  Optional[float] = None
    q1f:  Optional[float] = None
    q2p:  Optional[float] = None
    q2f:  Optional[float] = None
    q3p:  Optional[float] = None
    q3f:  Optional[float] = None
    q4p:  Optional[float] = None
    q4f:  Optional[float] = None


class CompanyPatch(BaseModel):
    """Per-(company, year) plan/fact + metadata edit."""
    year: int = Field(..., ge=2000, le=2100)
    year_fields: Optional[YearPatch] = None
    plan_status:      Optional[str] = None   # "Утверждён" | "Не утверждён" | "" | None (skip)
    forensic_status:  Optional[str] = None
    auditor:          Optional[str] = None
    audit_years:      Optional[str] = None   # field 'aYears'


@router.put("/companies/{code}")
async def update_forensic_company(
    code: str,
    payload: CompanyPatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Patch plan/fact + meta for one company × year. Writes back to JSONB snapshot."""
    if not await has_effective_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")

    # Scope-check (non-admin only)
    if not has_unrestricted_view(user):
        scope_ids = await allowed_company_ids(db, user)
        if not scope_ids:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No company access")
        code_q = await db.execute(select(Company.code).where(Company.id.in_(scope_ids)))
        allowed_codes = {c.lower() for (c,) in code_q.all() if c}
        if code.lower() not in allowed_codes:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    snap = await _load_snapshot(db)
    if not snap:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, "Procurement snapshot not initialised")

    co = next((c for c in snap if isinstance(c, dict) and (c.get("k") or "").lower() == code.lower()), None)
    if co is None:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found in snapshot")

    # ── Moderation gate ────────────────────────────────────────
    from fastapi.responses import JSONResponse
    from app.services.moderation_service import gate_or_apply
    queued, sub = await gate_or_apply(
        db, user=user,
        module="procurement", action="update_company",
        entity_id=code, entity_label=f"Закупки: {co.get('n', code)} · {payload.year}",
        company_id=None, sector_id=None, year=payload.year,
        payload=payload.model_dump(mode="json", exclude_none=True),
        diff_summary=f"Forensic update: {code} year {payload.year}",
    )
    if queued:
        return JSONResponse(
            status_code=http_status.HTTP_202_ACCEPTED,
            content={"queued": True, "submission_id": str(sub.id), "status": sub.status},
        )

    # Apply meta fields (skip when None — explicit empty-string clears)
    if payload.plan_status is not None:
        co["plan"] = payload.plan_status
    if payload.forensic_status is not None:
        co["forensic"] = payload.forensic_status
    if payload.auditor is not None:
        co["auditor"] = payload.auditor
    if payload.audit_years is not None:
        co["aYears"] = payload.audit_years

    # Apply year-level numeric fields
    if payload.year_fields:
        yr = _ensure_year_row(co, payload.year)
        patch_dict = payload.year_fields.model_dump(exclude_none=True)
        for k, v in patch_dict.items():
            yr[k] = v

    await _save_snapshot(db, snap)
    return {"ok": True, "code": code, "year": payload.year, "company": co}


# =====================================================================
# DELETE /forensic/data?year=N — clear year data
# =====================================================================

@router.delete("/data")
async def clear_forensic_year(
    year: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Clear year-specific data from all companies' years[] arrays.
    If `year` is omitted, clears all years[] for all companies (preserves company
    metadata — plan/forensic/auditor stay). For full snapshot wipe, drop the
    system_config row directly via admin tools.
    """
    if not await has_effective_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")
    if not has_unrestricted_view(user):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Admin scope required to clear data")

    snap = await _load_snapshot(db)
    if not snap:
        return {"ok": True, "cleared": 0, "note": "snapshot empty"}

    cleared = 0
    for co in snap:
        if not isinstance(co, dict):
            continue
        if year is None:
            if isinstance(co.get("years"), list) and co["years"]:
                cleared += len(co["years"])
                co["years"] = []
            # Also clear legacy fields
            for f in ("yP24", "yF24", "nP24", "nF24",
                      "yP25", "yF25", "nP25", "nF25",
                      "yP26"):
                if f in co:
                    del co[f]
        else:
            if isinstance(co.get("years"), list):
                before = len(co["years"])
                co["years"] = [yr for yr in co["years"] if yr.get("y") != year]
                cleared += before - len(co["years"])
            # Clear matching legacy fields for that year
            if year == 2024:
                for f in ("yP24", "yF24", "nP24", "nF24"):
                    if f in co: del co[f]
            elif year == 2025:
                for f in ("yP25", "yF25", "nP25", "nF25"):
                    if f in co: del co[f]
            elif year == 2026:
                for f in ("yP26",):
                    if f in co: del co[f]

    await _save_snapshot(db, snap)
    return {"ok": True, "cleared": cleared, "year": year}


# =====================================================================
# POST /forensic/import-excel — parse xlsx + upsert into snapshot
# =====================================================================

@router.post("/import-excel")
async def import_forensic_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Parse uploaded xlsx (Данные sheet, headers per downloadForensicTemplate)
    and upsert each row into the snapshot."""
    if not await has_effective_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")
    if not has_unrestricted_view(user):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Admin scope required for bulk import")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Expected .xlsx or .xls file")

    raw = await file.read()
    if not raw:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Empty file")

    # Parse via openpyxl
    try:
        import openpyxl  # noqa: WPS433
    except ImportError:
        raise HTTPException(http_status.HTTP_500_INTERNAL_SERVER_ERROR, "openpyxl not installed")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, f"Failed to parse xlsx: {e}")

    sheet_name = next((n for n in wb.sheetnames if "данные" in n.lower()), wb.sheetnames[0] if wb.sheetnames else None)
    if not sheet_name:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Workbook has no sheets")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"ok": True, "inserted": 0, "updated": 0, "skipped": 0, "note": "no data rows"}

    headers = [str(c or "").strip() for c in rows[0]]
    # Required headers
    if not any("код" in h.lower() for h in headers):
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Missing 'Код компании' column")

    # Column index map
    def col(*needles: str) -> int:
        for i, h in enumerate(headers):
            low = h.lower()
            if all(n in low for n in needles):
                return i
        return -1

    col_code = col("код")
    col_year = col("год")
    col_plan_y = col("план", "год")
    col_fact_y = col("факт", "год")
    col_plan_9 = col("план", "9")
    col_fact_9 = col("факт", "9")
    quarter_cols = {
        "q1p": col("q1", "план"), "q1f": col("q1", "факт"),
        "q2p": col("q2", "план"), "q2f": col("q2", "факт"),
        "q3p": col("q3", "план"), "q3f": col("q3", "факт"),
        "q4p": col("q4", "план"), "q4f": col("q4", "факт"),
    }
    col_plan_st = col("статус", "плана")
    col_for_st  = col("статус", "форензик")
    col_aud     = col("аудитор")
    col_period  = col("период")

    snap = await _load_snapshot(db)
    by_code = {(c.get("k") or "").lower(): c for c in snap if isinstance(c, dict)}

    inserted = updated = skipped = 0

    def _num(v: Any) -> Optional[float]:
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    for raw_row in rows[1:]:
        if not raw_row or all(c in (None, "") for c in raw_row):
            continue

        def cell(i: int) -> Any:
            return raw_row[i] if i >= 0 and i < len(raw_row) else None

        code_v = cell(col_code)
        if code_v is None or str(code_v).strip() == "":
            skipped += 1
            continue
        code_lc = str(code_v).strip().lower()
        co = by_code.get(code_lc)
        if co is None:
            skipped += 1
            continue

        year_v = cell(col_year)
        try:
            year = int(year_v) if year_v not in (None, "") else None
        except (TypeError, ValueError):
            skipped += 1
            continue
        if not year:
            skipped += 1
            continue

        # Numeric fields
        yr = _ensure_year_row(co, year)
        had_change = False
        # Find numeric columns and write
        for key, src in [
            ("plan", col_plan_y), ("fact", col_fact_y),
            ("n9p", col_plan_9), ("n9f", col_fact_9),
        ]:
            v = _num(cell(src))
            if v is not None:
                if yr.get(key) != v:
                    yr[key] = v
                    had_change = True
        for q_key, q_col in quarter_cols.items():
            v = _num(cell(q_col))
            if v is not None:
                if yr.get(q_key) != v:
                    yr[q_key] = v
                    had_change = True

        # Meta (only set if cell non-empty; empty preserves existing)
        for field, src in [
            ("plan", col_plan_st), ("forensic", col_for_st),
            ("auditor", col_aud), ("aYears", col_period),
        ]:
            v = cell(src)
            if v is not None and str(v).strip() != "":
                new_val = str(v).strip()
                if co.get(field) != new_val:
                    co[field] = new_val
                    had_change = True

        if had_change:
            updated += 1
        else:
            # Touched but nothing changed — count as no-op
            pass

    await _save_snapshot(db, snap)
    return {
        "ok": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "rows_processed": len(rows) - 1,
        "sheet": sheet_name,
    }
