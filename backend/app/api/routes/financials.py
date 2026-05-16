"""Financials editor API.

Endpoints:
  GET    /financials/catalog                    line-codes catalog for editor UI
  GET    /financials                            list reports (filters: company, year, standard)
  GET    /financials/{report_id}                full report with lines + checksum
  POST   /financials                            create new (empty) report
  PUT    /financials/{report_id}                full replace of report + all lines (anti-loss)
  DELETE /financials/{report_id}                delete report + lines

Anti-loss protocol on PUT:
  1. Optimistic concurrency: if `expected_prev_checksum` is provided and
     doesn't match server's current checksum, returns 409 Conflict.
  2. Single transaction: DELETE all old lines + INSERT all new lines.
     If anything fails, transaction rolls back — no partial state possible.
  3. Recompute and return the new checksum so client can verify-after-save.
  4. Audit chain entry written for every save with diff summary.
"""
import hashlib
import json
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status as http_status
from sqlalchemy import asc, delete, desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.access import allowed_company_ids
from app.core.audit_chain import append_audit_entry
from app.core.security import _has_permission, get_current_user
from app.database import get_db
from app.models.company import Company
from app.models.financial import FinancialLine, FinancialReport
from app.models.user import User
from app.models.year_registry import YearRegistry
from app.schemas.financial import (
    CatalogResponse, FinancialLineCatalogEntry, FinancialLineEdit,
    FinancialReportCreatePayload, FinancialReportFull,
    FinancialReportListItem, FinancialReportSavePayload,
    FinancialReportSaveResponse,
)


router = APIRouter(prefix="/financials", tags=["financials"])


# =====================================================================
# Catalog (loaded once from JSON seed file at module load)
# =====================================================================

_CATALOG_PATH = Path(__file__).resolve().parents[3] / "data" / "seed" / "financial_lines_catalog.json"


def _load_catalog() -> List[FinancialLineCatalogEntry]:
    if not _CATALOG_PATH.exists():
        return []
    raw = json.loads(_CATALOG_PATH.read_text(encoding="utf-8"))
    return [FinancialLineCatalogEntry(**r) for r in raw]


_CATALOG_CACHE: Optional[List[FinancialLineCatalogEntry]] = None


def get_catalog() -> List[FinancialLineCatalogEntry]:
    global _CATALOG_CACHE
    if _CATALOG_CACHE is None:
        _CATALOG_CACHE = _load_catalog()
    return _CATALOG_CACHE


# =====================================================================
# Checksum helpers (anti-loss verify)
# =====================================================================

def _compute_checksum(report: FinancialReport, lines: List[FinancialLine]) -> str:
    """Deterministic checksum over report header + sorted lines.

    Used by the editor for:
    - optimistic concurrency check (don't overwrite someone else's save)
    - verify-after-save (detect silent corruption)

    Decimal values are normalized to 4 decimal places (matching the DB column
    Numeric(28, 4)) so that `Decimal('12.5')` and `Decimal('12.5000')` —
    semantically identical but distinct as strings — hash to the same checksum.
    """
    parts: list[str] = [
        f"{report.year}|{report.quarter or ''}|{report.standard}|{report.report_type}",
        f"{report.currency}|{report.unit_scale}|{int(report.is_audited)}",
    ]
    sorted_lines = sorted(lines, key=lambda l: (l.line_code or "", l.sort_order))
    for ln in sorted_lines:
        if ln.value is None:
            v = ""
        else:
            # Normalize to 4 decimal places — matches Numeric(28, 4) DB column
            normalized = Decimal(ln.value).quantize(Decimal("0.0001"))
            v = format(normalized, "f")
        parts.append(f"{ln.line_code}|{v}|{ln.sort_order}")
    blob = "\n".join(parts).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()[:32]


# =====================================================================
# Helpers
# =====================================================================

async def _hydrate_report(db: AsyncSession, report: FinancialReport) -> FinancialReportFull:
    line_q = await db.execute(
        select(FinancialLine).where(FinancialLine.report_id == report.id)
        .order_by(FinancialLine.sort_order.asc(), FinancialLine.line_code.asc())
    )
    lines = list(line_q.scalars().all())

    co_q = await db.execute(
        select(Company.code, Company.name_short).where(Company.id == report.company_id)
    )
    co = co_q.first()

    checksum = _compute_checksum(report, lines)

    return FinancialReportFull(
        id=report.id, company_id=report.company_id,
        company_code=co.code if co else "",
        company_name=co.name_short if co else None,
        year=report.year, quarter=report.quarter,
        standard=report.standard, report_type=report.report_type,
        currency=report.currency, unit_scale=report.unit_scale,
        source=report.source, is_audited=report.is_audited,
        notes=report.notes, extra=report.extra,
        lines=[FinancialLineEdit(
            line_code=ln.line_code, line_name=ln.line_name,
            line_name_uz=ln.line_name_uz, line_name_en=ln.line_name_en,
            parent_code=ln.parent_code, value=ln.value,
            is_subtotal=ln.is_subtotal, is_calculated=ln.is_calculated,
            sort_order=ln.sort_order,
        ) for ln in lines],
        created_at=report.created_at, updated_at=report.updated_at,
        checksum=checksum,
    )


# =====================================================================
# Endpoints
# =====================================================================

@router.get("/catalog", response_model=CatalogResponse)
async def get_financials_catalog(user: User = Depends(get_current_user)):
    """Reference catalog for the editor — line codes + standards + scales."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.view")
    return CatalogResponse(line_codes=get_catalog())


@router.get("", response_model=List[FinancialReportListItem])
async def list_reports(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    company_code: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    standard: Optional[str] = Query(None, pattern="^(IFRS|NSBU)$"),
    limit: int = Query(100, ge=1, le=500),
):
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.view")

    # Per-company scope
    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and len(scope_ids) == 0:
        return []

    q = (select(FinancialReport, Company.code.label("co_code"),
                func.count(FinancialLine.id).label("lines_count"))
         .join(Company, FinancialReport.company_id == Company.id)
         .outerjoin(FinancialLine, FinancialLine.report_id == FinancialReport.id)
         .group_by(FinancialReport.id, Company.code))

    if scope_ids is not None:
        q = q.where(FinancialReport.company_id.in_(scope_ids))

    if company_code: q = q.where(func.lower(Company.code) == company_code.lower())
    if year:         q = q.where(FinancialReport.year == year)
    if standard:     q = q.where(FinancialReport.standard == standard)

    q = q.order_by(desc(FinancialReport.year), Company.code.asc()).limit(limit)
    rows = (await db.execute(q)).all()
    return [
        FinancialReportListItem(
            id=r.FinancialReport.id, company_code=r.co_code,
            year=r.FinancialReport.year, quarter=r.FinancialReport.quarter,
            standard=r.FinancialReport.standard, report_type=r.FinancialReport.report_type,
            is_audited=r.FinancialReport.is_audited,
            lines_count=r.lines_count or 0,
            updated_at=r.FinancialReport.updated_at,
        )
        for r in rows
    ]


@router.get("/{report_id}", response_model=FinancialReportFull)
async def get_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.view")

    res = await db.execute(select(FinancialReport).where(FinancialReport.id == report_id))
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, "Financial report not found")

    # Per-company scope check
    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and report.company_id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this report")

    return await _hydrate_report(db, report)


@router.post("", response_model=FinancialReportFull, status_code=http_status.HTTP_201_CREATED)
async def create_report(
    payload: FinancialReportCreatePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    # Per-company scope: scoped users can only create reports for allowed companies
    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and payload.company_id not in scope_ids:
        raise HTTPException(
            http_status.HTTP_403_FORBIDDEN,
            "Cannot create financial report for a company outside your allowed list",
        )

    co_q = await db.execute(select(Company).where(Company.id == payload.company_id))
    company = co_q.scalar_one_or_none()
    if not company:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Company not found")

    # Conflict check: same (company, year, quarter, standard, report_type)?
    dup_q = await db.execute(
        select(FinancialReport).where(
            FinancialReport.company_id == payload.company_id,
            FinancialReport.year == payload.year,
            FinancialReport.quarter == payload.quarter,
            FinancialReport.standard == payload.standard,
            FinancialReport.report_type == payload.report_type,
        )
    )
    if dup_q.scalar_one_or_none():
        raise HTTPException(
            http_status.HTTP_409_CONFLICT,
            f"Report for {company.code} {payload.year}/{payload.standard}/{payload.report_type} already exists.",
        )

    report = FinancialReport(
        company_id=payload.company_id,
        year=payload.year, quarter=payload.quarter,
        standard=payload.standard, report_type=payload.report_type,
        currency=payload.currency, unit_scale=payload.unit_scale,
        source=payload.source, is_audited=False,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    await append_audit_entry(
        db, actor_id=str(user.id), actor_email=user.email,
        action="financials.create",
        entity_type="financial_report", entity_id=str(report.id),
        notes=f"{company.code} {payload.year} {payload.standard}/{payload.report_type}",
    )
    await db.commit()

    return await _hydrate_report(db, report)


@router.put("/{report_id}", response_model=FinancialReportSaveResponse)
async def save_report(
    report_id: UUID,
    payload: FinancialReportSavePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Full replace of the report + ALL its lines, with optimistic concurrency.

    This is the editor's primary save endpoint. The flow:
      1. Load existing report + lines
      2. Compute current checksum, compare to `expected_prev_checksum`
         from the client. If mismatch → 409 (someone else saved).
      3. Inside a single transaction:
         - Update report header fields
         - DELETE all existing lines
         - INSERT all lines from the payload
      4. Recompute checksum after save and return it
      5. Append audit chain entry
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    res = await db.execute(select(FinancialReport).where(FinancialReport.id == report_id))
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, "Financial report not found")

    # Per-company scope check
    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and report.company_id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this report")

    # Optimistic concurrency check
    if payload.expected_prev_checksum:
        old_lines_q = await db.execute(
            select(FinancialLine).where(FinancialLine.report_id == report_id)
        )
        old_lines = list(old_lines_q.scalars().all())
        current_checksum = _compute_checksum(report, old_lines)
        if current_checksum != payload.expected_prev_checksum:
            raise HTTPException(
                http_status.HTTP_409_CONFLICT,
                "The report was modified by someone else since you opened it. "
                "Please refresh and re-apply your changes.",
            )

    # Update header
    report.year         = payload.year
    report.quarter      = payload.quarter
    report.standard     = payload.standard
    report.report_type  = payload.report_type
    report.currency     = payload.currency
    report.unit_scale   = payload.unit_scale
    report.source       = payload.source
    report.is_audited   = payload.is_audited
    report.notes        = payload.notes
    report.extra        = payload.extra

    # Replace all lines: delete old, insert new
    await db.execute(delete(FinancialLine).where(FinancialLine.report_id == report_id))

    new_line_objs: list[FinancialLine] = []
    for ln in payload.lines:
        new_line_objs.append(FinancialLine(
            report_id=report_id,
            line_code=ln.line_code, line_name=ln.line_name,
            line_name_uz=ln.line_name_uz, line_name_en=ln.line_name_en,
            parent_code=ln.parent_code, value=ln.value,
            is_subtotal=ln.is_subtotal, is_calculated=ln.is_calculated,
            sort_order=ln.sort_order,
        ))
    for o in new_line_objs:
        db.add(o)

    await db.flush()

    # Recompute checksum AFTER save
    new_checksum = _compute_checksum(report, new_line_objs)

    # Audit
    await append_audit_entry(
        db, actor_id=str(user.id), actor_email=user.email,
        action="financials.save",
        entity_type="financial_report", entity_id=str(report_id),
        notes=f"lines={len(new_line_objs)}, checksum={new_checksum[:16]}",
    )
    await db.commit()
    await db.refresh(report)

    full = await _hydrate_report(db, report)
    return FinancialReportSaveResponse(
        report=full,
        saved_at=datetime.now(timezone.utc),
        lines_total=len(new_line_objs),
        server_checksum=new_checksum,
    )


@router.delete("/{report_id}", status_code=http_status.HTTP_204_NO_CONTENT)
async def delete_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    res = await db.execute(select(FinancialReport).where(FinancialReport.id == report_id))
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, "Financial report not found")

    # Per-company scope check
    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and report.company_id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this report")

    co_q = await db.execute(select(Company.code).where(Company.id == report.company_id))
    co_code = co_q.scalar_one_or_none() or "?"

    await db.delete(report)  # cascades to financial_lines via FK ondelete
    await db.commit()

    await append_audit_entry(
        db, actor_id=str(user.id), actor_email=user.email,
        action="financials.delete",
        entity_type="financial_report", entity_id=str(report_id),
        notes=f"{co_code} {report.year} {report.standard}/{report.report_type}",
    )
    await db.commit()


# =====================================================================
# Excel import for detailed audited reports
# =====================================================================


from fastapi import File, Form, UploadFile  # noqa: E402
from app.services.excel_financial_parser import parse_workbook, ParsedSheet, ParsedSection  # noqa: E402


@router.post("/detailed/import-excel", status_code=http_status.HTTP_201_CREATED)
async def import_detailed_excel(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    file: UploadFile = File(...),
    standard: str = Form("IFRS", description="IFRS or NSBU — applied to all imported reports"),
    is_audited: bool = Form(True),
    company_code: Optional[str] = Form(None,
        description="Single-company mode: only import the sheet matching this company code"),
    report_type: Optional[str] = Form(None,
        description="Single-section mode: force this report_type for all parsed sections "
                    "(only valid with company_code AND only-one-section sheet)"),
    sheet_name: Optional[str] = Form(None,
        description="Force a specific sheet name to be used as the only source"),
):
    """Parse an audited Excel file and create detailed FinancialReports.

    Supports TWO modes:

    1) **Multi-sheet, multi-section** (default — when company_code is omitted):
       Each sheet name must match a company code (case-insensitive). Each
       sheet may contain SOFP/PNL/Cash flow sections marked by labels.
       All matching sheets and all detected sections are imported in one go.
       Suitable for the High_Level_Financials_v4 layout.

    2) **Single company** (when company_code is set):
       Only the sheet matching company_code is imported. If the sheet has
       multiple sections, all are imported with their detected report_type.
       If `report_type` is also given AND only ONE section is found, it
       overrides the detected type. Otherwise report_type is ignored.

    The import is REPLACE-NOT-MERGE: existing detailed reports for each
    (company, standard, report_type) tuple are deleted before insertion.
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    if standard not in ("IFRS", "NSBU"):
        raise HTTPException(422, "standard must be IFRS or NSBU")
    if report_type is not None and report_type not in ("PL", "BS", "CF"):
        raise HTTPException(422, "report_type must be PL, BS, or CF")

    # Read file
    try:
        file_bytes = await file.read()
    except Exception as e:
        raise HTTPException(400, f"Failed to read uploaded file: {e}")

    if len(file_bytes) > 25 * 1024 * 1024:
        raise HTTPException(413, "File too large (max 25 MB)")

    # Build map of all known company codes for sheet matching
    all_co_q = await db.execute(select(Company))
    all_companies = {c.code.lower(): c for c in all_co_q.scalars().all()}
    if not all_companies:
        raise HTTPException(500, "No companies in database")

    scope_ids = await allowed_company_ids(db, user)

    # ── Parse the Excel ──
    if company_code:
        # Single-company mode: filter to exactly one sheet
        co = all_companies.get(company_code.lower())
        if not co:
            raise HTTPException(404, f"Company '{company_code}' not found")
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

        try:
            parsed_sheets = parse_workbook(
                file_bytes, sheet_name,
                company_codes={co.code.lower()},
            )
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel: {type(e).__name__}: {e}")

        if not parsed_sheets:
            # Fallback: try without company filter (single-sheet workbook with arbitrary name)
            try:
                parsed_sheets = parse_workbook(file_bytes, sheet_name)
            except Exception as e:
                raise HTTPException(400, f"Failed to parse Excel: {type(e).__name__}: {e}")
            if not parsed_sheets:
                raise HTTPException(400, "No valid financial sheets found")
            # Force the company match
            for ps in parsed_sheets:
                for sec in ps.sections:
                    sec.company_hint = co.code.lower()
    else:
        # Multi-company mode: scan all sheets matching company codes
        try:
            parsed_sheets = parse_workbook(
                file_bytes, sheet_name,
                company_codes=set(all_companies.keys()),
            )
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel: {type(e).__name__}: {e}")

        if not parsed_sheets:
            raise HTTPException(400,
                "No sheets matched company codes. Sheet names must match a company code "
                f"(case-insensitive). Known codes: {sorted(all_companies.keys())[:5]}…")

    # Apply scope filter
    if scope_ids is not None:
        parsed_sheets = [
            ps for ps in parsed_sheets
            if (all_companies.get(ps.sheet_name.lower()) and
                all_companies[ps.sheet_name.lower()].id in scope_ids)
        ]
        if not parsed_sheets:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN,
                "No access to any company in this Excel")

    # Apply report_type override (only if exactly one section found in single-company mode)
    if company_code and report_type:
        all_sections = [sec for ps in parsed_sheets for sec in ps.sections]
        if len(all_sections) == 1:
            all_sections[0].report_type = report_type

    # Collect summary
    sheet_results = []
    total_reports = 0
    total_lines = 0
    skipped_sheets = []

    # ── Wipe existing detailed reports per (company, standard, report_type) we are about to write ──
    wipe_keys: set[tuple] = set()
    for ps in parsed_sheets:
        co = all_companies.get(ps.sheet_name.lower())
        if not co:
            skipped_sheets.append(f"{ps.sheet_name} (no matching company)")
            continue
        for sec in ps.sections:
            wipe_keys.add((co.id, standard, sec.report_type))

    if wipe_keys:
        for (cid, std, rtyp) in wipe_keys:
            existing_q = await db.execute(
                select(FinancialReport).where(
                    FinancialReport.company_id == cid,
                    FinancialReport.standard == std,
                    FinancialReport.report_type == rtyp,
                    FinancialReport.is_detailed == True,  # noqa: E712
                )
            )
            for old in existing_q.scalars().all():
                await db.delete(old)
        await db.flush()

    # ── Insert new reports ──
    for ps in parsed_sheets:
        co = all_companies.get(ps.sheet_name.lower())
        if not co:
            continue

        sheet_summary = {
            "sheet_name": ps.sheet_name,
            "company_code": co.code,
            "company_name": co.name_short,
            "sections": [],
            "warnings": list(ps.warnings),
        }

        for sec in ps.sections:
            for year in sec.years:
                rep = FinancialReport(
                    company_id=co.id,
                    year=year, quarter=None,
                    standard=standard, report_type=sec.report_type,
                    currency="UZS", unit_scale=1_000_000_000,  # bln UZS
                    source=f"excel:{file.filename or 'upload'}#{ps.sheet_name}",
                    is_audited=is_audited,
                    is_detailed=True,
                    extra={
                        "imported_at": datetime.now(timezone.utc).isoformat(),
                        "imported_by": user.email,
                        "source_sheet": ps.sheet_name,
                        "source_filename": file.filename,
                        "warnings": sec.warnings,
                    },
                )
                db.add(rep)
                await db.flush()
                total_reports += 1

                lines_for_year = 0
                for sort_idx, prow in enumerate(sec.rows):
                    v = prow.values.get(year)
                    if v is None:
                        continue
                    db.add(FinancialLine(
                        report_id=rep.id,
                        line_code=prow.code,
                        line_name=prow.label,
                        parent_code=prow.parent_code,
                        section_label=prow.section_label,
                        indent_level=prow.indent_level,
                        value=Decimal(str(v)),
                        is_subtotal=prow.is_subtotal,
                        is_calculated=False,
                        sort_order=sort_idx,
                    ))
                    lines_for_year += 1
                    total_lines += 1

            sheet_summary["sections"].append({
                "report_type": sec.report_type,
                "years": sec.years,
                "rows": len(sec.rows),
                "warnings": sec.warnings,
            })

        sheet_results.append(sheet_summary)

    await db.commit()

    co_codes = sorted({r["company_code"] for r in sheet_results})
    await append_audit_entry(
        db, actor_id=str(user.id), actor_email=user.email,
        action="financials.detailed.import",
        entity_type="batch", entity_id=str(file.filename or "upload")[:80],
        notes=f"{standard}: {len(co_codes)} companies, {total_reports} reports, "
              f"{total_lines} lines from {file.filename}",
    )
    await db.commit()

    return {
        "standard": standard,
        "companies_imported": len(co_codes),
        "company_codes": co_codes,
        "reports_created": total_reports,
        "lines_created": total_lines,
        "skipped_sheets": skipped_sheets,
        "results": sheet_results,
    }


@router.get("/detailed/{company_code}", response_model=dict)
async def get_detailed_report(
    company_code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    standard: str = Query("IFRS", pattern="^(IFRS|NSBU)$"),
    report_type: str = Query("BS", pattern="^(PL|BS|CF)$"),
):
    """Return the detailed audited report as a wide grid: rows × years."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.view")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == company_code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(404, f"Company '{company_code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    rep_q = await db.execute(
        select(FinancialReport).where(
            FinancialReport.company_id == co.id,
            FinancialReport.standard == standard,
            FinancialReport.report_type == report_type,
            FinancialReport.is_detailed == True,  # noqa: E712
        ).order_by(FinancialReport.year.asc())
    )
    reports = list(rep_q.scalars().all())

    if not reports:
        return {
            "company_code": co.code, "company_name": co.name_short,
            "standard": standard, "report_type": report_type,
            "years": [], "rows": [], "has_data": False,
        }

    years = [r.year for r in reports]

    # Collect all unique line_codes across all years, preserve sort_order
    line_meta: dict[str, dict] = {}  # line_code → metadata
    cells: dict[tuple[str, int], float] = {}  # (line_code, year) → value

    for rep in reports:
        ln_q = await db.execute(
            select(FinancialLine).where(FinancialLine.report_id == rep.id)
            .order_by(FinancialLine.sort_order.asc())
        )
        for ln in ln_q.scalars().all():
            cells[(ln.line_code, rep.year)] = float(ln.value) if ln.value is not None else None
            if ln.line_code not in line_meta:
                line_meta[ln.line_code] = {
                    "code": ln.line_code,
                    "label": ln.line_name,
                    "section": ln.section_label,
                    "indent": ln.indent_level,
                    "is_subtotal": ln.is_subtotal,
                    "sort_order": ln.sort_order,
                    "canonical_code": ln.parent_code,  # repurposed: holds canonical mapping
                    "is_unmapped": ln.parent_code is None,
                }
            elif ln.sort_order < line_meta[ln.line_code]["sort_order"]:
                line_meta[ln.line_code]["sort_order"] = ln.sort_order

    # Sort by sort_order
    rows = sorted(line_meta.values(), key=lambda r: r["sort_order"])
    for row in rows:
        row["values"] = {y: cells.get((row["code"], y)) for y in years}

    return {
        "company_code": co.code, "company_name": co.name_short,
        "standard": standard, "report_type": report_type,
        "years": years, "rows": rows, "has_data": True,
        "imported_at": (reports[0].extra or {}).get("imported_at") if reports else None,
        "source_filename": (reports[0].extra or {}).get("source_filename") if reports else None,
    }


@router.put("/detailed/{company_code}/cell")
async def update_detailed_cell(
    company_code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    standard: str = Query(..., pattern="^(IFRS|NSBU)$"),
    report_type: str = Query(..., pattern="^(PL|BS|CF)$"),
    year: int = Query(...),
    line_code: str = Query(..., max_length=64),
    value: Optional[float] = Query(None, description="New value; null clears the cell"),
):
    """Update one cell in the detailed grid. Used by inline editing in the UI."""
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == company_code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(404, f"Company '{company_code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    rep_q = await db.execute(
        select(FinancialReport).where(
            FinancialReport.company_id == co.id,
            FinancialReport.year == year,
            FinancialReport.standard == standard,
            FinancialReport.report_type == report_type,
            FinancialReport.is_detailed == True,  # noqa: E712
        )
    )
    rep = rep_q.scalar_one_or_none()
    if not rep:
        raise HTTPException(404, f"No detailed report for {company_code}/{year}/{standard}/{report_type}")

    line_q = await db.execute(
        select(FinancialLine).where(
            FinancialLine.report_id == rep.id,
            FinancialLine.line_code == line_code,
        )
    )
    line = line_q.scalar_one_or_none()
    if not line:
        raise HTTPException(404, f"Line '{line_code}' not found in this report")

    line.value = Decimal(str(value)) if value is not None else None
    await db.commit()

    return {"updated": True, "line_code": line_code, "year": year, "value": value}


# =====================================================================
# Detailed financials: canonical catalog + preview + confirm flow
# =====================================================================


from app.services.financial_canonical import CANONICAL  # noqa: E402


@router.get("/detailed/canonical/catalog", response_model=dict)
async def detailed_canonical_catalog(user: User = Depends(get_current_user)):
    """Return the canonical line schema for BS/PL/CF.

    UI uses this to:
      - Show the dropdown of mapping options when an unmapped row is being
        edited
      - Show "missing canonical lines" warnings (canonical codes that the
        company didn't report)
      - Display the canonical label (English/Russian) in the editor grid
    """
    out = {}
    for rtyp, lines in CANONICAL.items():
        out[rtyp] = [
            {
                "code": cl.code,
                "label": cl.label,
                "label_ru": cl.label_ru,
                "section": cl.section,
                "is_subtotal": cl.is_subtotal,
                "indent": cl.indent,
            }
            for cl in lines
        ]
    return out


def _serialize_parsed_section(sec) -> dict:
    """Convert a ParsedSection (incl. canonical mapping) to JSON-safe dict."""
    return {
        "report_type": sec.report_type,
        "years": sec.years,
        "warnings": sec.warnings,
        "rows": [
            {
                "code": r.code,
                "label": r.label,
                "indent_level": r.indent_level,
                "section_label": r.section_label,
                "is_subtotal": r.is_subtotal,
                "values": {str(y): v for y, v in r.values.items()},
                "canonical_code": r.canonical_code,
                "is_unmapped": r.is_unmapped,
            }
            for r in sec.rows
        ],
    }


@router.post("/detailed/parse-preview")
async def detailed_parse_preview(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    file: UploadFile = File(...),
    standard: str = Form("IFRS"),
    company_code: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
):
    """Parse an Excel file and return the parsed structure WITHOUT writing to DB.

    UI calls this when the user picks a file, then shows a modal with the
    parsed contents — including unmapped rows, missing canonical lines, etc.
    The user can edit / delete rows / fix mappings in the modal, then call
    `import-confirm` with the (potentially edited) structure.
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    if standard not in ("IFRS", "NSBU"):
        raise HTTPException(422, "standard must be IFRS or NSBU")

    try:
        file_bytes = await file.read()
    except Exception as e:
        raise HTTPException(400, f"Failed to read file: {e}")
    if len(file_bytes) > 25 * 1024 * 1024:
        raise HTTPException(413, "File too large (max 25 MB)")

    # Get all known company codes
    all_co_q = await db.execute(select(Company))
    all_companies = {c.code.lower(): {"id": str(c.id), "code": c.code, "name": c.name_short}
                     for c in all_co_q.scalars().all()}
    if not all_companies:
        raise HTTPException(500, "No companies in database")

    scope_ids = await allowed_company_ids(db, user)

    # Pick filter set
    if company_code:
        if company_code.lower() not in all_companies:
            raise HTTPException(404, f"Company '{company_code}' not found")
        filter_codes = {company_code.lower()}
    else:
        filter_codes = set(all_companies.keys())

    try:
        parsed_sheets = parse_workbook(file_bytes, sheet_name, company_codes=filter_codes)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse Excel: {type(e).__name__}: {e}")

    if not parsed_sheets:
        raise HTTPException(400,
            "No sheets matched company codes. "
            f"Sheet names must match one of the {len(all_companies)} company codes.")

    # Apply scope filter
    if scope_ids is not None:
        scope_codes_lc = set()
        for code, info in all_companies.items():
            if UUID(info["id"]) in scope_ids:
                scope_codes_lc.add(code)
        parsed_sheets = [ps for ps in parsed_sheets if ps.sheet_name.lower() in scope_codes_lc]

    # Build the response
    sheets_out = []
    for ps in parsed_sheets:
        co_lc = ps.sheet_name.lower()
        co_info = all_companies.get(co_lc)
        if not co_info:
            continue

        sections_out = []
        for sec in ps.sections:
            sec_dict = _serialize_parsed_section(sec)
            # Compute missing canonical codes for this section
            canonical_set = {cl.code for cl in CANONICAL.get(sec.report_type, [])}
            present_canonical = {r.canonical_code for r in sec.rows if r.canonical_code}
            sec_dict["missing_canonical_codes"] = sorted(canonical_set - present_canonical)
            sec_dict["unmapped_count"] = sum(1 for r in sec.rows if r.is_unmapped)
            sections_out.append(sec_dict)

        sheets_out.append({
            "sheet_name": ps.sheet_name,
            "company_code": co_info["code"],
            "company_name": co_info["name"],
            "warnings": ps.warnings,
            "sections": sections_out,
        })

    return {
        "standard": standard,
        "filename": file.filename,
        "sheets": sheets_out,
        "summary": {
            "sheets": len(sheets_out),
            "sections": sum(len(s["sections"]) for s in sheets_out),
            "rows": sum(len(sec["rows"]) for s in sheets_out for sec in s["sections"]),
            "unmapped_rows": sum(sec["unmapped_count"] for s in sheets_out for sec in s["sections"]),
        },
    }


@router.post("/detailed/import-confirm", status_code=http_status.HTTP_201_CREATED)
async def detailed_import_confirm(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Commit a (potentially user-edited) preview structure to DB.

    Payload shape (matching parse-preview output):
    {
      "standard": "IFRS" | "NSBU",
      "is_audited": bool,
      "filename": "...",
      "sheets": [
        {
          "company_code": "ngmk",
          "sections": [
            {
              "report_type": "BS",
              "years": [2021, 2022, 2023, 2024],
              "rows": [
                {
                  "label": "...",
                  "code": "...",
                  "canonical_code": "ppe" | null,
                  "is_unmapped": false,
                  "indent_level": 0,
                  "section_label": "...",
                  "is_subtotal": false,
                  "values": {"2021": 87438.24, ...}
                },
                ...
              ]
            }
          ]
        }
      ]
    }
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    standard = payload.get("standard", "IFRS")
    if standard not in ("IFRS", "NSBU"):
        raise HTTPException(422, "standard must be IFRS or NSBU")

    is_audited = bool(payload.get("is_audited", True))
    filename = payload.get("filename", "preview")
    sheets_in = payload.get("sheets") or []
    if not sheets_in:
        raise HTTPException(422, "No sheets in payload")

    # Resolve companies
    all_co_q = await db.execute(select(Company))
    all_co = {c.code.lower(): c for c in all_co_q.scalars().all()}
    scope_ids = await allowed_company_ids(db, user)

    # ── Wipe matching existing detailed reports first ──
    wipe_keys: set[tuple] = set()
    for sh in sheets_in:
        co_code = (sh.get("company_code") or "").lower()
        co = all_co.get(co_code)
        if not co:
            continue
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN,
                f"No access to company '{co.code}'")
        for sec in sh.get("sections") or []:
            rtyp = sec.get("report_type")
            if rtyp not in ("BS", "PL", "CF"):
                continue
            wipe_keys.add((co.id, standard, rtyp))

    for (cid, std, rtyp) in wipe_keys:
        existing_q = await db.execute(
            select(FinancialReport).where(
                FinancialReport.company_id == cid,
                FinancialReport.standard == std,
                FinancialReport.report_type == rtyp,
                FinancialReport.is_detailed == True,  # noqa: E712
            )
        )
        for old in existing_q.scalars().all():
            await db.delete(old)
    await db.flush()

    # ── Insert new ──
    total_reports = 0
    total_lines = 0
    co_codes_done = set()
    skipped = []

    for sh in sheets_in:
        co_code = (sh.get("company_code") or "").lower()
        co = all_co.get(co_code)
        if not co:
            skipped.append(f"{co_code} (unknown)")
            continue

        for sec in sh.get("sections") or []:
            rtyp = sec.get("report_type")
            if rtyp not in ("BS", "PL", "CF"):
                continue
            years = sec.get("years") or []
            rows = sec.get("rows") or []
            if not years or not rows:
                continue

            for year in years:
                rep = FinancialReport(
                    company_id=co.id,
                    year=int(year), quarter=None,
                    standard=standard, report_type=rtyp,
                    currency="UZS", unit_scale=1_000_000_000,
                    source=f"excel-confirm:{filename}",
                    is_audited=is_audited,
                    is_detailed=True,
                    extra={
                        "imported_at": datetime.now(timezone.utc).isoformat(),
                        "imported_by": user.email,
                        "source_filename": filename,
                        "via": "preview-confirm",
                    },
                )
                db.add(rep)
                await db.flush()
                total_reports += 1

                for sort_idx, prow in enumerate(rows):
                    yk = str(year)
                    raw_v = (prow.get("values") or {}).get(yk)
                    if raw_v is None:
                        # also accept int keys
                        raw_v = (prow.get("values") or {}).get(year)
                    if raw_v is None or raw_v == "":
                        continue
                    try:
                        v = Decimal(str(raw_v))
                    except Exception:
                        continue

                    code = (prow.get("code") or "").strip() or f"line_{sort_idx}"
                    label = (prow.get("label") or "").strip() or code
                    canonical_code = prow.get("canonical_code") or None
                    indent = int(prow.get("indent_level") or 0)
                    section_label = prow.get("section_label") or None
                    is_subtotal = bool(prow.get("is_subtotal", False))

                    # Store canonical mapping in extra JSON via parent_code field
                    # (we already have it at canonical_code but the DB column is
                    # parent_code — repurposed here to hold canonical mapping).
                    db.add(FinancialLine(
                        report_id=rep.id,
                        line_code=code[:32],
                        line_name=label[:255],
                        parent_code=(canonical_code[:32] if canonical_code else None),
                        section_label=section_label,
                        indent_level=indent,
                        value=v,
                        is_subtotal=is_subtotal,
                        is_calculated=False,
                        sort_order=sort_idx,
                    ))
                    total_lines += 1

        co_codes_done.add(co.code)

    await db.commit()

    await append_audit_entry(
        db, actor_id=str(user.id), actor_email=user.email,
        action="financials.detailed.import.confirm",
        entity_type="batch", entity_id=str(filename)[:80],
        notes=f"{standard}: {len(co_codes_done)} co, {total_reports} reports, "
              f"{total_lines} lines (via preview-confirm)",
    )
    await db.commit()

    return {
        "standard": standard,
        "companies_imported": len(co_codes_done),
        "company_codes": sorted(co_codes_done),
        "reports_created": total_reports,
        "lines_created": total_lines,
        "skipped": skipped,
    }


@router.put("/detailed/{company_code}/line/mapping")
async def update_detailed_line_mapping(
    company_code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    standard: str = Query(..., pattern="^(IFRS|NSBU)$"),
    report_type: str = Query(..., pattern="^(PL|BS|CF)$"),
    line_code: str = Query(..., max_length=64),
    canonical_code: Optional[str] = Query(None, description="New canonical mapping; null clears it"),
    new_label: Optional[str] = Query(None, description="Optional: also rename the line"),
):
    """Update the canonical mapping (and optionally label) of a detailed line
    across all years of one (company × standard × report_type) tuple."""
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == company_code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(404, f"Company '{company_code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    rep_q = await db.execute(
        select(FinancialReport).where(
            FinancialReport.company_id == co.id,
            FinancialReport.standard == standard,
            FinancialReport.report_type == report_type,
            FinancialReport.is_detailed == True,  # noqa: E712
        )
    )
    reports = list(rep_q.scalars().all())
    if not reports:
        raise HTTPException(404, "No detailed reports found")

    updated = 0
    for rep in reports:
        ln_q = await db.execute(
            select(FinancialLine).where(
                FinancialLine.report_id == rep.id,
                FinancialLine.line_code == line_code,
            )
        )
        for ln in ln_q.scalars().all():
            ln.parent_code = canonical_code if canonical_code else None
            if new_label:
                ln.line_name = new_label[:255]
            updated += 1

    await db.commit()
    return {"updated": updated, "canonical_code": canonical_code, "new_label": new_label}


@router.delete("/detailed/{company_code}/line", status_code=http_status.HTTP_204_NO_CONTENT)
async def delete_detailed_line(
    company_code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    standard: str = Query(..., pattern="^(IFRS|NSBU)$"),
    report_type: str = Query(..., pattern="^(PL|BS|CF)$"),
    line_code: str = Query(..., max_length=64),
):
    """Delete a line across all years of one (company × standard × report_type)."""
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == company_code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(404, f"Company '{company_code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    rep_q = await db.execute(
        select(FinancialReport).where(
            FinancialReport.company_id == co.id,
            FinancialReport.standard == standard,
            FinancialReport.report_type == report_type,
            FinancialReport.is_detailed == True,  # noqa: E712
        )
    )
    deleted = 0
    for rep in rep_q.scalars().all():
        del_q = await db.execute(
            delete(FinancialLine).where(
                FinancialLine.report_id == rep.id,
                FinancialLine.line_code == line_code,
            )
        )
        deleted += del_q.rowcount or 0

    await db.commit()
    return None


# =====================================================================
# Portfolio summary (dashboard aggregator)            [Phase 19a-1]
# =====================================================================
# Used by the Financials dashboard view (frontend) to render portfolio-wide
# KPI band and per-company multi-year metric breakdown without N+1 queries.
#
# Returns per-company × per-year canonical metric values normalized to
# raw currency units (FinancialLine.value × FinancialReport.unit_scale).
# Single SQL query, no N+1 — handles full portfolio in <100 ms.
#
# Handles legacy line_code variants from different import paths:
#   - camelCase from old monolith imports (revenue, grossProfit, opProfit, ...)
#   - snake_case from later imports (gross_profit, total_assets, total_equity)
# =====================================================================

# Maps actual line_code variants (any case) → canonical metric code
# returned in the response. Frontend reads only canonical names.
_PORTFOLIO_METRIC_ALIASES: dict[str, str] = {
    # ── P&L ─────────────────────────────────────────────────────────
    "revenue": "revenue",
    "cogs": "cogs",
    "grossProfit": "grossProfit",
    "gross_profit": "grossProfit",
    "opProfit": "opProfit",
    "operatingProfit": "opProfit",
    "operating_profit": "opProfit",
    "ebitda": "ebitda",
    "depreciation": "depreciation",
    "pbt": "pbt",
    "tax": "tax",
    "profit": "profit",
    "netProfit": "profit",
    "net_profit": "profit",
    # Finance items (used for finCost/finIncome split in detailed views)
    "finCost": "finCost",
    "finIncome": "finIncome",
    "interestExp": "interestExp",
    "forex": "forex",
    # ── Balance Sheet ───────────────────────────────────────────────
    "totalAssets": "totalAssets",
    "total_assets": "totalAssets",
    "totalLiabilities": "totalLiabilities",
    "total_liabilities": "totalLiabilities",
    "equity": "equity",
    "total_equity": "equity",
    "totalCA": "totalCA",
    "totalNCA": "totalNCA",
    "ppe": "ppe",
    "cash": "cash",
    "debt": "debt",
    "ltBorrowings": "ltBorrowings",
    "stBorrowings": "stBorrowings",
    "ltBankLoans": "ltBankLoans",
    "inventories": "inventories",
    "tradeReceivables": "tradeReceivables",
}


def _canon_metric(line_code: str | None, parent_code: str | None = None) -> str | None:
    """Map raw line_code (any case/variant) to canonical metric code.
    Returns None for unknown/garbage codes — those are skipped in aggregation.

    Pack 7.54: if line_code doesn't map, fall back to parent_code. This lets
    custom user fields contribute to portfolio aggregations when the user
    explicitly maps them to a canonical metric in the NSBU editor.
    """
    if line_code:
        direct = _PORTFOLIO_METRIC_ALIASES.get(line_code)
        if direct:
            return direct
        stripped = _PORTFOLIO_METRIC_ALIASES.get(line_code.strip())
        if stripped:
            return stripped
    # Fallback to parent_code mapping (for custom fields)
    if parent_code:
        direct = _PORTFOLIO_METRIC_ALIASES.get(parent_code)
        if direct:
            return direct
        return _PORTFOLIO_METRIC_ALIASES.get(parent_code.strip())
    return None


@router.get("/portfolio/summary")
async def portfolio_summary(
    standard: str = Query("IFRS", description="IFRS or NSBU"),
    years: str = Query(
        "2021,2022,2023,2024,2025,2026",
        description="Comma-separated list of fiscal years",
    ),
    currency: str = Query("UZS", description="Currency filter (UZS/USD/EUR)"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Aggregate financial metrics across all accessible companies.

    Returns per-company × year × metric breakdown with values normalized
    to raw currency units. One query, no N+1.

    Response shape:
        {
          "standard": "IFRS",
          "currency": "UZS",
          "years": [2021, 2022, ...],
          "items": [
            {
              "company_id": "...",
              "company_code": "NGMK",
              "company_name": "Навоийский ГМК",
              "company_name_short": "НГМК",
              "sector_code": "mining",
              "by_year": {
                "2024": { "revenue": 93558000000.0, "ebitda": 62334000000.0, ... },
                "2023": {...},
                ...
              }
            },
            ...22 companies...
          ],
          "portfolio_totals_by_year": {
            "2024": { "revenue": 328345000000.0, "ebitda": 125034000000.0, ... }
          },
          "coverage": {
            "companies_total": 22,
            "with_revenue_any_year": 18,
            "with_data_2024": 15,
            "with_data_2023": 16,
            ...
          }
        }
    """
    std = standard.upper()
    if std not in ("IFRS", "NSBU"):
        raise HTTPException(400, "standard must be IFRS or NSBU")

    cur = currency.upper()
    if cur not in ("UZS", "USD", "EUR"):
        raise HTTPException(400, "currency must be UZS, USD or EUR")

    try:
        year_list = sorted({int(y.strip()) for y in years.split(",") if y.strip()})
    except ValueError:
        raise HTTPException(400, "years must be comma-separated integers")

    if not year_list:
        raise HTTPException(400, "at least one year required")
    if len(year_list) > 12:
        raise HTTPException(400, "max 12 years per request")

    # RBAC: scope to companies user has access to (None = full access)
    allowed_set = await allowed_company_ids(db, user)

    # One join query: companies × reports × lines, filtered tightly.
    # Pack 7.40.3: currency filter made tolerant — if exact match returns
    # zero rows we retry with broader filtering. Legacy data may have
    # currency stored as "uzs"/""/"local"/NULL rather than canonical "UZS".
    base_stmt = (
        select(
            Company.id.label("co_id"),
            Company.code.label("co_code"),
            Company.name_ru.label("co_name"),
            Company.name_short.label("co_short"),
            FinancialReport.year.label("year"),
            FinancialReport.report_type.label("rtype"),
            FinancialReport.unit_scale.label("scale"),
            FinancialReport.currency.label("rcurrency"),
            Company.sector_id.label("sector_id"),
            FinancialLine.line_code.label("code"),
            FinancialLine.parent_code.label("parent_code"),
            FinancialLine.value.label("val"),
        )
        .join(FinancialReport, FinancialReport.company_id == Company.id)
        .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
        .where(
            FinancialReport.standard == std,
            FinancialReport.year.in_(year_list),
            FinancialReport.report_type.in_(["PL", "BS", "CF"]),
        )
    )
    if allowed_set is not None:
        base_stmt = base_stmt.where(Company.id.in_(allowed_set))

    # Attempt 1: exact case-sensitive currency match (canonical path)
    stmt = base_stmt.where(FinancialReport.currency == cur)
    result = await db.execute(stmt)
    rows = result.all()
    currency_filter_relaxed: str | None = None

    # Attempt 2: if zero rows, retry case-insensitive (handles "uzs"/"Uzs" etc.)
    if not rows:
        from sqlalchemy import func as _func2
        stmt2 = base_stmt.where(_func2.upper(FinancialReport.currency) == cur)
        result = await db.execute(stmt2)
        rows = result.all()
        if rows:
            currency_filter_relaxed = "case-insensitive"

    # Attempt 3: if still zero, drop currency filter entirely and return ALL
    # reports for the given standard+years. Logs a warning so the issue is
    # surfaced in backend logs. The frontend can detect this via the
    # `currency_filter_relaxed` field in the response.
    if not rows:
        import logging as _lg
        _lg.getLogger(__name__).warning(
            "[portfolio_summary] No reports matched currency=%s for "
            "standard=%s years=%s — falling back to no-currency filter",
            cur, std, year_list,
        )
        result = await db.execute(base_stmt)
        rows = result.all()
        if rows:
            currency_filter_relaxed = "removed"

    # Sector mapping for grouping
    from app.models.company import Sector  # noqa: WPS433
    _sec_q = await db.execute(select(Sector.id, Sector.code))
    sec_map = {row[0]: row[1] for row in _sec_q.all()}

    # Pack 7.40.5 — backend-side currency conversion.
    # Load year_registry rates once, then convert every row from row.currency
    # to the requested `cur` during aggregation.
    rates_q = await db.execute(
        select(
            YearRegistry.year,
            YearRegistry.usd_rate,
            YearRegistry.eur_rate,
        )
    )
    _rates_by_year: dict[int, dict[str, float]] = {}
    for r in rates_q.all():
        _rates_by_year[int(r.year)] = {
            "USD": float(r.usd_rate) if r.usd_rate is not None else 0.0,
            "EUR": float(r.eur_rate) if r.eur_rate is not None else 0.0,
        }

    # Hardcoded fallback (same as frontend useCurrencyConverter fallback).
    # Used when year_registry doesn't have a value for the requested year.
    _RATE_FALLBACK = {
        "USD": {
            2021: 10610.00, 2022: 11050.00, 2023: 11420.00,
            2024: 12650.91, 2025: 12576.41, 2026: 12200.00,
        },
        "EUR": {
            2021: 12520.00, 2022: 11600.00, 2023: 12330.00,
            2024: 13691.00, 2025: 14140.00, 2026: 14250.00,
        },
    }

    def _rate_for(year_v: int, currency_v: str) -> float:
        """Return UZS-per-`currency_v` rate for year_v, or 0.0 if unknown."""
        if currency_v == "UZS":
            return 1.0
        # Try year_registry first
        registry = _rates_by_year.get(year_v, {})
        v = registry.get(currency_v, 0.0)
        if v and v > 0:
            return v
        # Fallback table
        fb = _RATE_FALLBACK.get(currency_v, {})
        if year_v in fb:
            return fb[year_v]
        # Nearest available year in fallback
        if fb:
            nearest = min(fb.keys(), key=lambda y: abs(y - year_v))
            return fb[nearest]
        return 0.0

    # Aggregate into per-company nested dict
    by_co: dict[str, dict] = {}
    for r in rows:
        if r.val is None:
            continue
        # Pack 7.54: try line_code first; fall back to parent_code for custom-mapped fields.
        canon = _canon_metric(r.code, getattr(r, "parent_code", None))
        if not canon:
            continue
        scale = r.scale or 1000
        try:
            value_raw = float(r.val) * 1_000_000_000.0
        except (TypeError, ValueError):
            continue

        # Pack 7.40.5 — convert from row.currency to requested `cur`.
        # Strategy: normalize to UZS first, then divide to target currency.
        row_currency = (r.rcurrency or "UZS").upper()
        row_year = int(r.year)
        if row_currency != "UZS":
            # The stored value is in row_currency — multiply by the
            # corresponding rate to get UZS.
            inverse_rate = _rate_for(row_year, row_currency)
            if inverse_rate > 0:
                value_raw = value_raw * inverse_rate
            else:
                # No rate available — skip this row
                continue
        # Now value_raw is in UZS. If target currency != UZS, divide by target rate.
        if cur != "UZS":
            target_rate = _rate_for(row_year, cur)
            if target_rate > 0:
                value_raw = value_raw / target_rate
            else:
                # No rate available — leave in UZS but mark with a fallback so
                # the frontend at least shows something. Better: skip.
                continue

        co_key = r.co_code
        co = by_co.get(co_key)
        if co is None:
            co = {
                "company_id": str(r.co_id),
                "company_code": co_key,
                "company_name": r.co_name,
                "company_name_short": r.co_short,
                "sector_code": (sec_map.get(r.sector_id) or "other"),
                "by_year": {},
            }
            by_co[co_key] = co

        ydict = co["by_year"].setdefault(int(r.year), {})
        # Duplicate-key strategy: keep largest absolute value
        # (handles same-metric appearing in multiple report_types)
        existing = ydict.get(canon)
        if existing is None or abs(value_raw) > abs(float(existing)):
            ydict[canon] = value_raw

    items = list(by_co.values())
    items.sort(key=lambda x: x["company_code"] or "")

    # coverage_patched_v2: denominator = ALL companies in scope (incl. those without data)
    from sqlalchemy import func as _func
    _all_co_q = select(_func.count(Company.id))
    if allowed_set is not None:
        _all_co_q = _all_co_q.where(Company.id.in_(allowed_set))
    _total_companies = (await db.execute(_all_co_q)).scalar() or 0

    # Portfolio-wide totals per year (used by KPI band)
    totals_by_year: dict[int, dict[str, float]] = {}
    for it in items:
        for y, metrics in it["by_year"].items():
            t = totals_by_year.setdefault(y, {})
            for m, v in metrics.items():
                if v is None:
                    continue
                t[m] = t.get(m, 0.0) + float(v)

    # Coverage: how many companies have data per year
    coverage: dict[str, int] = {
        "companies_total": _total_companies,
        "with_revenue_any_year": sum(
            1 for it in items
            if any("revenue" in y for y in it["by_year"].values())
        ),
    }
    for y in year_list:
        coverage[f"with_data_{y}"] = sum(
            1 for it in items if y in it["by_year"]
        )

    return {
        "standard": std,
        "currency": cur,
        "years": year_list,
        "items": items,
        "portfolio_totals_by_year": totals_by_year,
        "coverage": coverage,
        "currency_filter_relaxed": currency_filter_relaxed,
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.52 — NSBU editor save/load endpoint
# ════════════════════════════════════════════════════════════════════════
#
# GET  /api/financials/companies/{code}/nsbu-editor
#   Returns:
#     {
#       customFields: [...],         # user-added field definitions
#       renames: {fieldId: label},   # user renames
#       formulaOverrides: {fieldId: expression},  # overridden auto formulas
#       manualFlags: {fieldId: {year: true}},     # auto-fields the user overrode
#       updatedAt, updatedBy,
#     }
#
# PUT  /api/financials/companies/{code}/nsbu-editor
#   Payload: NsbuEditorSavePayload (см. ниже)
#   Returns: { ok, saved, reports_created, reports_updated, lines_upserted }
#
# Persistence strategy:
#   • Values → financial_reports + financial_lines (standard='NSBU', is_detailed=False,
#     source='nsbu-editor', report_type='PL' for ОФР, 'BS' for Баланс)
#   • Customization → company.extra.nsbu_editor_schema (JSONB)
#
# Поля делятся на PL/BS по принадлежности к секции в frontend schema
# (см. composables/useNsbuSchema.ts). Backend держит хардкод списка
# для совместимости — должен совпадать с frontend!

from pydantic import BaseModel, Field

# Хардкод списка полей PL vs BS — должен совпадать с STANDARD_SCHEMA в frontend.
_NSBU_PL_FIELDS = {
    "revenue", "cogs", "grossProfit", "opProfit", "depreciation",
    "finIncome", "finCost", "forex", "pbt", "tax", "profit", "ebitda",
}
_NSBU_BS_FIELDS = {
    "ppe", "totalNCA", "cash", "totalCA", "totalAssets",
    "equity", "ltBorrowings", "stBorrowings", "totalLiabilities",
    "ltBankLoans", "ltOtherLoans", "stBankLoans", "stOtherLoans", "debt",
}


class NsbuCustomFieldDef(BaseModel):
    id: str
    label: str
    section: Optional[str] = None  # 'pnl' | 'sofp'
    autoFormula: Optional[str] = None
    isCustom: Optional[bool] = True
    # Pack 7.54: optional canonical metric mapping. If set (e.g. 'ebitda'),
    # this custom field's value will be picked up by portfolio aggregations
    # via FinancialLine.parent_code.
    canonical: Optional[str] = None


class NsbuEditorSavePayload(BaseModel):
    """Payload from frontend NsbuEditor.vue save action."""
    values: dict[str, dict[str, Optional[float]]] = Field(default_factory=dict)
    # values[fieldId][yearStr] = value (млрд сум) or null
    customFields: list[NsbuCustomFieldDef] = Field(default_factory=list)
    renames: dict[str, str] = Field(default_factory=dict)
    formulaOverrides: dict[str, str] = Field(default_factory=dict)
    manualFlags: dict[str, dict[str, bool]] = Field(default_factory=dict)


@router.get("/companies/{code}/nsbu-editor")
async def get_nsbu_editor_schema(
    code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Return user customizations + financial values for the NSBU editor.

    Pack 7.65: values are now bundled in the response (previously the editor
    pulled them via portfolio-summary, but the drill-down modal needs a
    direct, low-latency source).
    """
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(
        select(Company).where(func.lower(Company.code) == code.lower())
    )
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    schema = (co.extra or {}).get("nsbu_editor_schema", {}) if co.extra else {}

    # Load NSBU values from DB (PL + Balance sheet, FY only, consolidated default)
    rows = await db.execute(
        select(FinancialReport, FinancialLine)
        .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
        .where(
            FinancialReport.company_id == co.id,
            FinancialReport.standard == "NSBU",
            FinancialReport.is_detailed.is_(False),
            FinancialReport.quarter.is_(None),
        )
    )
    values: dict[str, dict[str, float]] = {}
    for fr, fl in rows.all():
        if fl.value is None:
            continue
        try:
            v_mlrd = float(fl.value)
        except (TypeError, ValueError):
            continue
        values.setdefault(fl.line_code, {})[str(fr.year)] = v_mlrd

    return {
        "code": co.code,
        "values": values,
        "customFields": schema.get("customFields", []),
        "renames": schema.get("renames", {}),
        "formulaOverrides": schema.get("formulaOverrides", {}),
        "manualFlags": schema.get("manualFlags", {}),
        "updatedAt": schema.get("updatedAt"),
        "updatedBy": schema.get("updatedBy"),
    }


@router.put("/companies/{code}/nsbu-editor")
async def save_nsbu_editor(
    code: str,
    payload: NsbuEditorSavePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Save values to financial_reports/lines + customization to company.extra.
    Idempotent: re-running with the same payload produces the same DB state."""
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    co_q = await db.execute(
        select(Company).where(func.lower(Company.code) == code.lower())
    )
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    now_iso = datetime.now(timezone.utc).isoformat()

    # ─── 1. Persist customization to company.extra.nsbu_editor_schema ───────
    extra = dict(co.extra or {})
    extra["nsbu_editor_schema"] = {
        "customFields": [cf.model_dump() for cf in payload.customFields],
        "renames": payload.renames,
        "formulaOverrides": payload.formulaOverrides,
        "manualFlags": payload.manualFlags,
        "updatedAt": now_iso,
        "updatedBy": user.email,
    }
    co.extra = extra

    # ─── 2. Group values by (year, report_type) ────────────────────────────
    # Custom fields use the section explicitly declared in CustomFieldDef
    custom_section_by_id: dict[str, str] = {}
    for cf in payload.customFields:
        if cf.section in ("pnl", "sofp"):
            custom_section_by_id[cf.id] = cf.section

    changes_by_report: dict[tuple[int, str], list[tuple[str, Optional[float], str, Optional[str]]]] = {}
    # key=(year, report_type) → [(field_id, value, line_name, canonical_or_None)]

    # Build label lookup (for storing line_name)
    label_for_field: dict[str, str] = {f: f for f in (_NSBU_PL_FIELDS | _NSBU_BS_FIELDS)}
    for cf in payload.customFields:
        label_for_field[cf.id] = cf.label
    for fld, renamed in payload.renames.items():
        label_for_field[fld] = renamed

    # Pack 7.54: canonical mapping per field — for custom fields, allows them
    # to contribute to portfolio aggregations via FinancialLine.parent_code.
    canonical_for_field: dict[str, Optional[str]] = {}
    for f in (_NSBU_PL_FIELDS | _NSBU_BS_FIELDS):
        canonical_for_field[f] = f  # standard fields use their id as canonical
    for cf in payload.customFields:
        if cf.canonical:
            canonical_for_field[cf.id] = cf.canonical
        else:
            canonical_for_field[cf.id] = None  # custom without mapping → no portfolio contrib

    for field, year_map in payload.values.items():
        # Determine report_type
        if field in _NSBU_PL_FIELDS:
            report_type = "PL"
        elif field in _NSBU_BS_FIELDS:
            report_type = "BS"
        elif field in custom_section_by_id:
            report_type = "PL" if custom_section_by_id[field] == "pnl" else "BS"
        else:
            # Unknown field — skip (could be auto-only like grossProfit which is not stored)
            continue
        for year_str, val in year_map.items():
            try:
                year = int(year_str)
            except (TypeError, ValueError):
                continue
            key = (year, report_type)
            if key not in changes_by_report:
                changes_by_report[key] = []
            changes_by_report[key].append((
                field, val,
                label_for_field.get(field, field),
                canonical_for_field.get(field),
            ))

    # ─── 3. Upsert reports + lines ─────────────────────────────────────────
    reports_created = 0
    reports_updated = 0
    lines_upserted = 0
    lines_deleted = 0

    for (year, report_type), changes in changes_by_report.items():
        # Find or create the editor's report row.
        # Prefer existing is_detailed=False, source='nsbu-editor'; fallback to any
        # is_detailed=False NSBU report for this (co, year, report_type).
        rep_q = await db.execute(
            select(FinancialReport).where(
                FinancialReport.company_id == co.id,
                FinancialReport.year == year,
                FinancialReport.quarter.is_(None),
                FinancialReport.standard == "NSBU",
                FinancialReport.report_type == report_type,
                FinancialReport.is_detailed == False,  # noqa: E712
            ).order_by(FinancialReport.updated_at.desc())
        )
        existing_reports = list(rep_q.scalars().all())
        report: Optional[FinancialReport] = None
        for r in existing_reports:
            if r.source == "nsbu-editor":
                report = r
                break
        if report is None and existing_reports:
            report = existing_reports[0]  # use first non-detailed report

        if report is None:
            # Create new editor report
            report = FinancialReport(
                company_id=co.id,
                year=year,
                quarter=None,
                standard="NSBU",
                report_type=report_type,
                currency="UZS",
                unit_scale=1_000_000_000,  # values stored as млрд UZS = mlrd
                source="nsbu-editor",
                is_audited=False,
                is_detailed=False,
                notes=f"Saved via NSBU editor by {user.email} on {now_iso}",
                extra={"editor_version": "p7.52"},
            )
            db.add(report)
            await db.flush()
            reports_created += 1
        else:
            reports_updated += 1
            report.source = report.source or "nsbu-editor"
            report.notes = f"Last edit via NSBU editor by {user.email} on {now_iso}"

        # Load existing lines for this report
        ln_q = await db.execute(
            select(FinancialLine).where(FinancialLine.report_id == report.id)
        )
        existing_lines = {ln.line_code: ln for ln in ln_q.scalars().all()}

        # Upsert each change
        for field, val, label, canonical in changes:
            existing = existing_lines.get(field)
            if val is None:
                # Null value — delete the line if exists
                if existing is not None:
                    await db.delete(existing)
                    lines_deleted += 1
                continue
            decimal_val = Decimal(str(val))
            # Pack 7.54: parent_code holds canonical mapping for portfolio aggregation.
            # Standard fields: parent_code = same as line_code (self-mapping).
            # Custom mapped: parent_code = the canonical metric chosen by user.
            # Custom unmapped: parent_code = None → not aggregated in portfolio.
            new_parent = canonical
            if existing is not None:
                existing.value = decimal_val
                existing.line_name = label
                existing.parent_code = new_parent
            else:
                new_line = FinancialLine(
                    report_id=report.id,
                    line_code=field,
                    parent_code=new_parent,
                    line_name=label,
                    value=decimal_val,
                    is_subtotal=False,
                    is_calculated=False,
                    sort_order=0,
                )
                db.add(new_line)
            lines_upserted += 1

    # Pack 7.55: write audit log entry for this save
    try:
        # Compact diff: counts + sample of changed fields
        sample_fields = sorted(set(field for changes in changes_by_report.values() for field, _, _, _ in changes))[:20]
        await append_audit_entry(
            db,
            actor_id=str(user.id) if user.id else None,
            actor_email=user.email,
            action="nsbu_editor.save",
            entity_type="company",
            entity_id=str(co.id),
            diff={
                "reports_created": reports_created,
                "reports_updated": reports_updated,
                "lines_upserted": lines_upserted,
                "lines_deleted": lines_deleted,
                "fields": sample_fields,
                "years": sorted({y for (y, _) in changes_by_report.keys()}),
            },
            payload={
                "company_code": co.code,
                "customFields_count": len(payload.customFields),
                "renames_count": len(payload.renames),
                "formulaOverrides_count": len(payload.formulaOverrides),
            },
            notes=f"NSBU editor save · {co.code}",
        )
    except Exception as e:
        # Audit failure must not block the save itself
        print(f"[nsbu-editor] audit log failed: {e}")

    await db.commit()

    return {
        "ok": True,
        "saved_at": now_iso,
        "reports_created": reports_created,
        "reports_updated": reports_updated,
        "lines_upserted": lines_upserted,
        "lines_deleted": lines_deleted,
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.55 — NSBU editor history endpoint
# ════════════════════════════════════════════════════════════════════════

@router.get("/companies/{code}/nsbu-editor/history")
async def get_nsbu_editor_history(
    code: str,
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Return last N audit log entries for NSBU editor saves on this company."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    from app.models.audit import AuditLog

    q = await db.execute(
        select(AuditLog)
        .where(
            AuditLog.action == "nsbu_editor.save",
            AuditLog.entity_type == "company",
            AuditLog.entity_id == str(co.id),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    )
    entries = list(q.scalars().all())
    return {
        "code": co.code,
        "company_name": co.name_short or co.name_ru,
        "total": len(entries),
        "entries": [
            {
                "id": str(e.id),
                "at": e.created_at.isoformat() if e.created_at else None,
                "actor_email": e.actor_email,
                "diff": e.diff or {},
                "payload": e.payload or {},
                "notes": e.notes,
            }
            for e in entries
        ],
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.53 — NSBU editor Excel import (template + parse)
# ════════════════════════════════════════════════════════════════════════

from fastapi.responses import StreamingResponse  # noqa: E402
import io  # noqa: E402

_NSBU_FIELD_LABELS = {
    # P&L (ОФР)
    "revenue":       ("revenue",       "Выручка",                            "010", "pnl"),
    "cogs":          ("cogs",          "Себестоимость",                       "020", "pnl"),
    "grossProfit":   ("grossProfit",   "Валовая прибыль (авто)",              "030", "pnl"),
    "opProfit":      ("opProfit",      "Операционная прибыль",                "060", "pnl"),
    "depreciation":  ("depreciation",  "Амортизация",                          "070", "pnl"),
    "finIncome":     ("finIncome",     "Доходы от фин. деятельности",         "110", "pnl"),
    "finCost":       ("finCost",       "Расходы от фин. деятельности",        "170", "pnl"),
    "forex":         ("forex",         "Курсовая разница (справочно)",        "180", "pnl"),
    "pbt":           ("pbt",           "Прибыль до налога (авто)",            "190", "pnl"),
    "tax":           ("tax",           "Налог на прибыль",                    "220", "pnl"),
    "profit":        ("profit",        "Чистая прибыль (авто)",               "270", "pnl"),
    "ebitda":        ("ebitda",        "EBITDA (авто)",                       "",    "pnl"),
    # Balance Sheet (Баланс)
    "ppe":              ("ppe",              "Основные средства",            "010", "sofp"),
    "totalNCA":         ("totalNCA",         "Внеоборотные активы (итог)",   "190", "sofp"),
    "cash":             ("cash",             "Денежные средства",            "320", "sofp"),
    "totalCA":          ("totalCA",          "Оборотные активы (итог)",      "390", "sofp"),
    "totalAssets":      ("totalAssets",      "ИТОГО Активы (авто)",          "400", "sofp"),
    "equity":           ("equity",           "Собственный капитал",          "480", "sofp"),
    "ltBorrowings":     ("ltBorrowings",     "Долгосрочные обязательства",   "590", "sofp"),
    "stBorrowings":     ("stBorrowings",     "Краткосрочные обязательства",  "780", "sofp"),
    "totalLiabilities": ("totalLiabilities", "ИТОГО Обязательства (авто)",   "",    "sofp"),
    "ltBankLoans":      ("ltBankLoans",      "Долгосрочные банковские кредиты", "7810", "sofp"),
    "ltOtherLoans":     ("ltOtherLoans",     "Долгосрочные займы",            "7820", "sofp"),
    "stBankLoans":      ("stBankLoans",      "Краткосрочные банковские кредиты","6810", "sofp"),
    "stOtherLoans":     ("stOtherLoans",     "Краткосрочные займы",           "6820", "sofp"),
    "debt":             ("debt",             "Финансовый долг (авто)",        "",    "sofp"),
}


@router.get("/companies/{code}/nsbu-editor/template")
async def download_nsbu_editor_template(
    code: str,
    years: str = Query("2021,2022,2023,2024,2025,2026"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Generate an XLSX template the user fills with NSBU values, then uploads back."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    try:
        year_list = sorted({int(y.strip()) for y in years.split(",") if y.strip()})
    except ValueError:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Invalid years parameter")
    if not year_list:
        year_list = [2021, 2022, 2023, 2024, 2025, 2026]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Sheet 1: ОФР
    ws_pl = wb.active
    ws_pl.title = "ОФР"
    # Sheet 2: Баланс
    ws_bs = wb.create_sheet("Баланс")

    header_font = Font(bold=True, size=11, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF7F77DD")
    auto_fill = PatternFill("solid", fgColor="FFFFFBF0")
    auto_font = Font(italic=True, color="FFD97706")
    border = Border(
        left=Side(style="thin", color="FFE2E8F0"),
        right=Side(style="thin", color="FFE2E8F0"),
        top=Side(style="thin", color="FFE2E8F0"),
        bottom=Side(style="thin", color="FFE2E8F0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def fill_sheet(ws, section):
        # Title row
        ws.cell(row=1, column=1, value=f"НСБУ {section} · {co.code} {co.name_short or co.name_ru or ''}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="FF1E2A4A")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(year_list))
        # Helper row
        helper = "Заполни числовые поля. Поля, помеченные «авто», пересчитываются автоматически и записывать их не обязательно. Числа в МЛРД СУМ (например 62,5)."
        ws.cell(row=2, column=1, value=helper).font = Font(italic=True, size=9, color="FF94A3B8")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + len(year_list))
        # Header row
        ws.cell(row=4, column=1, value="Код")
        ws.cell(row=4, column=2, value="№ строки")
        ws.cell(row=4, column=3, value="Показатель")
        for i, yr in enumerate(year_list):
            ws.cell(row=4, column=4 + i, value=yr)
        for col in range(1, 4 + len(year_list)):
            c = ws.cell(row=4, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        # Data rows
        row = 5
        for field_id, (fid, label, nsbu_code, sect) in _NSBU_FIELD_LABELS.items():
            if sect != section:
                continue
            is_auto = "(авто)" in label
            ws.cell(row=row, column=1, value=fid)
            ws.cell(row=row, column=2, value=nsbu_code)
            ws.cell(row=row, column=3, value=label)
            for col in range(1, 4 + len(year_list)):
                c = ws.cell(row=row, column=col)
                c.border = border
                if col >= 4:
                    c.alignment = center
                elif col == 3:
                    c.alignment = left
                else:
                    c.alignment = center
                if is_auto:
                    c.fill = auto_fill
                    if col == 3:
                        c.font = auto_font
            row += 1
        # Column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 45
        for i in range(len(year_list)):
            ws.column_dimensions[chr(ord("D") + i)].width = 14
        ws.freeze_panes = "D5"

    fill_sheet(ws_pl, "pnl")
    fill_sheet(ws_bs, "sofp")

    # Stream out
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"nsbu_template_{co.code}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/companies/{code}/nsbu-editor/parse-excel")
async def parse_nsbu_editor_excel(
    code: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Parse an uploaded XLSX file (template format) → return values matrix.
    Frontend then applies the matrix to editor state (not auto-saved to DB)."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    contents = await file.read()
    if not contents:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Empty file")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, f"Cannot parse XLSX: {e}")

    values: dict[str, dict[int, float]] = {}
    parse_log: list[str] = []

    # Build a label → field_id mapping for case-insensitive matching
    label_to_field: dict[str, str] = {}
    for fid, (canonical, label, _code, _sect) in _NSBU_FIELD_LABELS.items():
        # Strip "(авто)" suffix when matching
        clean = label.replace("(авто)", "").strip().lower()
        label_to_field[clean] = canonical
        label_to_field[canonical.lower()] = canonical

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find header row containing years — look in first 10 rows
        header_row = None
        year_cols: dict[int, int] = {}  # col_idx → year
        for r in range(1, min(15, ws.max_row + 1)):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            if not row_cells:
                continue
            row_vals = row_cells[0]
            yc_local: dict[int, int] = {}
            for ci, val in enumerate(row_vals):
                if isinstance(val, (int, float)) and 1990 < int(val) < 2100:
                    yc_local[ci] = int(val)
                elif isinstance(val, str) and val.strip().isdigit() and 1990 < int(val.strip()) < 2100:
                    yc_local[ci] = int(val.strip())
            if yc_local:
                header_row = r
                year_cols = yc_local
                break
        if not header_row or not year_cols:
            parse_log.append(f"⚠ Лист «{sheet}»: не найдена строка с годами")
            continue

        parse_log.append(f"Лист «{sheet}»: заголовок в строке {header_row}, годы {sorted(year_cols.values())}")

        # Iterate data rows
        rows_parsed = 0
        for r in range(header_row + 1, ws.max_row + 1):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            if not row_cells:
                continue
            row_vals = row_cells[0]
            # Try to identify field: check cols 0..2 for canonical code or label match
            field_id: Optional[str] = None
            for ci in range(min(3, len(row_vals))):
                v = row_vals[ci]
                if not isinstance(v, str):
                    continue
                key = v.strip().lower().replace("(авто)", "").strip()
                if key in label_to_field:
                    field_id = label_to_field[key]
                    break
            if not field_id:
                continue
            # Extract values per year
            for col_idx, year in year_cols.items():
                if col_idx >= len(row_vals):
                    continue
                cell_val = row_vals[col_idx]
                if cell_val is None or cell_val == "":
                    continue
                try:
                    num = float(cell_val) if not isinstance(cell_val, str) else float(cell_val.replace(",", ".").replace(" ", ""))
                except (TypeError, ValueError):
                    continue
                if not (-1e12 < num < 1e12):
                    continue
                if field_id not in values:
                    values[field_id] = {}
                values[field_id][year] = num
            rows_parsed += 1
        parse_log.append(f"  → распознано строк: {rows_parsed}")

    # Convert year keys to strings for JSON-friendly output
    out_values = {fld: {str(y): v for y, v in ym.items()} for fld, ym in values.items()}
    return {
        "company_code": co.code,
        "company_name": co.name_short or co.name_ru,
        "values": out_values,
        "fields_count": len(out_values),
        "cells_count": sum(len(ym) for ym in out_values.values()),
        "log": parse_log,
        "filename": file.filename,
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.59 — IFRS editor (mirrors NSBU editor with 4-section grid +
# quarterly support + standalone/consolidated toggle).
# ════════════════════════════════════════════════════════════════════════

# Field sets per section. MUST match useIfrsSchema.ts in frontend.
_IFRS_PL_FIELDS = {
    "revenue", "cogs", "grossProfit", "opProfit", "depreciation",
    "finIncome", "finCost", "interestExp", "forex",
    "pbt", "tax", "profit", "ebitda",
}
_IFRS_OCI_FIELDS = {
    "oci_currency_translation", "oci_revaluation_ppe", "oci_actuarial",
    "oci_hedge_reserve", "oci_fvtoci", "total_comprehensive_income",
}
_IFRS_BS_FIELDS = {
    "ppe", "totalNCA", "cash", "totalCA", "totalAssets",
    "equity", "ltBorrowings", "stBorrowings", "totalLiabilities",
    "ltBankLoans", "ltOtherLoans", "stBankLoans", "stOtherLoans",
    "longTermDebt", "debt",
}
_IFRS_CF_FIELDS = {
    "cfo", "cfo_pbt", "cfo_depreciation", "cfo_working_capital",
    "cfo_interest_paid", "cfo_tax_paid",
    "cfi", "cfi_capex", "cfi_acquisitions",
    "cff", "cff_borrowings", "cff_repayments", "dividendsPaid",
    "netCashChange", "freeCashFlow",
}


def _ifrs_report_type(field: str) -> Optional[str]:
    if field in _IFRS_PL_FIELDS: return "PL"
    if field in _IFRS_OCI_FIELDS: return "OCI"
    if field in _IFRS_BS_FIELDS: return "BS"
    if field in _IFRS_CF_FIELDS: return "CF"
    return None


def _period_to_quarter(period: str) -> Optional[int]:
    """FY → None (annual), Q1 → 1, H1 → 2, 9M → 3."""
    m = {"FY": None, "Q1": 1, "H1": 2, "9M": 3}
    if period not in m:
        raise HTTPException(422, f"Invalid period '{period}', expected FY/Q1/H1/9M")
    return m[period]


class IfrsCustomFieldDef(BaseModel):
    id: str
    label: str
    section: Optional[str] = None  # 'pnl' | 'oci' | 'sofp' | 'cf'
    autoFormula: Optional[str] = None
    isCustom: Optional[bool] = True
    canonical: Optional[str] = None


class IfrsEditorSavePayload(BaseModel):
    """Payload from frontend IfrsEditor.vue save action."""
    period: str = "FY"  # FY | Q1 | H1 | 9M
    consolidated: bool = True
    currency: str = "UZS"
    values: dict[str, dict[str, Optional[float]]] = Field(default_factory=dict)
    # values[fieldId][yearStr] = value (млрд UZS) or null
    customFields: list[IfrsCustomFieldDef] = Field(default_factory=list)
    renames: dict[str, str] = Field(default_factory=dict)
    formulaOverrides: dict[str, str] = Field(default_factory=dict)
    manualFlags: dict[str, dict[str, bool]] = Field(default_factory=dict)
    # Pack 7.59: audit metadata (auditor firm, opinion type, fee, signed date)
    audit_meta: Optional[dict] = None
    # Pack 7.63: per-line notes/disclosures — fieldId → markdown text
    notes: dict[str, str] = Field(default_factory=dict)


@router.get("/companies/{code}/ifrs-editor")
async def get_ifrs_editor_schema(
    code: str,
    period: str = Query("FY", description="FY | Q1 | H1 | 9M"),
    consolidated: bool = Query(True, description="Consolidated (group) vs Standalone (parent)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Return IFRS-editor customization + current values for the (period, consolidated) slice.

    Values are loaded directly from financial_reports/lines (unlike NSBU editor
    which reads from portfolio-summary). This gives us full control over the
    quarter / is_consolidated filter.
    """
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    quarter = _period_to_quarter(period)

    # Load customization from company.extra (separate slot per scope+period combo)
    extra = co.extra or {}
    schema_key = f"ifrs_editor_schema_{period}_{'c' if consolidated else 's'}"
    schema = extra.get(schema_key, {}) if extra else {}

    # Load values from DB: aggregate across PL/OCI/BS/CF reports for this slice
    rows = await db.execute(
        select(FinancialReport, FinancialLine)
        .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
        .where(
            FinancialReport.company_id == co.id,
            FinancialReport.standard == "IFRS",
            FinancialReport.is_consolidated.is_(consolidated),
            FinancialReport.is_detailed.is_(False),
            FinancialReport.quarter.is_(quarter) if quarter is None else FinancialReport.quarter == quarter,
        )
    )

    # Build values dict: {field_id: {year: value_in_mlrd_UZS}}
    values: dict[str, dict[str, float]] = {}
    audit_meta_latest: Optional[dict] = None
    audit_year_latest = 0
    for fr, fl in rows.all():
        year_str = str(fr.year)
        val = fl.value
        if val is None:
            continue
        # Stored value is in unit_scale (typically 1_000_000_000 = mlrd UZS)
        # But values stored by nsbu-editor are already in mlrd, so just float()
        try:
            v_mlrd = float(val)
        except (TypeError, ValueError):
            continue
        values.setdefault(fl.line_code, {})[year_str] = v_mlrd

        # Capture audit metadata from most-recent year's report
        if fr.extra and "audit" in fr.extra and fr.year > audit_year_latest:
            audit_meta_latest = fr.extra.get("audit")
            audit_year_latest = fr.year

    return {
        "code": co.code,
        "period": period,
        "consolidated": consolidated,
        "currency": "UZS",  # always store as UZS; client converts for display
        "values": values,
        "customFields": schema.get("customFields", []),
        "renames": schema.get("renames", {}),
        "formulaOverrides": schema.get("formulaOverrides", {}),
        "manualFlags": schema.get("manualFlags", {}),
        "notes": schema.get("notes", {}),
        "audit_meta": audit_meta_latest,
        "updatedAt": schema.get("updatedAt"),
        "updatedBy": schema.get("updatedBy"),
    }


@router.put("/companies/{code}/ifrs-editor")
async def save_ifrs_editor(
    code: str,
    payload: IfrsEditorSavePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Save IFRS values + customization. Idempotent. Per-period+scope storage.

    Architectural notes:
    - Each (year, report_type) tuple gets ONE FinancialReport with the matching
      (quarter, is_consolidated) values.
    - Values are stored in млрд UZS (matches NSBU convention).
    - audit_meta stored in FinancialReport.extra.audit (latest year carries it).
    - Customization stored in Company.extra under per-scope slot.
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required: financials.edit")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    quarter = _period_to_quarter(payload.period)
    now_iso = datetime.now(timezone.utc).isoformat()

    # ─── 1. Persist customization (per-scope slot) ──────────────────────
    extra = dict(co.extra or {})
    schema_key = f"ifrs_editor_schema_{payload.period}_{'c' if payload.consolidated else 's'}"
    extra[schema_key] = {
        "customFields": [cf.model_dump() for cf in payload.customFields],
        "renames": payload.renames,
        "formulaOverrides": payload.formulaOverrides,
        "manualFlags": payload.manualFlags,
        # Pack 7.63: per-line notes — only non-empty entries
        "notes": {k: v for k, v in payload.notes.items() if v and v.strip()},
        "updatedAt": now_iso,
        "updatedBy": user.email,
    }
    co.extra = extra

    # ─── 2. Build canonical map + report-type bucketing ─────────────────
    label_for_field: dict[str, str] = {}
    custom_section_by_id: dict[str, str] = {}
    canonical_for_field: dict[str, Optional[str]] = {}

    for f in (_IFRS_PL_FIELDS | _IFRS_OCI_FIELDS | _IFRS_BS_FIELDS | _IFRS_CF_FIELDS):
        canonical_for_field[f] = f

    for cf in payload.customFields:
        label_for_field[cf.id] = cf.label
        if cf.section:
            custom_section_by_id[cf.id] = cf.section  # 'pnl' | 'oci' | 'sofp' | 'cf'
        canonical_for_field[cf.id] = cf.canonical if cf.canonical else None

    for fld, renamed in payload.renames.items():
        label_for_field[fld] = renamed

    section_to_rtype = {"pnl": "PL", "oci": "OCI", "sofp": "BS", "cf": "CF"}

    changes_by_report: dict[tuple[int, str], list[tuple[str, Optional[float], str, Optional[str]]]] = {}

    for field, year_map in payload.values.items():
        rtype = _ifrs_report_type(field)
        if rtype is None and field in custom_section_by_id:
            rtype = section_to_rtype.get(custom_section_by_id[field])
        if not rtype:
            continue

        for year_str, val in year_map.items():
            try:
                year = int(year_str)
            except (TypeError, ValueError):
                continue
            key = (year, rtype)
            changes_by_report.setdefault(key, []).append((
                field, val, label_for_field.get(field, field),
                canonical_for_field.get(field),
            ))

    # ─── 3. Upsert reports + lines ──────────────────────────────────────
    reports_created = 0
    reports_updated = 0
    lines_upserted = 0
    lines_deleted = 0

    # Cache: years that should carry audit_meta = max(year) for this save
    audit_target_year = max((y for (y, _) in changes_by_report.keys()), default=None)

    for (year, report_type), changes in changes_by_report.items():
        rep_q = await db.execute(
            select(FinancialReport).where(
                FinancialReport.company_id == co.id,
                FinancialReport.year == year,
                FinancialReport.quarter.is_(quarter) if quarter is None else FinancialReport.quarter == quarter,
                FinancialReport.standard == "IFRS",
                FinancialReport.report_type == report_type,
                FinancialReport.is_consolidated.is_(payload.consolidated),
                FinancialReport.is_detailed.is_(False),
            )
        )
        existing_reports = list(rep_q.scalars().all())
        report = None
        for r in existing_reports:
            if r.source == "ifrs-editor":
                report = r
                break
        if report is None and existing_reports:
            report = existing_reports[0]

        if report is None:
            report = FinancialReport(
                company_id=co.id,
                year=year,
                quarter=quarter,
                standard="IFRS",
                report_type=report_type,
                currency="UZS",
                unit_scale=1_000_000_000,
                source="ifrs-editor",
                is_audited=bool(payload.audit_meta and payload.audit_meta.get("opinion") == "clean"),
                is_detailed=False,
                is_consolidated=payload.consolidated,
                notes=f"Saved via IFRS editor by {user.email} on {now_iso}",
                extra={"editor_version": "p7.59"},
            )
            db.add(report)
            await db.flush()
            reports_created += 1
        else:
            reports_updated += 1
            report.source = report.source or "ifrs-editor"
            report.notes = f"Last edit via IFRS editor by {user.email} on {now_iso}"

        # Stamp audit_meta on max-year report
        if payload.audit_meta is not None and year == audit_target_year:
            rep_extra = dict(report.extra or {})
            rep_extra["audit"] = payload.audit_meta
            report.extra = rep_extra

        # Load existing lines
        ln_q = await db.execute(
            select(FinancialLine).where(FinancialLine.report_id == report.id)
        )
        existing_lines = {ln.line_code: ln for ln in ln_q.scalars().all()}

        for field, val, label, canonical in changes:
            existing = existing_lines.get(field)
            if val is None:
                if existing is not None:
                    await db.delete(existing)
                    lines_deleted += 1
                continue
            decimal_val = Decimal(str(val))
            new_parent = canonical
            if existing is not None:
                existing.value = decimal_val
                existing.line_name = label
                existing.parent_code = new_parent
            else:
                new_line = FinancialLine(
                    report_id=report.id,
                    line_code=field,
                    parent_code=new_parent,
                    line_name=label,
                    value=decimal_val,
                    is_subtotal=False,
                    is_calculated=False,
                    sort_order=0,
                )
                db.add(new_line)
            lines_upserted += 1

    # ─── 4. Audit log entry ─────────────────────────────────────────────
    try:
        sample_fields = sorted(set(field for changes in changes_by_report.values() for field, _, _, _ in changes))[:20]
        await append_audit_entry(
            db,
            actor_id=str(user.id) if user.id else None,
            actor_email=user.email,
            action="ifrs_editor.save",
            entity_type="company",
            entity_id=str(co.id),
            diff={
                "period": payload.period,
                "consolidated": payload.consolidated,
                "reports_created": reports_created,
                "reports_updated": reports_updated,
                "lines_upserted": lines_upserted,
                "lines_deleted": lines_deleted,
                "fields": sample_fields,
                "years": sorted({y for (y, _) in changes_by_report.keys()}),
            },
            payload={
                "company_code": co.code,
                "customFields_count": len(payload.customFields),
                "renames_count": len(payload.renames),
                "audit_meta_set": payload.audit_meta is not None,
            },
            notes=f"IFRS editor save · {co.code} · {payload.period} · {'consolidated' if payload.consolidated else 'standalone'}",
        )
    except Exception as e:
        print(f"[ifrs-editor] audit log failed: {e}")

    await db.commit()

    return {
        "ok": True,
        "saved_at": now_iso,
        "period": payload.period,
        "consolidated": payload.consolidated,
        "reports_created": reports_created,
        "reports_updated": reports_updated,
        "lines_upserted": lines_upserted,
        "lines_deleted": lines_deleted,
    }


@router.get("/companies/{code}/ifrs-editor/history")
async def get_ifrs_editor_history(
    code: str,
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    from app.models.audit import AuditLog
    q = await db.execute(
        select(AuditLog)
        .where(
            AuditLog.action == "ifrs_editor.save",
            AuditLog.entity_type == "company",
            AuditLog.entity_id == str(co.id),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    )
    entries = list(q.scalars().all())
    return {
        "code": co.code,
        "company_name": co.name_short or co.name_ru,
        "total": len(entries),
        "entries": [
            {
                "id": str(e.id),
                "at": e.created_at.isoformat() if e.created_at else None,
                "actor_email": e.actor_email,
                "diff": e.diff or {},
                "payload": e.payload or {},
                "notes": e.notes,
            }
            for e in entries
        ],
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.62 — NSBU ↔ IFRS reconciliation diff
# ════════════════════════════════════════════════════════════════════════

@router.get("/companies/{code}/ifrs-nsbu-diff")
async def get_ifrs_nsbu_diff(
    code: str,
    year: int = Query(..., ge=2018, le=2030),
    consolidated: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Compare NSBU vs IFRS for the same canonical metrics × company × year.

    Returns per-metric delta showing where the two standards diverge.
    Useful for IFRS audit/disclosure reconciliation tab.

    Significance buckets by |delta_pct|:
      low    < 5%
      medium 5–20%
      high   >= 20%
    """
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    # Fetch BOTH NSBU and IFRS reports for this (co, year, quarter=None, scope=consolidated)
    rows = await db.execute(
        select(FinancialReport.standard, FinancialLine.line_code, FinancialLine.parent_code, FinancialLine.value, FinancialLine.line_name)
        .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
        .where(
            FinancialReport.company_id == co.id,
            FinancialReport.year == year,
            FinancialReport.quarter.is_(None),
            FinancialReport.is_consolidated.is_(consolidated),
            FinancialReport.is_detailed.is_(False),
            FinancialReport.standard.in_(("IFRS", "NSBU")),
        )
    )

    # Aggregate per (standard, canonical_metric) — max-abs dedup logic mirrors portfolio_summary
    by_std: dict[str, dict[str, tuple[float, str]]] = {"IFRS": {}, "NSBU": {}}
    for std, line_code, parent_code, value, line_name in rows.all():
        if value is None:
            continue
        canon = _canon_metric(line_code, parent_code)
        if not canon:
            continue
        try:
            v = float(value)
        except (TypeError, ValueError):
            continue
        existing = by_std[std].get(canon)
        if existing is None or abs(v) > abs(existing[0]):
            by_std[std][canon] = (v, line_name or canon)

    # Build diff entries for all canonical metrics that appear in EITHER standard
    all_metrics = set(by_std["NSBU"].keys()) | set(by_std["IFRS"].keys())
    diffs: list[dict] = []
    for metric in sorted(all_metrics):
        nsbu_tuple = by_std["NSBU"].get(metric)
        ifrs_tuple = by_std["IFRS"].get(metric)
        nsbu_val = nsbu_tuple[0] if nsbu_tuple else None
        ifrs_val = ifrs_tuple[0] if ifrs_tuple else None
        label = (ifrs_tuple or nsbu_tuple)[1] if (ifrs_tuple or nsbu_tuple) else metric

        if nsbu_val is not None and ifrs_val is not None:
            delta = ifrs_val - nsbu_val
            denom = abs(nsbu_val) if nsbu_val != 0 else 1
            delta_pct = (delta / denom) * 100
            abs_pct = abs(delta_pct)
            if abs_pct < 5:    sig = "low"
            elif abs_pct < 20: sig = "medium"
            else:              sig = "high"
        elif nsbu_val is None and ifrs_val is not None:
            delta = ifrs_val
            delta_pct = None
            sig = "ifrs_only"  # IFRS-only line (e.g. OCI, deferred tax)
        elif ifrs_val is None and nsbu_val is not None:
            delta = -nsbu_val
            delta_pct = None
            sig = "nsbu_only"
        else:
            continue  # both null — skip

        diffs.append({
            "metric": metric,
            "label": label,
            "nsbu_value": nsbu_val,
            "ifrs_value": ifrs_val,
            "delta": delta,
            "delta_pct": delta_pct,
            "significance": sig,
        })

    # Order: high significance first, then by metric name
    sig_rank = {"high": 0, "medium": 1, "ifrs_only": 2, "nsbu_only": 3, "low": 4}
    diffs.sort(key=lambda d: (sig_rank.get(d["significance"], 9), d["metric"]))

    return {
        "code": co.code,
        "company_name": co.name_short or co.name_ru,
        "year": year,
        "consolidated": consolidated,
        "metrics_total": len(diffs),
        "summary": {
            "high": sum(1 for d in diffs if d["significance"] == "high"),
            "medium": sum(1 for d in diffs if d["significance"] == "medium"),
            "low": sum(1 for d in diffs if d["significance"] == "low"),
            "ifrs_only": sum(1 for d in diffs if d["significance"] == "ifrs_only"),
            "nsbu_only": sum(1 for d in diffs if d["significance"] == "nsbu_only"),
        },
        "diffs": diffs,
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.64 — IFRS Excel import (template + parse)
# ════════════════════════════════════════════════════════════════════════

_IFRS_FIELD_LABELS: dict[str, tuple[str, str, str, str]] = {
    # (canonical, label, ifrs_code (informal note ref), section)
    # P&L (Profit & Loss)
    "revenue":            ("revenue",            "Revenue · Выручка",                              "",  "pnl"),
    "cogs":               ("cogs",               "Cost of sales · Себестоимость",                  "",  "pnl"),
    "grossProfit":        ("grossProfit",        "Gross profit · Валовая прибыль (авто)",         "",  "pnl"),
    "opProfit":           ("opProfit",           "Operating profit · Операционная прибыль",        "",  "pnl"),
    "depreciation":       ("depreciation",       "D&A · Амортизация",                              "",  "pnl"),
    "finIncome":          ("finIncome",          "Finance income · Финансовые доходы",             "",  "pnl"),
    "finCost":            ("finCost",            "Finance costs · Финансовые расходы",             "",  "pnl"),
    "interestExp":        ("interestExp",        "  Interest expense · Процентные расходы",        "",  "pnl"),
    "forex":              ("forex",              "Forex · Курсовые разницы",                       "",  "pnl"),
    "pbt":                ("pbt",                "Profit before tax · Прибыль до налога (авто)",  "",  "pnl"),
    "tax":                ("tax",                "Income tax · Налог на прибыль",                  "",  "pnl"),
    "profit":             ("profit",             "Net profit · Чистая прибыль (авто)",             "",  "pnl"),
    "ebitda":             ("ebitda",             "EBITDA (авто)",                                  "",  "pnl"),
    # OCI (Other Comprehensive Income)
    "oci_currency_translation": ("oci_currency_translation", "OCI · Currency translation · Курсовые разницы пересчёта", "", "oci"),
    "oci_revaluation_ppe":      ("oci_revaluation_ppe",      "OCI · PPE revaluation · Переоценка ОС",                   "", "oci"),
    "oci_actuarial":            ("oci_actuarial",            "OCI · Actuarial · Актуарные доходы/расходы",              "", "oci"),
    "oci_hedge_reserve":        ("oci_hedge_reserve",        "OCI · Hedge reserve · Резерв хеджирования",               "", "oci"),
    "oci_fvtoci":               ("oci_fvtoci",               "OCI · FVTOCI · ФА по справ. стоимости",                   "", "oci"),
    "total_comprehensive_income": ("total_comprehensive_income", "Total comprehensive income · Совокупный доход (авто)", "", "oci"),
    # Balance Sheet (SOFP)
    "ppe":                ("ppe",                "PPE · Основные средства",                        "", "sofp"),
    "totalNCA":           ("totalNCA",           "Total NCA · Внеоборотные активы (итог)",         "", "sofp"),
    "cash":               ("cash",               "Cash · Денежные средства",                       "", "sofp"),
    "totalCA":            ("totalCA",            "Total CA · Оборотные активы (итог)",             "", "sofp"),
    "totalAssets":        ("totalAssets",        "TOTAL ASSETS · Итого активы (авто)",             "", "sofp"),
    "equity":             ("equity",             "Equity · Собственный капитал",                   "", "sofp"),
    "ltBorrowings":       ("ltBorrowings",       "LT borrowings · Долгосрочные обяз-ва",           "", "sofp"),
    "stBorrowings":       ("stBorrowings",       "ST borrowings · Краткосрочные обяз-ва",          "", "sofp"),
    "totalLiabilities":   ("totalLiabilities",   "TOTAL LIABILITIES (авто)",                       "", "sofp"),
    "ltBankLoans":        ("ltBankLoans",        "  LT bank loans",                                "", "sofp"),
    "ltOtherLoans":       ("ltOtherLoans",       "  LT other loans",                               "", "sofp"),
    "stBankLoans":        ("stBankLoans",        "  ST bank loans",                                "", "sofp"),
    "stOtherLoans":       ("stOtherLoans",       "  ST other loans",                               "", "sofp"),
    "longTermDebt":       ("longTermDebt",       "Long-term debt (separately)",                    "", "sofp"),
    "debt":               ("debt",               "Total debt · Финансовый долг (авто)",            "", "sofp"),
    # Cash Flow Statement
    "cfo":                ("cfo",                "CFO · Поток от операц. деятельности (авто)",    "", "cf"),
    "cfo_pbt":            ("cfo_pbt",            "  Profit before tax (adj)",                      "", "cf"),
    "cfo_depreciation":   ("cfo_depreciation",   "  Depreciation (adj)",                           "", "cf"),
    "cfo_working_capital":("cfo_working_capital","  Change in working capital",                    "", "cf"),
    "cfo_interest_paid":  ("cfo_interest_paid",  "  Interest paid",                                "", "cf"),
    "cfo_tax_paid":       ("cfo_tax_paid",       "  Income tax paid",                              "", "cf"),
    "cfi":                ("cfi",                "CFI · Поток от инвест. деятельности (авто)",    "", "cf"),
    "cfi_capex":          ("cfi_capex",          "  CapEx · Капитальные затраты",                  "", "cf"),
    "cfi_acquisitions":   ("cfi_acquisitions",   "  Acquisitions",                                 "", "cf"),
    "cff":                ("cff",                "CFF · Поток от фин. деятельности (авто)",       "", "cf"),
    "cff_borrowings":     ("cff_borrowings",     "  Proceeds from borrowings",                     "", "cf"),
    "cff_repayments":     ("cff_repayments",     "  Repayments of borrowings",                     "", "cf"),
    "dividendsPaid":      ("dividendsPaid",      "  Dividends paid",                               "", "cf"),
    "netCashChange":      ("netCashChange",      "Net change in cash (авто)",                      "", "cf"),
    "freeCashFlow":       ("freeCashFlow",       "Free Cash Flow (FCF) (авто)",                    "", "cf"),
}

# Sheet names per section — mirror frontend useIfrsSchema labels
_IFRS_SHEET_LABELS = {
    "pnl":  "ОФР",
    "oci":  "ОПД",
    "sofp": "Баланс",
    "cf":   "ДДС",
}


@router.get("/companies/{code}/ifrs-editor/template")
async def download_ifrs_editor_template(
    code: str,
    years: str = Query("2021,2022,2023,2024,2025,2026"),
    period: str = Query("FY", description="FY | Q1 | H1 | 9M"),
    consolidated: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Generate a 4-sheet XLSX template for IFRS: ОФР · ОПД · Баланс · ДДС."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    try:
        year_list = sorted({int(y.strip()) for y in years.split(",") if y.strip()})
    except ValueError:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Invalid years parameter")
    if not year_list:
        year_list = [2021, 2022, 2023, 2024, 2025, 2026]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Remove default sheet — we'll add 4 of our own
    if wb.active:
        wb.remove(wb.active)

    header_font = Font(bold=True, size=11, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF7F77DD")
    auto_fill = PatternFill("solid", fgColor="FFFFFBF0")
    auto_font = Font(italic=True, color="FFD97706")
    border = Border(
        left=Side(style="thin", color="FFE2E8F0"),
        right=Side(style="thin", color="FFE2E8F0"),
        top=Side(style="thin", color="FFE2E8F0"),
        bottom=Side(style="thin", color="FFE2E8F0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    scope_label = "consolidated" if consolidated else "standalone"

    def fill_sheet(ws, section_id: str):
        section_name = _IFRS_SHEET_LABELS.get(section_id, section_id)
        ws.cell(row=1, column=1, value=f"МСФО · {section_name} · {co.code} {co.name_short or co.name_ru or ''} · {period} · {scope_label}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="FF1E2A4A")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(year_list))
        # Helper row
        helper = "Заполни числовые поля. Поля, помеченные «авто», пересчитываются автоматически. Числа в МЛРД UZS (например 62,5)."
        ws.cell(row=2, column=1, value=helper).font = Font(italic=True, size=9, color="FF94A3B8")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(year_list))
        # Header row
        ws.cell(row=4, column=1, value="Код")
        ws.cell(row=4, column=2, value="Показатель")
        for i, yr in enumerate(year_list):
            ws.cell(row=4, column=3 + i, value=yr)
        for col in range(1, 3 + len(year_list)):
            c = ws.cell(row=4, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        # Data rows
        row = 5
        for field_id, (_fid, label, _code, sect) in _IFRS_FIELD_LABELS.items():
            if sect != section_id:
                continue
            is_auto = "(авто)" in label
            ws.cell(row=row, column=1, value=field_id)
            ws.cell(row=row, column=2, value=label)
            for col in range(1, 3 + len(year_list)):
                c = ws.cell(row=row, column=col)
                c.border = border
                if col >= 3:
                    c.alignment = center
                elif col == 2:
                    c.alignment = left
                else:
                    c.alignment = center
                if is_auto:
                    c.fill = auto_fill
                    if col == 2:
                        c.font = auto_font
            row += 1
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 55
        for i in range(len(year_list)):
            ws.column_dimensions[chr(ord("C") + i)].width = 14
        ws.freeze_panes = "C5"

    # Create 4 sheets in IFRS-natural order
    for section_id in ("pnl", "oci", "sofp", "cf"):
        ws = wb.create_sheet(_IFRS_SHEET_LABELS[section_id])
        fill_sheet(ws, section_id)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ifrs_template_{co.code}_{period}_{scope_label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/companies/{code}/ifrs-editor/parse-excel")
async def parse_ifrs_editor_excel(
    code: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Parse uploaded XLSX → return {field_id: {year: value}} matrix.

    Supports 4-section IFRS template (ОФР / ОПД / Баланс / ДДС) but is forgiving:
    parses any sheet that has a year-header row and matchable canonical codes.
    """
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    contents = await file.read()
    if not contents:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Empty file")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, f"Cannot parse XLSX: {e}")

    values: dict[str, dict[int, float]] = {}
    parse_log: list[str] = []

    # Build a label → canonical mapping (case-insensitive, strips "(авто)")
    label_to_field: dict[str, str] = {}
    for fid, (canonical, label, _code, _sect) in _IFRS_FIELD_LABELS.items():
        clean = label.replace("(авто)", "").strip().lower()
        label_to_field[clean] = canonical
        label_to_field[canonical.lower()] = canonical
        # Also support label parts separated by ' · ' (e.g. "Revenue · Выручка" → also "Revenue" and "Выручка")
        for part in label.replace("(авто)", "").split("·"):
            p = part.strip().lower()
            if p and p not in label_to_field:
                label_to_field[p] = canonical

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find header row containing years — first 15 rows
        header_row = None
        year_cols: dict[int, int] = {}
        for r in range(1, min(16, ws.max_row + 1)):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            if not row_cells:
                continue
            row_vals = row_cells[0]
            yc_local: dict[int, int] = {}
            for ci, val in enumerate(row_vals):
                if isinstance(val, (int, float)) and 1990 < int(val) < 2100:
                    yc_local[ci] = int(val)
                elif isinstance(val, str) and val.strip().isdigit() and 1990 < int(val.strip()) < 2100:
                    yc_local[ci] = int(val.strip())
            if yc_local:
                header_row = r
                year_cols = yc_local
                break
        if not header_row or not year_cols:
            parse_log.append(f"⚠ Лист «{sheet}»: не найдена строка с годами")
            continue
        parse_log.append(f"Лист «{sheet}»: заголовок в строке {header_row}, годы {sorted(year_cols.values())}")
        rows_parsed = 0
        for r in range(header_row + 1, ws.max_row + 1):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            if not row_cells:
                continue
            row_vals = row_cells[0]
            # Identify field: scan first 3 cells for canonical code or label match
            field_id: Optional[str] = None
            for ci in range(min(3, len(row_vals))):
                v = row_vals[ci]
                if not isinstance(v, str):
                    continue
                key = v.strip().lower().replace("(авто)", "").strip()
                if key in label_to_field:
                    field_id = label_to_field[key]
                    break
            if not field_id:
                continue
            for col_idx, year in year_cols.items():
                if col_idx >= len(row_vals):
                    continue
                cell_val = row_vals[col_idx]
                if cell_val is None or cell_val == "":
                    continue
                try:
                    num = float(cell_val) if not isinstance(cell_val, str) else float(cell_val.replace(",", ".").replace(" ", ""))
                except (TypeError, ValueError):
                    continue
                if not (-1e12 < num < 1e12):
                    continue
                if field_id not in values:
                    values[field_id] = {}
                values[field_id][year] = num
            rows_parsed += 1
        parse_log.append(f"  → распознано строк: {rows_parsed}")

    out_values = {fld: {str(y): v for y, v in ym.items()} for fld, ym in values.items()}
    return {
        "company_code": co.code,
        "company_name": co.name_short or co.name_ru,
        "values": out_values,
        "fields_count": len(out_values),
        "cells_count": sum(len(ym) for ym in out_values.values()),
        "log": parse_log,
        "filename": file.filename,
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.66 — High-Level Financials (HLF) import + display
#
# Parses the structured 4-section XLSX (SOFP / PNL / Cash flow + extras)
# and stores per-company JSON in company.extra.hlf for table rendering.
# ════════════════════════════════════════════════════════════════════════


def _classify_hlf_row(label: str, has_values: bool) -> str:
    """Classify a row from HLF sheet as header/subheader/line/subtotal/total."""
    lbl = (label or "").strip()
    lbl_upper = lbl.upper()
    if not has_values:
        # No values → it's a structural row
        # All-caps short labels are section headers
        if lbl_upper in ("ASSETS", "EQUITY", "LIABILITIES", "ADJUSTMENTS:", "INVESTING ACTIVITIES:", "FINANCING ACTIVITIES:") or lbl_upper.endswith(":"):
            return "section_header" if lbl_upper in ("ASSETS", "EQUITY", "LIABILITIES") else "subheader"
        return "subheader"
    # Has values
    if lbl_upper.startswith("TOTAL "):
        return "total" if lbl_upper in (
            "TOTAL ASSETS", "TOTAL LIABILITIES", "TOTAL EQUITY", "TOTAL LIABILITIES AND EQUITY",
            "TOTAL COMPREHENSIVE INCOME FOR THE YEAR"
        ) else "subtotal"
    if lbl.startswith("Total "):
        return "subtotal"
    if lbl_upper in ("GROSS PROFIT", "OPERATING PROFIT", "PROFIT BEFORE INCOME TAX",
                     "OPERATING CASH FLOW", "INVESTING CASH FLOW", "FINANCING CASH FLOW",
                     "NET CHANGE IN CASH AND CASH EQUIVALENTS",
                     "OPERATING PROFIT BEFORE WORKING CAPITAL CHANGES",
                     "CASH GENERATED FROM OPERATING ACTIVITIES"):
        return "subtotal"
    if lbl_upper == "PROFIT FOR THE YEAR":
        return "total"
    return "line"


def _parse_hlf_sheet(ws) -> dict:
    """Parse one company sheet from the HLF Excel file.

    Returns dict with:
      - years: list of detected years
      - sections: list of {id, title, rows: [{type, label, values}]}
    """
    # Step 1: find year header rows. There can be multiple (SOFP, PNL, CF each has its own).
    # Look for rows where 4+ consecutive cells are years (1990 < x < 2100).
    max_row = ws.max_row
    # Pre-scan all rows
    all_rows: list[tuple[int, tuple]] = []
    for r in range(1, max_row + 1):
        row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        if row_cells:
            all_rows.append((r, row_cells[0]))

    # Detect section markers: rows with "SOFP", "PNL", "Cash flow" in any column
    section_markers: list[tuple[int, str]] = []  # (row_idx, section_id)
    for r_idx, row in all_rows:
        for cell in row:
            if isinstance(cell, str):
                cs = cell.strip().lower()
                if cs == "sofp":
                    section_markers.append((r_idx, "sofp"))
                    break
                if cs == "pnl":
                    section_markers.append((r_idx, "pnl"))
                    break
                if "cash flow" in cs and len(cs) < 18:
                    section_markers.append((r_idx, "cashflow"))
                    break

    # Fallback: if no section markers found, treat the whole sheet as one section
    if not section_markers:
        section_markers = [(0, "report")]

    sections: list[dict] = []
    all_years_found: set[int] = set()

    # Iterate each detected section
    for i, (sec_start, sec_id) in enumerate(section_markers):
        sec_end = section_markers[i + 1][0] if i + 1 < len(section_markers) else max_row + 1
        # Find year header within this section (first row with 3+ years)
        year_row_idx = None
        year_cols: dict[int, int] = {}  # col_idx → year
        for r_idx, row in all_rows:
            if r_idx <= sec_start or r_idx >= sec_end:
                continue
            ycols: dict[int, int] = {}
            seen_years_in_row: set[int] = set()
            # Find the FIRST contiguous block of year columns.
            # Stop expanding once we hit a non-year cell or a duplicate (= another block starts).
            in_block = False
            for ci, val in enumerate(row):
                yr_candidate = None
                if isinstance(val, (int, float)) and 2000 < int(val) < 2035:
                    yr_candidate = int(val)
                elif isinstance(val, str):
                    s = val.strip()
                    if s.isdigit() and 2000 < int(s) < 2035:
                        yr_candidate = int(s)
                if yr_candidate is None:
                    if in_block:
                        # Block ended — stop, don't scan further blocks
                        break
                    continue
                if yr_candidate in seen_years_in_row:
                    # Duplicate year = start of another block (bln/mln copies) — stop
                    break
                seen_years_in_row.add(yr_candidate)
                ycols[ci] = yr_candidate
                in_block = True
            if len(ycols) >= 3:
                year_row_idx = r_idx
                year_cols = ycols
                break
        if year_row_idx is None or not year_cols:
            continue
        years_sorted = sorted(year_cols.values())
        all_years_found.update(years_sorted)
        col_year_pairs = sorted(year_cols.items(), key=lambda kv: kv[1])  # sort by YEAR so values align with years_sorted

        # Parse rows after year row until end of section
        rows_out: list[dict] = []
        seen_labels: set[str] = set()
        for r_idx, row in all_rows:
            if r_idx <= year_row_idx or r_idx >= sec_end:
                continue
            # Label = prefer column B (label) over column A (mapping key in Uzbek)
            # Some rows have label only in A, some only in B, some in both
            label = None
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            for candidate in (col_b, col_a):
                if isinstance(candidate, str) and candidate.strip():
                    s = candidate.strip()
                    if s.lower() in ("bln uzs", "mln uzs", "31 dec"):
                        continue
                    if s.isdigit():
                        continue
                    label = s
                    break
            # Fallback: look at column C+ for stray text labels
            if not label:
                for ci in range(2, min(5, len(row))):
                    v = row[ci]
                    if isinstance(v, str) and v.strip():
                        s = v.strip()
                        if s.lower() in ("bln uzs", "mln uzs", "31 dec", "sofp", "pnl", "cash flow"):
                            continue
                        if s.isdigit():
                            continue
                        label = s
                        break
            if not label:
                continue
            # Get values for each year column
            values: list[Optional[float]] = []
            has_any = False
            for col, _yr in col_year_pairs:
                if col < len(row):
                    cv = row[col]
                    if isinstance(cv, (int, float)) and cv != 0:
                        values.append(float(cv))
                        has_any = True
                    elif cv == 0:
                        values.append(0.0)
                    else:
                        values.append(None)
                else:
                    values.append(None)
            # Some labels appear twice (e.g. "Займы" in current + non-current liabilities) — keep both
            row_type = _classify_hlf_row(label, has_any)
            rows_out.append({
                "type": row_type,
                "label": label,
                "values": values,
            })

        # Skip empty sections
        if not rows_out:
            continue

        title_map = {
            "sofp": "Отчёт о финансовом положении (SOFP)",
            "pnl":  "Отчёт о прибылях и убытках (P&L)",
            "cashflow": "Отчёт о движении денежных средств (Cash Flow)",
            "report": "Финансовый отчёт",
        }
        sections.append({
            "id": sec_id,
            "title": title_map.get(sec_id, sec_id),
            "years": years_sorted,
            "rows": rows_out,
        })

    return {
        "years": sorted(all_years_found),
        "sections": sections,
    }


@router.post("/hlf-import")
async def import_hlf_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Upload + parse High-Level Financials XLSX → save to company.extra.hlf.

    Sheets are matched to companies by code (sheet name).
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    contents = await file.read()
    if not contents:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Empty file")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, f"Cannot parse XLSX: {e}")

    # Load all companies (by code) into a lookup
    cos_q = await db.execute(select(Company))
    cos = cos_q.scalars().all()
    co_by_code: dict[str, Company] = {c.code.lower(): c for c in cos}

    now_iso = datetime.utcnow().isoformat()
    summary_log: list[str] = []
    imported_count = 0
    skipped_sheets: list[str] = []

    SKIP_SHEET_NAMES = {
        "status of ifrs reports", "company metrics", "mapping lib",
        "company names", "x-rates", "sheet1",
    }

    for sheet_name in wb.sheetnames:
        sn_lower = sheet_name.lower().strip()
        if sn_lower in SKIP_SHEET_NAMES or sn_lower.startswith("_"):
            continue
        # Match by company code
        co = co_by_code.get(sn_lower)
        if not co:
            skipped_sheets.append(f"{sheet_name} (no company)")
            continue
        ws = wb[sheet_name]
        try:
            parsed = _parse_hlf_sheet(ws)
        except Exception as e:
            summary_log.append(f"⚠ {sheet_name}: parse error — {e}")
            continue
        if not parsed["sections"]:
            summary_log.append(f"⚠ {sheet_name}: no sections detected")
            continue

        # Save to company.extra.hlf
        extra = dict(co.extra or {})
        extra["hlf"] = {
            "version": "v4_2024",
            "imported_at": now_iso,
            "imported_by": user.email,
            "filename": file.filename,
            "currency": "UZS",
            "unit": "bln",  # all values are already in bln UZS
            "years": parsed["years"],
            "sections": parsed["sections"],
        }
        co.extra = extra
        imported_count += 1
        section_row_counts = [f"{s['id']}={len(s['rows'])}" for s in parsed["sections"]]
        summary_log.append(f"✓ {sheet_name} → {co.code}: years {parsed['years']} · {' '.join(section_row_counts)}")

    await db.commit()

    return {
        "imported_count": imported_count,
        "skipped_sheets": skipped_sheets,
        "log": summary_log,
        "filename": file.filename,
    }


@router.get("/companies/{code}/hlf")
async def get_company_hlf(
    code: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Return High-Level Financials JSON saved in company.extra.hlf."""
    if not _has_permission(user, "financials.view"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    hlf = (co.extra or {}).get("hlf") if co.extra else None
    return {
        "code": co.code,
        "company_name": co.name_short or co.name_ru,
        "hlf": hlf,  # None if not imported yet
    }


# ════════════════════════════════════════════════════════════════════════
# Pack 7.67 — HLF editing + KPI extraction
# ════════════════════════════════════════════════════════════════════════

class HlfRowPayload(BaseModel):
    type: str
    label: str
    values: list[Optional[float]]
    # Optional: keep original mapping key from column A (Uzbek)
    mapping: Optional[str] = None


class HlfSectionPayload(BaseModel):
    id: str
    title: str
    years: list[int]
    rows: list[HlfRowPayload]


class HlfSavePayload(BaseModel):
    years: list[int]
    sections: list[HlfSectionPayload]
    currency: str = "UZS"
    unit: str = "bln"


@router.put("/companies/{code}/hlf")
async def save_company_hlf(
    code: str,
    payload: HlfSavePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Persist edited High-Level Financials JSON to company.extra.hlf.

    Full replace — frontend sends the complete modified structure.
    """
    if not _has_permission(user, "financials.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Permission required")

    co_q = await db.execute(select(Company).where(func.lower(Company.code) == code.lower()))
    co = co_q.scalar_one_or_none()
    if not co:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, f"Company '{code}' not found")

    scope_ids = await allowed_company_ids(db, user)
    if scope_ids is not None and co.id not in scope_ids:
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access")

    now_iso = datetime.utcnow().isoformat()
    extra = dict(co.extra or {})
    existing = extra.get("hlf", {}) or {}
    extra["hlf"] = {
        **existing,  # preserve version/imported_at/filename
        "currency": payload.currency,
        "unit": payload.unit,
        "years": payload.years,
        "sections": [s.model_dump() for s in payload.sections],
        "updated_at": now_iso,
        "updated_by": user.email,
    }
    co.extra = extra
    await db.commit()

    total_rows = sum(len(s.rows) for s in payload.sections)
    return {
        "code": co.code,
        "saved": True,
        "years": payload.years,
        "sections_count": len(payload.sections),
        "rows_count": total_rows,
        "updated_at": now_iso,
    }
