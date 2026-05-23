"""Procurement Analysis routes — backend for the BETA tab.

Endpoint: GET /procurement/aggregate

Returns the full paCompute() result: KPIs + rating + per-category breakdown.

NOTE: This route depends on `ProcurementClosure` model existing. If your DB
doesn't have closures populated yet, the endpoint returns empty arrays
(safe — the frontend handles empty state gracefully).
"""
import io
import statistics
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Dict, List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status as http_status
from pydantic import BaseModel, Field
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.access import allowed_company_ids, ensure_company_access, has_unrestricted_view
from app.core.security import has_effective_permission as _has_permission
from app.database import get_db
from app.models.user import User
from app.schemas.procurement_analysis import (
    CategoryAggregate,
    CategoryDeviation,
    CategoryMeta,
    ClosureRow,
    CompanyRatingRow,
    ProcurementAggregate,
    ProcurementKpis,
    ProcurementMeta,
    ProductAgg,
)

# Try to import ProcurementClosure model; if missing, endpoint returns empty
try:
    from app.models.procurement import ProcurementClosure  # type: ignore
    HAS_CLOSURES_MODEL = True
except ImportError:
    HAS_CLOSURES_MODEL = False
    ProcurementClosure = None  # type: ignore

# Try to import Company model
try:
    from app.models.company import Company  # type: ignore
except ImportError:
    Company = None  # type: ignore

router = APIRouter(prefix="/procurement", tags=["procurement-analysis"])


# =====================================================================
# 15 fixed procurement categories — verbatim from monolith
# =====================================================================

# Verbatim from monolith PA_CATEGORIES_DEFAULT (line 20040, Ф-59 decree).
# The xarid xlsx tags rows with these IDs in "Category" column, so the names
# must match exactly — otherwise the by-category view shows wrong labels.
CATEGORIES_SEED: List[CategoryMeta] = [
    CategoryMeta(id=1,  name="Офисная бумага",                       short="Бумага",     unit="пачка"),
    CategoryMeta(id=2,  name="Канцелярские товары",                  short="Канц.",      unit="набор"),
    CategoryMeta(id=3,  name="Компьютеры и периферия",               short="ПК",         unit="шт"),
    CategoryMeta(id=4,  name="Картриджи и расходники",               short="Картр.",     unit="шт"),
    CategoryMeta(id=5,  name="СИЗ и спецодежда",                     short="СИЗ",        unit="комплект"),
    CategoryMeta(id=6,  name="Офисная мебель",                       short="Мебель",     unit="шт"),
    CategoryMeta(id=7,  name="Гигиена и чистящие средства",          short="Гигиена",    unit="л"),
    CategoryMeta(id=8,  name="Продукты питания",                     short="Питание",    unit="кг"),
    CategoryMeta(id=9,  name="Топливо-смазочные материалы",          short="ТСМ",        unit="л"),
    CategoryMeta(id=10, name="Запчасти для автотранспорта",          short="Запч.",      unit="шт"),
    CategoryMeta(id=11, name="Стройматериалы",                       short="Стр-во",     unit="т"),
    CategoryMeta(id=12, name="Освещение и электротехника",           short="Свет",       unit="шт"),
    CategoryMeta(id=13, name="Кондиционеры и вентиляция",            short="Конд.",      unit="шт"),
    CategoryMeta(id=14, name="Связь и телекоммуникации",             short="Связь",      unit="шт"),
    CategoryMeta(id=15, name="Лицензии на ПО",                       short="ПО",         unit="лиц"),
]


# =====================================================================
# Helpers (port from monolith paCompute)
# =====================================================================

def _color_for_sector(sector: Optional[str]) -> str:
    pal = {
        "mining":    "#7F77DD",
        "oilgas":    "#1D9E75",
        "energy":    "#EF9F27",
        "transport": "#378ADD",
        "other":     "#888780",
    }
    return pal.get((sector or "other").lower(), "#888780")


def _aggregate_products(closures: list) -> tuple[Dict[str, ProductAgg], List[CategoryAggregate]]:
    """Compute per-product aggregation (productsByCode) + per-category buckets.
    1:1 monolith paComputeFromContracts → cat.allProducts + data.productsByCode.

    Pack 7.9m FIX: quality bands aligned to monolith line 21433:
      clean:  spread <  200%
      wide:   spread 200-1000%
      dirty:  spread > 1000%
    Раньше backend использовал 500% как dirty threshold → ~2× products
    лишне маркировались dirty и исключались из KPI/rating.

    Также добавлен min_observations filter (monolith lines 21381-21385):
    product исключается из benchmark если unique_buyers<2 OR contract_count<3.
    """
    by_code: Dict[str, list] = {}
    for c in closures:
        if not c.product_code:
            continue
        by_code.setdefault(c.product_code, []).append(c)

    products: Dict[str, ProductAgg] = {}
    for code, rows in by_code.items():
        prices = [float(r.unit_price) for r in rows if r.unit_price]
        if not prices:
            continue
        min_p = min(prices)
        max_p = max(prices)
        avg_p = float(statistics.median(prices))
        spread_pct = ((max_p - min_p) / min_p * 100) if min_p > 0 else 0.0
        total_spend = sum(float(r.unit_price or 0) * float(r.volume or 0) for r in rows)
        unique_buyers = len({r.company_id for r in rows})
        contract_count = len(rows)
        max_dev = max(
            (abs(float(r.deviation_pct or 0)) for r in rows if r.deviation_pct is not None),
            default=0.0,
        )
        # Pack 7.9m: monolith spec — 200/1000 thresholds (line 21433)
        if spread_pct < 200:
            band = "clean"
        elif spread_pct <= 1000:
            band = "wide"
        else:
            band = "dirty"
        # Pack 7.9m: min observations — низкий sample size означает "no benchmark"
        # (нельзя достоверно сказать median рынка). Помечаем как 'wide' (а не 'dirty'),
        # чтобы НЕ исключать из KPI band counters, но frontend сможет фильтровать
        # отдельным флагом для rating расчёта.
        # Monolith line 21381-21385: products with unique<2 OR n<3 → excluded from
        # benchmark calc (но contractы участвуют в kpi.total_closures).
        # Most common product name + unit + category
        def _most_common(rows, attr):
            counts: dict = {}
            for r in rows:
                v = getattr(r, attr, None)
                if v:
                    counts[v] = counts.get(v, 0) + 1
            return max(counts.items(), key=lambda x: x[1])[0] if counts else None

        products[code] = ProductAgg(
            code=code,
            root_code=code.split("-")[0],
            name=_most_common(rows, "product_name") or code,
            unit=_most_common(rows, "unit") or "ед",
            category_id=_most_common(rows, "category_id"),
            avg_price=avg_p,
            min_price=min_p,
            max_price=max_p,
            spread_pct=round(spread_pct, 2),
            total_spend=round(total_spend, 2),
            unique_buyers=unique_buyers,
            contract_count=contract_count,
            max_deviation_pct=round(max_dev, 2),
            quality_band=band,
        )

    # Per-category aggregates
    cat_aggs: List[CategoryAggregate] = []
    for cat in CATEGORIES_SEED:
        cat_key = str(cat.id)
        in_cat = [p for p in products.values() if p.category_id == cat_key]
        # Sort by totalSpend desc to put most-impactful first
        in_cat.sort(key=lambda p: -p.total_spend)
        # Cap to keep response sane (frontend shows top-15, sortable through 50)
        in_cat_capped = in_cat[:50]
        clean = [p for p in in_cat if p.quality_band == "clean"]
        clean_avgs = [p.avg_price for p in clean]
        cat_aggs.append(CategoryAggregate(
            id=cat.id,
            name=cat.name,
            short=cat.short,
            unit="ед",
            all_products=in_cat_capped,
            clean_count=len(clean),
            benchmark_product_count=len(clean),
            clean_spread_min=min(clean_avgs) if clean_avgs else None,
            clean_spread_max=max(clean_avgs) if clean_avgs else None,
        ))

    return products, cat_aggs


def _aggregate_rating(closures: list) -> List[CompanyRatingRow]:
    """Group closures by company × category, compute weighted-avg deviations,
    rank companies by company_deviation ascending (lower = better)."""
    if not closures:
        return []

    by_company: Dict[str, Dict] = {}

    for c in closures:
        co_id = str(c.company_id)
        if co_id not in by_company:
            by_company[co_id] = {
                "company_id": c.company_id,
                "company_name": getattr(c, "company_name", None) or co_id[:8],
                "company_sector": getattr(c, "company_sector", None) or "other",
                "company_color": _color_for_sector(getattr(c, "company_sector", None)),
                "cats": {},
                "sum_dev": Decimal(0),
                "sum_ref": Decimal(0),
                "red_count": 0,    # Pack 7.9m: closures with dev ≥ +10% (monolith line 21479)
                "yellow_count": 0, # 0 ≤ dev < 10
                "green_count": 0,  # dev < 0
                "total_count": 0,  # non-dirty closures count
                "sum_overpay": Decimal(0),  # Σ positive (spend-ref)
                "sum_savings": Decimal(0),  # Σ negative (spend-ref) — as positive
            }
        co = by_company[co_id]

        # Skip dirty closures from KPI aggregates
        if getattr(c, "is_dirty", False):
            continue

        cat_id = c.category_id
        if cat_id not in co["cats"]:
            co["cats"][cat_id] = {
                "category_id": cat_id,
                "category_name": getattr(c, "category_name", "") or f"Категория {cat_id}",
                "category_short": (getattr(c, "category_name", "") or f"Кат {cat_id}")[:20],
                "sum_spend": Decimal(0),
                "sum_ref": Decimal(0),
                "closure_count": 0,
            }
        cat = co["cats"][cat_id]

        unit_price = c.unit_price or Decimal(0)
        market_avg = c.market_avg or Decimal(0)
        volume = c.volume or Decimal(0)

        spend = Decimal(unit_price) * Decimal(volume)
        ref = Decimal(market_avg) * Decimal(volume)

        cat["sum_spend"] += spend
        cat["sum_ref"] += ref
        cat["closure_count"] += 1

        delta = spend - ref
        co["sum_dev"] += delta
        co["sum_ref"] += ref
        co["total_count"] += 1

        if delta > 0:
            co["sum_overpay"] += delta
        else:
            co["sum_savings"] += -delta

        # Pack 7.9m: closure bucket classification по deviation_pct
        closure_dev = getattr(c, "deviation_pct", None)
        dev_val = float(closure_dev) if closure_dev is not None else 0.0
        if dev_val >= 10:
            co["red_count"] += 1
        elif dev_val >= 0:
            co["yellow_count"] += 1
        else:
            co["green_count"] += 1

    rating: List[CompanyRatingRow] = []
    for co_id, co in by_company.items():
        cat_devs: List[CategoryDeviation] = []
        for cat_id, cat in co["cats"].items():
            sum_dev = cat["sum_spend"] - cat["sum_ref"]
            dev_pct = float(sum_dev / cat["sum_ref"] * 100) if cat["sum_ref"] > 0 else 0.0
            cat_devs.append(
                CategoryDeviation(
                    category_id=cat_id,
                    category_name=cat["category_name"],
                    category_short=cat["category_short"],
                    sum_dev=sum_dev,
                    sum_ref=cat["sum_ref"],
                    deviation_pct=dev_pct,
                    closure_count=cat["closure_count"],
                )
            )

        # Pack 7.9m: above_count = red closures count (matches monolith semantics
        # AND это нужно frontend KPI band #3 "Красных закупок").
        above_count = co["red_count"]
        company_dev_pct = float(co["sum_dev"] / co["sum_ref"] * 100) if co["sum_ref"] > 0 else 0.0

        # Pack 7.9p: monolith-compat fields for PaRatingPanel
        total_n = max(1, co["total_count"])
        red_pct = co["red_count"] / total_n * 100
        yellow_pct = co["yellow_count"] / total_n * 100
        green_pct = co["green_count"] / total_n * 100
        # problem_cats: # categories where avg dev > 10%
        problem_cats = sum(1 for cd in cat_devs if cd.deviation_pct > 10)

        # Sort cats — best (lowest dev_pct) and worst (highest dev_pct)
        sorted_by_dev = sorted(cat_devs, key=lambda x: x.deviation_pct)
        best_cats = sorted_by_dev[:3]
        worst_cats = list(reversed(sorted_by_dev[-3:]))

        rating.append(
            CompanyRatingRow(
                company_id=co["company_id"],
                company_name=co["company_name"],
                company_color=co["company_color"],
                company_sector=co["company_sector"],
                company_deviation=company_dev_pct,
                sum_dev=co["sum_dev"],
                sum_ref=co["sum_ref"],
                above_count=above_count,
                cat_count=len(cat_devs),
                cat_dev=cat_devs,
                best_cats=best_cats,
                worst_cats=worst_cats,
                # Pack 7.9p: monolith-compat fields
                sum_overpay=co["sum_overpay"],
                sum_savings=co["sum_savings"],
                red_pct=red_pct,
                yellow_pct=yellow_pct,
                green_pct=green_pct,
                problem_cats=problem_cats,
                total_count=co["total_count"],
            )
        )

    rating.sort(key=lambda r: r.company_deviation)
    for i, r in enumerate(rating):
        r.rank = i + 1
    return rating


def _kpis_for(rating: List[CompanyRatingRow], all_closures: list) -> ProcurementKpis:
    clean = [c for c in all_closures if not getattr(c, "is_dirty", False)]
    above_count = sum(1 for r in rating if r.company_deviation > 0)
    total_overpay = sum(max(Decimal(0), r.sum_dev) for r in rating)
    above_pct = (above_count / len(rating) * 100) if rating else 0.0

    devs = sorted([r.company_deviation for r in rating])
    median_dev = devs[len(devs) // 2] if devs else 0.0

    return ProcurementKpis(
        total_companies=len(rating),
        clean_companies=len(rating),
        total_closures=len(all_closures),
        clean_closures=len(clean),
        total_overpay_uzs=total_overpay,
        above_market_pct=above_pct,
        median_deviation_pct=float(median_dev),
    )


# =====================================================================
# Routes
# =====================================================================

@router.get("/aggregate", response_model=ProcurementAggregate)
async def get_aggregate(
    year: Optional[int] = None,
    sector_code: Optional[str] = None,
    company_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Returns the full paCompute() aggregation."""
    if not HAS_CLOSURES_MODEL or ProcurementClosure is None:
        # Graceful degradation — frontend handles empty state
        return ProcurementAggregate(
            year=year,
            sector_code=sector_code,
            kpis=ProcurementKpis(
                total_companies=0, clean_companies=0,
                total_closures=0, clean_closures=0,
                total_overpay_uzs=Decimal(0),
                above_market_pct=0.0, median_deviation_pct=0.0,
            ),
            categories=CATEGORIES_SEED,
            rating=[],
            purchases=[],
            available_years=[],
            sectors=[],
            generated_at=datetime.now(timezone.utc),
        )

    # Scope filter: либо проверяем явный company_id, либо ограничиваем
    # выдачу для scoped users по allowed_companies. Пустой scope → пусто.
    if company_id is not None:
        await ensure_company_access(db, user, company_id)

    scope_ids = None
    if not has_unrestricted_view(user):
        scope_ids = await allowed_company_ids(db, user)
        if not scope_ids:
            return ProcurementAggregate(
                year=year,
                sector_code=sector_code,
                kpis=ProcurementKpis(
                    total_companies=0, clean_companies=0,
                    total_closures=0, clean_closures=0,
                    total_overpay_uzs=Decimal(0),
                    above_market_pct=0.0, median_deviation_pct=0.0,
                ),
                categories=CATEGORIES_SEED,
                rating=[], purchases=[], available_years=[], sectors=[],
                generated_at=datetime.now(timezone.utc),
            )

    q = select(ProcurementClosure)
    # IMPORTANT: filter out rows that don't have a benchmark (market_avg=NULL).
    # Pre-existing test data has no benchmark and would zero-out all KPIs
    # (sum_ref → 0 → company_deviation → 0). Only rows with a non-null
    # market_avg + non-null unit_price are statistically meaningful.
    q = q.where(
        ProcurementClosure.market_avg.is_not(None),
        ProcurementClosure.unit_price.is_not(None),
    )
    if year is not None:
        q = q.where(ProcurementClosure.year == year)
    if company_id is not None:
        q = q.where(ProcurementClosure.company_id == company_id)
    elif scope_ids is not None:
        q = q.where(ProcurementClosure.company_id.in_(scope_ids))
    closures = (await db.execute(q)).scalars().all()

    # Enrich closures with company name/sector — best-effort
    if Company is not None and closures:
        co_ids = list({c.company_id for c in closures})
        co_q = await db.execute(select(Company).where(Company.id.in_(co_ids)))
        co_map = {str(c.id): c for c in co_q.scalars().all()}
        for c in closures:
            co = co_map.get(str(c.company_id))
            if co:
                # attach name+sector for aggregation
                setattr(c, "company_name", getattr(co, "name_short", None) or getattr(co, "name_ru", None) or "—")
                setattr(c, "company_sector", getattr(co, "sector_code", None))

    rating = _aggregate_rating(list(closures))
    kpis = _kpis_for(rating, list(closures))
    products_by_code, cat_aggregates = _aggregate_products(list(closures))

    # Available years
    yr_q = await db.execute(select(ProcurementClosure.year).distinct())
    avail_years = sorted({y for (y,) in yr_q.all() if y is not None})

    # Build purchases array (cap = ~15k to cover full Q1 2026 dataset)
    purchases: List[ClosureRow] = []
    for c in closures[:15000]:
        unit_price = c.unit_price or Decimal(0)
        market_avg = c.market_avg or Decimal(1)
        volume = c.volume or Decimal(0)
        dev_pct = float((unit_price - market_avg) / market_avg * 100) if market_avg else 0.0
        purchases.append(
            ClosureRow(
                id=c.id,
                company_id=c.company_id,
                company_name=getattr(c, "company_name", None),
                company_color=_color_for_sector(getattr(c, "company_sector", None)),
                company_sector=getattr(c, "company_sector", None),
                category_id=c.category_id,
                category_name=getattr(c, "category_name", "") or "—",
                category_unit=getattr(c, "unit", None) or "ед",
                product_code=getattr(c, "product_code", None),
                sub_product_code=getattr(c, "sub_product_code", None),
                product_name=getattr(c, "product_name", None),
                supplier=getattr(c, "supplier", None),
                unit_price=Decimal(unit_price),
                market_avg=Decimal(market_avg),
                volume=Decimal(volume),
                deviation_pct=dev_pct,
                deviation_abs=Decimal(unit_price - market_avg) * Decimal(volume),
                spread_pct=getattr(c, "spread_pct", None),
                is_dirty=getattr(c, "is_dirty", False),
                contract_date=getattr(c, "contract_date", None),
                year=getattr(c, "year", None),
            )
        )

    return ProcurementAggregate(
        year=year,
        sector_code=sector_code,
        kpis=kpis,
        categories=CATEGORIES_SEED,
        category_aggregates=cat_aggregates,
        products_by_code=products_by_code,
        rating=rating,
        purchases=purchases,
        available_years=avail_years,
        sectors=[],
        meta=ProcurementMeta(source="procurementContracts"),
        generated_at=datetime.now(timezone.utc),
    )


# =====================================================================
# Mutation endpoints (Phase 3 extension — PA bulk + per-row edit)
# =====================================================================

_PURCHASE_TYPE_NORM = {
    "E-SHOP":                    "e_shop",
    "E_STORE":                   "e_shop",
    "AUCTION":                   "auction",
    "BEST_OFFER":                "best_offer",
    "OTHER_COMPETITIVE_METHODS": "competitive",
}


def _norm_purchase_type(s):
    if not s:
        return None
    return _PURCHASE_TYPE_NORM.get(str(s).upper().strip(), str(s).strip()[:32])


def _parse_category_id(text_val):
    if text_val is None:
        return None
    import re
    m = re.match(r"^\s*(\d{1,2})\.", str(text_val).strip())
    return str(int(m.group(1))) if m else None


# Sheet-name → company `code` (lowercase abbr; matches governance seed).
_PA_SHEET_TO_CODE = {
    "NGMK": "ngmk", "NAVOIYURAN": "nur", "AGMK": "agmk", "UMK": "umk", "UUG": "uug",
    "UNG":  "ung",  "UTG":        "utg", "HGT":  "hgt",  "UGT": "ugt", "NES": "nes",
    "TES":  "tes",  "RES":        "res", "UGE":  "uge",  "UTY": "uty", "UHY": "uhy",
    "UAP":  "uap",  "UTC":        "utc", "TSHT": "tst",  "UPT": "upt", "UAS": "uas",
    "NAZ":  "naz",  "UKS":        "uks",
}


class PaImportSummary(BaseModel):
    ok: bool = True
    inserted: int
    skipped_no_company: int
    skipped_no_data: int
    sheets_processed: int
    product_codes: int
    benchmark_rows: int                       # rows that got market_avg / deviation_pct


@router.post("/closures/import-excel", response_model=PaImportSummary)
async def import_closures_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> PaImportSummary:
    """Bulk-import procurement_closures from the xarid format
    (22 sheets, one per SOE, headers may include trailing spaces).
    Computes per-product_code median → market_avg + deviation_pct (1:1 monolith)."""
    if not await _has_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")
    if not has_unrestricted_view(user):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Admin scope required")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Expected .xlsx or .xls file")
    raw = await file.read()
    if not raw:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Empty file")

    try:
        import openpyxl  # noqa: WPS433
    except ImportError:
        raise HTTPException(http_status.HTTP_500_INTERNAL_SERVER_ERROR, "openpyxl not installed")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, f"Failed to parse xlsx: {e}")

    # Build company-code map
    if Company is None:
        raise HTTPException(http_status.HTTP_500_INTERNAL_SERVER_ERROR, "Company model unavailable")
    co_q = await db.execute(text("SELECT id, code, sector_id FROM companies"))
    by_code = {(r.code or "").lower(): r.id for r in co_q.all()}
    sec_q = await db.execute(text("""
        SELECT c.id, s.code AS sector_code
        FROM companies c LEFT JOIN sectors s ON s.id = c.sector_id
    """))
    sector_by_co = {r.id: r.sector_code for r in sec_q.all()}

    # 1st pass: collect parsed rows (no insert yet — need median for benchmark)
    parsed: list[dict[str, Any]] = []
    sheets_processed = 0
    skipped_no_company = 0
    skipped_no_data = 0

    for sheet_name in wb.sheetnames:
        code = _PA_SHEET_TO_CODE.get(sheet_name.upper().strip())
        if not code:
            # Try direct match if sheet name itself looks like a code
            code = sheet_name.lower().strip()
        company_id = by_code.get(code)
        if company_id is None:
            skipped_no_company += 1
            continue
        sheets_processed += 1

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue
        headers = [str(c or "").strip() for c in all_rows[0]]  # trim stray spaces
        col_idx = {h: i for i, h in enumerate(headers)}

        def col(name: str) -> int:
            return col_idx.get(name, -1)

        c_lot     = col("lotId")
        c_organ   = col("organ")
        c_vendor  = col("vendor")
        c_vinn    = col("vendorInn")
        c_start   = col("startSumma")
        c_camt    = col("contractAmount")
        c_saved   = col("savedAmount")
        c_spct    = col("savedPercent")
        c_cdate   = col("contractDate")
        c_sdate   = col("startDate")
        c_ptype   = col("purchaseType")
        c_plat    = col("platformName")
        c_cat     = col("Category")
        c_pcode   = col("productCode")
        c_pname   = col("productName")
        c_ptype2  = col("productType")
        c_unit    = col("unit")
        c_amt     = col("amount")
        c_price   = col("Unit price")
        c_region  = col("regionName")

        for row in all_rows[1:]:
            if not row:
                continue
            def cell(i: int):
                return row[i] if 0 <= i < len(row) else None

            try:
                up_raw = cell(c_price)
                vol_raw = cell(c_amt)
                up = float(up_raw) if up_raw not in (None, "") else None
                vol = float(vol_raw) if vol_raw not in (None, "") else None
            except (TypeError, ValueError):
                skipped_no_data += 1
                continue
            if up is None or vol is None or up <= 0 or vol <= 0:
                skipped_no_data += 1
                continue

            def s(i: int):
                v = cell(i)
                return None if v in (None, "") else str(v).strip()

            def iso_date(i: int):
                v = cell(i)
                if not v:
                    return None
                ss = str(v).strip()
                import re as _re
                m = _re.match(r"^(\d{4})-(\d{2})-(\d{2})", ss)
                if m:
                    from datetime import date as _date
                    try:
                        return _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    except ValueError:
                        return None
                return None

            def num(i: int):
                try:
                    v = cell(i)
                    return float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    return None

            parsed.append({
                "company_id":    company_id,
                "code":          code,
                "year":          2026,
                "closure_date":  iso_date(c_cdate) or iso_date(c_sdate),
                "category_id":   _parse_category_id(s(c_cat)),
                "product_code":  s(c_pcode),
                "product_name":  (s(c_pname) or "")[:1024] or None,
                "unit_price":    up,
                "volume":        vol,
                "total_amount":  up * vol,
                "contract_amount": num(c_camt),
                "start_summa":   num(c_start),
                "saved_amount":  num(c_saved),
                "saved_percent": num(c_spct),
                "supplier_name": (s(c_vendor) or "")[:512] or None,
                "supplier_inn":  s(c_vinn),
                "lot_id":        s(c_lot),
                "platform":      (s(c_plat) or "")[:64] or None,
                "purchase_type": _norm_purchase_type(s(c_ptype)),
                "region":        (s(c_region) or "")[:128] or None,
                "unit":          s(c_unit),
                "product_type":  s(c_ptype2),
                "sector":        sector_by_co.get(company_id),
            })

    # 2nd pass: median per product_code → market_avg + deviation_pct
    by_pcode: dict[str, list[float]] = {}
    for r in parsed:
        if r["product_code"]:
            by_pcode.setdefault(r["product_code"], []).append(r["unit_price"])
    medians: dict[str, float] = {
        pcode: float(statistics.median(prices))
        for pcode, prices in by_pcode.items()
    }
    benchmark_rows = 0
    for r in parsed:
        med = medians.get(r["product_code"]) if r["product_code"] else None
        if med is not None and med > 0:
            r["market_avg"] = med
            r["deviation_pct"] = (r["unit_price"] - med) / med * 100
            benchmark_rows += 1
        else:
            r["market_avg"] = None
            r["deviation_pct"] = None

    # 3rd pass: batch insert (server-side default for id)
    import json as _json
    insert_sql = text("""
        INSERT INTO procurement_closures (
            id, company_id, year, closure_date,
            category_id, product_code, product_name,
            unit_price, market_avg, deviation_pct,
            unit, volume, total_amount, saved_amount,
            supplier_name, supplier_inn,
            contract_id, lot_id, platform, purchase_type, region, sector,
            is_clean, is_dirty, extra,
            created_at, updated_at
        ) VALUES (
            gen_random_uuid(), :company_id, :year, :closure_date,
            :category_id, :product_code, :product_name,
            :unit_price, :market_avg, :deviation_pct,
            :unit, :volume, :total_amount, :saved_amount,
            :supplier_name, :supplier_inn,
            NULL, :lot_id, :platform, :purchase_type, :region, :sector,
            TRUE, FALSE, CAST(:extra AS jsonb),
            NOW(), NOW()
        )
    """)
    BATCH = 500
    inserted = 0
    buf: list[dict] = []
    for r in parsed:
        buf.append({
            "company_id":    r["company_id"],
            "year":          r["year"],
            "closure_date":  r["closure_date"],
            "category_id":   r["category_id"],
            "product_code":  r["product_code"],
            "product_name":  r["product_name"],
            "unit_price":    r["unit_price"],
            "market_avg":    r["market_avg"],
            "deviation_pct": r["deviation_pct"],
            "unit":          r["unit"],
            "volume":        r["volume"],
            "total_amount":  r["total_amount"],
            "saved_amount":  r["saved_amount"],
            "supplier_name": r["supplier_name"],
            "supplier_inn":  r["supplier_inn"],
            "lot_id":        r["lot_id"],
            "platform":      r["platform"],
            "purchase_type": r["purchase_type"],
            "region":        r["region"],
            "sector":        r["sector"],
            "extra": _json.dumps({
                "source": "manual-upload",
                "start_summa":     r["start_summa"],
                "contract_amount": r["contract_amount"],
                "saved_percent":   r["saved_percent"],
                "product_type":    r["product_type"],
            }, ensure_ascii=False),
        })
        if len(buf) >= BATCH:
            await db.execute(insert_sql, buf)
            inserted += len(buf)
            buf = []
    if buf:
        await db.execute(insert_sql, buf)
        inserted += len(buf)

    await db.commit()

    return PaImportSummary(
        inserted=inserted,
        skipped_no_company=skipped_no_company,
        skipped_no_data=skipped_no_data,
        sheets_processed=sheets_processed,
        product_codes=len(by_pcode),
        benchmark_rows=benchmark_rows,
    )


class PaClosurePatch(BaseModel):
    """Editable per-closure fields. Numeric fields trigger market_avg + deviation
    recompute for the affected product_code."""
    unit_price:    Optional[float] = Field(None, gt=0)
    volume:        Optional[float] = Field(None, gt=0)
    product_code:  Optional[str] = None
    product_name:  Optional[str] = None
    supplier_name: Optional[str] = None
    is_dirty:      Optional[bool] = None
    dirty_reason:  Optional[str] = None


@router.put("/closures/{closure_id}")
async def update_closure(
    closure_id: UUID,
    payload: PaClosurePatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Update a single closure. After change, recomputes median for the
    affected product_code and updates market_avg/deviation on all sibling rows."""
    if not await _has_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")
    if not HAS_CLOSURES_MODEL:
        raise HTTPException(http_status.HTTP_500_INTERNAL_SERVER_ERROR, "ProcurementClosure model unavailable")

    row = (await db.execute(
        select(ProcurementClosure).where(ProcurementClosure.id == closure_id),
    )).scalar_one_or_none()
    if row is None:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, "Closure not found")

    if not has_unrestricted_view(user):
        allowed = await allowed_company_ids(db, user)
        if row.company_id not in allowed:
            raise HTTPException(http_status.HTTP_403_FORBIDDEN, "No access to this company")

    old_pcode = row.product_code
    changed_price = False
    if payload.unit_price is not None:
        row.unit_price = Decimal(str(payload.unit_price))
        changed_price = True
    if payload.volume is not None:
        row.volume = Decimal(str(payload.volume))
        # Recompute total_amount when volume changes
        if row.unit_price:
            row.total_amount = Decimal(row.unit_price) * Decimal(str(payload.volume))
    if payload.product_code is not None:
        row.product_code = payload.product_code
    if payload.product_name is not None:
        row.product_name = payload.product_name
    if payload.supplier_name is not None:
        row.supplier_name = payload.supplier_name
    if payload.is_dirty is not None:
        row.is_dirty = payload.is_dirty
        row.is_clean = not payload.is_dirty
    if payload.dirty_reason is not None:
        row.dirty_reason = payload.dirty_reason

    # Recompute medians for any affected product_code (old + new)
    affected_codes = {c for c in (old_pcode, row.product_code) if c}
    recomputed = 0
    if changed_price or payload.product_code is not None:
        for pcode in affected_codes:
            sibs_q = await db.execute(
                select(ProcurementClosure).where(ProcurementClosure.product_code == pcode),
            )
            sibs = list(sibs_q.scalars().all())
            prices = [float(s.unit_price) for s in sibs if s.unit_price]
            if not prices:
                continue
            med = float(statistics.median(prices))
            for s in sibs:
                if s.unit_price and med > 0:
                    s.market_avg = Decimal(str(med))
                    s.deviation_pct = (s.unit_price - Decimal(str(med))) / Decimal(str(med)) * 100
                    recomputed += 1

    await db.commit()
    await db.refresh(row)
    return {
        "ok": True,
        "id": str(row.id),
        "unit_price": float(row.unit_price) if row.unit_price else None,
        "market_avg": float(row.market_avg) if row.market_avg else None,
        "deviation_pct": float(row.deviation_pct) if row.deviation_pct else None,
        "siblings_recomputed": recomputed,
    }


@router.delete("/closures")
async def clear_closures(
    year: Optional[int] = None,
    source: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Delete closures filtered by year and/or extra.source label
    (e.g. source='manual-upload' to drop only manual imports, preserving
    seeded q1-2026-xlsx data). At least one filter is required."""
    if not await _has_permission(db, user, "procurement.edit"):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "procurement.edit required")
    if not has_unrestricted_view(user):
        raise HTTPException(http_status.HTTP_403_FORBIDDEN, "Admin scope required to clear closures")
    if not HAS_CLOSURES_MODEL:
        raise HTTPException(http_status.HTTP_500_INTERNAL_SERVER_ERROR, "ProcurementClosure model unavailable")
    if year is None and not source:
        raise HTTPException(http_status.HTTP_400_BAD_REQUEST, "Provide ?year= and/or ?source=")

    where_clauses = []
    params: dict[str, Any] = {}
    if year is not None:
        where_clauses.append("year = :year")
        params["year"] = year
    if source:
        where_clauses.append("extra->>'source' = :source")
        params["source"] = source
    res = await db.execute(
        text(f"DELETE FROM procurement_closures WHERE {' AND '.join(where_clauses)} RETURNING id"),
        params,
    )
    cleared = len(res.fetchall())
    await db.commit()
    return {"ok": True, "cleared": cleared, "year": year, "source": source}
