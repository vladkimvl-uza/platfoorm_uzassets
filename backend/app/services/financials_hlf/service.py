"""High-Level Financials use-cases (Pack 7.66/7.67).

Imports the structured 4-section XLSX (SOFP / PNL / Cash flow), parses
into normalized JSON, and persists per-company in `company.extra.hlf`.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi import status as http_status
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.access import allowed_company_ids
from app.core.editor_lock import check_editor_token, token_from_isos
from app.core.security import has_effective_permission
from app.models.user import User
from app.repositories.financials_repository import FinancialsRepository

_SKIP_SHEET_NAMES = {
    "status of ifrs reports", "company metrics", "mapping lib",
    "company names", "x-rates", "sheet1",
}


# ─── Payload schemas ──────────────────────────────────────────────

class HlfRowPayload(BaseModel):
    type: str
    label: str
    values: list[Optional[float]]
    mapping: Optional[str] = None
    # Автосумма для итоговых строк (total/subtotal): значения = сумма строк-line
    # в области. None = эвристика по умолчанию (балансовые «Total ...» — авто).
    auto: Optional[bool] = None


class HlfSectionPayload(BaseModel):
    id: str
    title: str
    years: list[int]
    rows: list[HlfRowPayload]


class HlfSavePayload(BaseModel):
    years: list[int]
    sections: list[HlfSectionPayload]
    currency: str = "UZS"
    unit: str = "bln"


# ─── Parser helpers ───────────────────────────────────────────────

def _classify_hlf_row(label: str, has_values: bool) -> str:
    """Classify row as header/subheader/line/subtotal/total."""
    lbl = (label or "").strip()
    lbl_upper = lbl.upper()
    if not has_values:
        if lbl_upper in (
            "ASSETS", "EQUITY", "LIABILITIES", "ADJUSTMENTS:",
            "INVESTING ACTIVITIES:", "FINANCING ACTIVITIES:",
        ) or lbl_upper.endswith(":"):
            return (
                "section_header"
                if lbl_upper in ("ASSETS", "EQUITY", "LIABILITIES")
                else "subheader"
            )
        return "subheader"
    if lbl_upper.startswith("TOTAL "):
        return "total" if lbl_upper in (
            "TOTAL ASSETS", "TOTAL LIABILITIES", "TOTAL EQUITY",
            "TOTAL LIABILITIES AND EQUITY",
            "TOTAL COMPREHENSIVE INCOME FOR THE YEAR",
        ) else "subtotal"
    if lbl.startswith("Total "):
        return "subtotal"
    if lbl_upper in (
        "GROSS PROFIT", "OPERATING PROFIT", "PROFIT BEFORE INCOME TAX",
        "OPERATING CASH FLOW", "INVESTING CASH FLOW", "FINANCING CASH FLOW",
        "NET CHANGE IN CASH AND CASH EQUIVALENTS",
        "OPERATING PROFIT BEFORE WORKING CAPITAL CHANGES",
        "CASH GENERATED FROM OPERATING ACTIVITIES",
    ):
        return "subtotal"
    if lbl_upper == "PROFIT FOR THE YEAR":
        return "total"
    return "line"


_SECTION_TITLES = {
    "sofp":     "Отчёт о финансовом положении (SOFP)",
    "pnl":      "Отчёт о прибылях и убытках (P&L)",
    "cashflow": "Отчёт о движении денежных средств (Cash Flow)",
    "report":   "Финансовый отчёт",
}


def _parse_hlf_sheet(ws) -> dict:
    """Parse one company sheet from the HLF Excel file.

    Returns: {years, sections: [{id, title, years, rows: [{type, label, values}]}]}
    """
    max_row = ws.max_row
    all_rows: list[tuple[int, tuple]] = []
    for r in range(1, max_row + 1):
        row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        if row_cells:
            all_rows.append((r, row_cells[0]))

    # Detect section markers
    section_markers: list[tuple[int, str]] = []
    for r_idx, row in all_rows:
        for cell in row:
            if isinstance(cell, str):
                cs = cell.strip().lower()
                if cs == "sofp":
                    section_markers.append((r_idx, "sofp"))
                    break
                if cs == "pnl":
                    section_markers.append((r_idx, "pnl"))
                    break
                if "cash flow" in cs and len(cs) < 18:
                    section_markers.append((r_idx, "cashflow"))
                    break
    if not section_markers:
        section_markers = [(0, "report")]

    sections: list[dict] = []
    all_years_found: set[int] = set()

    for i, (sec_start, sec_id) in enumerate(section_markers):
        sec_end = (
            section_markers[i + 1][0]
            if i + 1 < len(section_markers) else max_row + 1
        )
        year_row_idx = None
        year_cols: dict[int, int] = {}
        for r_idx, row in all_rows:
            if r_idx <= sec_start or r_idx >= sec_end:
                continue
            ycols: dict[int, int] = {}
            seen_years_in_row: set[int] = set()
            in_block = False
            for ci, val in enumerate(row):
                yr = None
                if isinstance(val, int | float) and 2000 < int(val) < 2035:
                    yr = int(val)
                elif isinstance(val, str):
                    s = val.strip()
                    if s.isdigit() and 2000 < int(s) < 2035:
                        yr = int(s)
                if yr is None:
                    if in_block:
                        break
                    continue
                if yr in seen_years_in_row:
                    break
                seen_years_in_row.add(yr)
                ycols[ci] = yr
                in_block = True
            if len(ycols) >= 3:
                year_row_idx = r_idx
                year_cols = ycols
                break
        if year_row_idx is None or not year_cols:
            continue
        years_sorted = sorted(year_cols.values())
        all_years_found.update(years_sorted)
        col_year_pairs = sorted(year_cols.items(), key=lambda kv: kv[1])

        rows_out: list[dict] = []
        for r_idx, row in all_rows:
            if r_idx <= year_row_idx or r_idx >= sec_end:
                continue
            label = None
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            for candidate in (col_b, col_a):
                if isinstance(candidate, str) and candidate.strip():
                    s = candidate.strip()
                    if s.lower() in ("bln uzs", "mln uzs", "31 dec"):
                        continue
                    if s.isdigit():
                        continue
                    label = s
                    break
            if not label:
                for ci in range(2, min(5, len(row))):
                    v = row[ci]
                    if isinstance(v, str) and v.strip():
                        s = v.strip()
                        if s.lower() in (
                            "bln uzs", "mln uzs", "31 dec",
                            "sofp", "pnl", "cash flow",
                        ):
                            continue
                        if s.isdigit():
                            continue
                        label = s
                        break
            if not label:
                continue
            values: list[Optional[float]] = []
            has_any = False
            for col, _yr in col_year_pairs:
                if col < len(row):
                    cv = row[col]
                    if isinstance(cv, int | float) and cv != 0:
                        values.append(float(cv))
                        has_any = True
                    elif cv == 0:
                        values.append(0.0)
                    else:
                        values.append(None)
                else:
                    values.append(None)
            rows_out.append({
                "type": _classify_hlf_row(label, has_any),
                "label": label,
                "values": values,
            })
        if not rows_out:
            continue
        sections.append({
            "id": sec_id,
            "title": _SECTION_TITLES.get(sec_id, sec_id),
            "years": years_sorted,
            "rows": rows_out,
        })

    return {
        "years": sorted(all_years_found),
        "sections": sections,
    }


# ─── Shared write-core (LIVE + APPLY) ─────────────────────────────
# The HLF store is a FULL-BLOB last-writer-wins replace of company.extra['hlf'].
# Both the live PUT path (save_company_hlf) and the moderation apply path
# (apply_submission) funnel through this ONE function so there is no drift.
#
# Optimistic-lock re-check lives HERE (not only on the route): on the live path
# `expected_token` is the author's If-Match; on the apply path it is the token
# captured at submit time and carried in proposed_value. Either way we refuse to
# clobber intervening edits (mirror of save_report.expected_prev_checksum, which
# is likewise re-checked inside the apply handler). No-op when the token is absent
# (legacy clients) — see check_editor_token.

async def _apply_hlf_core(
    db: AsyncSession,
    *,
    company,
    payload: "HlfSavePayload",
    user: User,
    expected_token: Optional[str] = None,
) -> dict:
    now_iso = datetime.utcnow().isoformat()
    extra = dict(company.extra or {})
    existing = extra.get("hlf", {}) or {}

    check_editor_token(
        scope_name=f"financials/hlf/{company.code}",
        expected_token=expected_token,
        current_token=token_from_isos(
            existing.get("updated_at"), existing.get("imported_at"),
        ),
    )

    extra["hlf"] = {
        **existing,
        "currency": payload.currency,
        "unit": payload.unit,
        "years": payload.years,
        "sections": [s.model_dump() for s in payload.sections],
        "updated_at": now_iso,
        "updated_by": user.email,
    }
    company.extra = extra
    await db.commit()

    total_rows = sum(len(s.rows) for s in payload.sections)
    return {
        "code": company.code,
        "saved": True,
        "years": payload.years,
        "sections_count": len(payload.sections),
        "rows_count": total_rows,
        "updated_at": now_iso,
        # Re-issue (matches a subsequent GET: updated_at=now_iso + preserved
        # imported_at via {**existing}) — same editor saves again w/o reload.
        "_editor_token": token_from_isos(now_iso, existing.get("imported_at")),
    }


# ─── Service ──────────────────────────────────────────────────────

@dataclass
class FinancialsHlfService:
    async def import_file(
        self,
        file: UploadFile,
        db: AsyncSession,
        user: User,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.edit"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        contents = await file.read()
        if not contents:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST, "Empty file",
            )
        from openpyxl import load_workbook
        try:
            wb = load_workbook(
                io.BytesIO(contents), data_only=True, read_only=True,
            )
        except Exception as e:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                f"Cannot parse XLSX: {e}",
            )

        repo = FinancialsRepository(db)
        cos = await repo.list_all_companies()
        co_by_code = {c.code.lower(): c for c in cos}
        # P0 (аудит фин-источников): per-company scope. Раньше импорт проверял
        # только financials.edit и писал во ВСЕ компании, чьи коды совпали с
        # листами — company-scoped пользователь мог перезаписать HLF чужих
        # компаний. Теперь чужие листы пропускаются (как в PUT/GET).
        from app.core.access import allowed_company_ids
        scope_ids = await allowed_company_ids(db, user)

        now_iso = datetime.utcnow().isoformat()
        summary_log: list[str] = []
        imported_count = 0
        skipped_sheets: list[str] = []

        for sheet_name in wb.sheetnames:
            sn_lower = sheet_name.lower().strip()
            if sn_lower in _SKIP_SHEET_NAMES or sn_lower.startswith("_"):
                continue
            co = co_by_code.get(sn_lower)
            if not co:
                skipped_sheets.append(f"{sheet_name} (no company)")
                continue
            if scope_ids is not None and co.id not in scope_ids:
                skipped_sheets.append(f"{sheet_name} (нет доступа к компании)")
                continue
            ws = wb[sheet_name]
            try:
                parsed = _parse_hlf_sheet(ws)
            except Exception as e:
                summary_log.append(f"⚠ {sheet_name}: parse error — {e}")
                continue
            if not parsed["sections"]:
                summary_log.append(f"⚠ {sheet_name}: no sections detected")
                continue
            extra = dict(co.extra or {})
            extra["hlf"] = {
                "version": "v4_2024",
                "imported_at": now_iso,
                "imported_by": user.email,
                "filename": file.filename,
                "currency": "UZS",
                "unit": "bln",
                "years": parsed["years"],
                "sections": parsed["sections"],
            }
            co.extra = extra
            imported_count += 1
            sec_counts = [
                f"{s['id']}={len(s['rows'])}" for s in parsed["sections"]
            ]
            summary_log.append(
                f"✓ {sheet_name} → {co.code}: "
                f"years {parsed['years']} · {' '.join(sec_counts)}"
            )

        await db.commit()
        return {
            "imported_count": imported_count,
            "skipped_sheets": skipped_sheets,
            "log": summary_log,
            "filename": file.filename,
        }

    async def get_company_hlf(
        self, code: str, db: AsyncSession, user: User,
    ) -> dict:
        if not await has_effective_permission(db, user, "financials.view"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )
        hlf = (co.extra or {}).get("hlf") if co.extra else None
        return {
            "code": co.code,
            "company_name": co.name_short or co.name_ru,
            "hlf": hlf,
            # Токен свёртывает updated_at (ручное сохранение) И imported_at (реимпорт)
            # — чтобы реимпорт тоже двигал токен и не затирался открытым редактором.
            "_editor_token": token_from_isos(
                (hlf or {}).get("updated_at"), (hlf or {}).get("imported_at"),
            ),
        }

    async def save_company_hlf(
        self,
        code: str,
        payload: HlfSavePayload,
        db: AsyncSession,
        user: User,
        *,
        expected_token: Optional[str] = None,
    ) -> tuple[Optional[dict], Optional[dict]]:
        """Returns either (result, None) for the normal write path, or
        (None, queued_dict) if the moderation gate held the change (HTTP 202).

        Moderation is deny-by-default and OFF in prod until an admin enables the
        'financials' module in the panel — so this is pure wiring, no behavior
        change today (owner/bypass/no-rule all take the (result, None) branch).
        """
        if not await has_effective_permission(db, user, "financials.edit"):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        # Moderation gate — AFTER permission + per-company scope, BEFORE any DB
        # mutation. The editor-token optimistic-lock check is deferred into
        # _apply_hlf_core so the LIVE and APPLY paths run it identically (full-
        # blob store): `expected_token` is carried inside proposed_value so the
        # apply path re-checks it against intervening edits.
        from app.services.moderation_service import gate_or_apply
        queued, sub = await gate_or_apply(
            db, user=user,
            module="financials", action="hlf_save",
            entity_id=co.code,
            entity_label=(
                "Высокоуровневая отчётность · "
                f"{co.name_short or co.name_ru or co.code}"
            ),
            company_id=co.id, sector_id=None, year=None,
            payload={
                "code": co.code,
                "years": payload.years,
                "sections": [s.model_dump() for s in payload.sections],
                "currency": payload.currency,
                "unit": payload.unit,
                "expected_token": expected_token,
            },
            editor_token=expected_token,
            diff_summary=(
                f"HLF · {len(payload.sections)} секц. · годы {payload.years}"
            ),
        )
        if queued:
            return None, {
                "queued": True,
                "submission_id": str(sub.id),
                "status": sub.status,
                "message": "Изменение отправлено на модерацию",
            }

        result = await _apply_hlf_core(
            db, company=co, payload=payload, user=user,
            expected_token=expected_token,
        )
        return result, None


# ─── Moderation apply entry point ─────────────────────────────────
# Dispatched by app.services.moderation_apply.financials when an approved
# submission has action "hlf_save". Resolves the company by the code carried in
# proposed_value, attributes the write to the ORIGINAL proposer (not the
# approving moderator), validates the payload schema, and funnels through the
# SAME _apply_hlf_core the live path uses.

async def apply_submission(db: AsyncSession, *, sub, user: User) -> dict:
    from sqlalchemy import select

    pv = dict(sub.proposed_value or {})
    if not pv:
        raise ValueError("proposed_value is empty")

    code = pv.get("code") or sub.target_entity_id
    if not code:
        raise ValueError("hlf apply requires a company code in proposed_value")

    repo = FinancialsRepository(db)
    co = await repo.find_company_by_code(str(code))
    if co is None:
        raise ValueError(f"Company '{code}' no longer exists")

    try:
        payload = HlfSavePayload.model_validate({
            "years": pv.get("years"),
            "sections": pv.get("sections"),
            "currency": pv.get("currency", "UZS"),
            "unit": pv.get("unit", "bln"),
        })
    except Exception as e:  # noqa: BLE001 — surface as apply_error, don't crash
        raise ValueError(
            f"proposed_value does not match HlfSavePayload: {e}"
        ) from e

    # Attribute the write to the proposer; fall back to the approving moderator.
    author: User = user
    if sub.proposer_user_id:
        proposer = (await db.execute(
            select(User).where(User.id == sub.proposer_user_id)
        )).scalar_one_or_none()
        if proposer is not None:
            author = proposer

    result = await _apply_hlf_core(
        db, company=co, payload=payload, user=author,
        expected_token=pv.get("expected_token"),
    )
    return {
        "code": co.code,
        "years": result.get("years"),
        "sections": result.get("sections_count"),
        "rows": result.get("rows_count"),
    }
