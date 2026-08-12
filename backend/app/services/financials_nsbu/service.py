"""NSBU editor use-cases (Pack 7.52/7.53/7.54/7.55).

5 endpoints encapsulated:
  GET    /companies/{code}/nsbu-editor                  schema + values
  PUT    /companies/{code}/nsbu-editor                  save (idempotent)
  GET    /companies/{code}/nsbu-editor/history          last N audit entries
  GET    /companies/{code}/nsbu-editor/template         download XLSX template
  POST   /companies/{code}/nsbu-editor/parse-excel      parse uploaded XLSX

Values are stored in `financial_reports` (one per (year, report_type) with
standard='NSBU', is_detailed=False, source='nsbu-editor') + `financial_lines`.
Customization (custom fields, renames, overrides, manual flags) lives in
`company.extra.nsbu_editor_schema` JSONB.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi import status as http_status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.access import allowed_company_ids
from app.core.audit_chain import append_audit_entry
from app.core.editor_lock import check_editor_token, compute_financials_editor_token
from app.core.security import has_effective_permission
from app.models.audit import AuditLog
from app.models.financial import FinancialLine, FinancialReport
from app.models.user import User
from app.repositories.financials_repository import FinancialsRepository

# Field sets — MUST match useNsbuSchema.ts STANDARD_SCHEMA in frontend.
_NSBU_PL_FIELDS = {
    "revenue", "govGrants", "cogs", "grossProfit", "opProfit", "depreciation",
    "finIncome", "finCost", "forex", "pbt", "tax", "profit", "ebitda",
}
_NSBU_BS_FIELDS = {
    "ppe", "totalNCA", "cash", "totalCA", "accountsReceivable", "totalAssets",
    "equity", "shareCapital", "retainedEarnings",
    "ltBorrowings", "stBorrowings", "accountsPayable", "totalLiabilities",
    "ltBankLoans", "ltOtherLoans", "stBankLoans", "stOtherLoans", "debt",
}

# Field labels for XLSX template. (canonical_id, label_ru, nsbu_code, section)
_NSBU_FIELD_LABELS: dict[str, tuple[str, str, str, str]] = {
    # P&L (ОФР)
    "revenue":       ("revenue",       "Выручка",                                  "010", "pnl"),
    "govGrants":     ("govGrants",     "Господдержка (трансферы)",                 "",    "pnl"),
    "cogs":          ("cogs",          "Себестоимость",                            "020", "pnl"),
    "grossProfit":   ("grossProfit",   "Валовая прибыль (авто)",                   "030", "pnl"),
    "opProfit":      ("opProfit",      "Операционная прибыль",                     "060", "pnl"),
    "depreciation":  ("depreciation",  "Амортизация",                              "070", "pnl"),
    "finIncome":     ("finIncome",     "Доходы от фин. деятельности",              "110", "pnl"),
    "finCost":       ("finCost",       "Расходы от фин. деятельности",             "170", "pnl"),
    "forex":         ("forex",         "Курсовая разница (справочно)",             "180", "pnl"),
    "pbt":           ("pbt",           "Прибыль до налога (авто)",                 "190", "pnl"),
    "tax":           ("tax",           "Налог на прибыль",                         "220", "pnl"),
    "profit":        ("profit",        "Чистая прибыль (авто)",                    "270", "pnl"),
    "ebitda":        ("ebitda",        "EBITDA (авто)",                            "",    "pnl"),
    # Balance Sheet (Баланс)
    "ppe":              ("ppe",              "Основные средства",                  "010", "sofp"),
    "totalNCA":         ("totalNCA",         "Внеоборотные активы (итог)",         "190", "sofp"),
    "cash":             ("cash",             "Денежные средства",                  "320", "sofp"),
    "totalCA":          ("totalCA",          "Оборотные активы (итог)",            "390", "sofp"),
    "accountsReceivable": ("accountsReceivable", "Дебиторская задолженность",       "210", "sofp"),
    "totalAssets":      ("totalAssets",      "ИТОГО Активы (авто)",                "400", "sofp"),
    "equity":           ("equity",           "Собственный капитал",                "480", "sofp"),
    "shareCapital":     ("shareCapital",     "Уставный капитал",                   "410", "sofp"),
    "retainedEarnings": ("retainedEarnings", "Нераспределённая прибыль",           "470", "sofp"),
    "ltBorrowings":     ("ltBorrowings",     "Долгосрочные обязательства",         "590", "sofp"),
    "stBorrowings":     ("stBorrowings",     "Краткосрочные обязательства",        "780", "sofp"),
    "accountsPayable":  ("accountsPayable",  "Кредиторская задолженность",         "601", "sofp"),
    "totalLiabilities": ("totalLiabilities", "ИТОГО Обязательства (авто)",         "",    "sofp"),
    "ltBankLoans":      ("ltBankLoans",      "Долгоср. банковские кредиты",        "7810", "sofp"),
    "ltOtherLoans":     ("ltOtherLoans",     "Долгоср. займы",                     "7820", "sofp"),
    "stBankLoans":      ("stBankLoans",      "Краткоср. банковские кредиты",       "6810", "sofp"),
    "stOtherLoans":     ("stOtherLoans",     "Краткоср. займы",                    "6820", "sofp"),
    "debt":             ("debt",             "Финансовый долг (авто)",             "",    "sofp"),
}


# ─── Payload schemas ──────────────────────────────────────────────

class NsbuCustomFieldDef(BaseModel):
    id: str
    label: str
    section: Optional[str] = None  # 'pnl' | 'sofp'
    autoFormula: Optional[str] = None
    isCustom: Optional[bool] = True
    canonical: Optional[str] = None
    # id строки, ПОСЛЕ которой пользователь вставил эту; без него — в конец секции
    afterId: Optional[str] = None


class NsbuEditorSavePayload(BaseModel):
    """Payload from frontend NsbuEditor.vue save action."""
    values: dict[str, dict[str, Optional[float]]] = Field(default_factory=dict)
    customFields: list[NsbuCustomFieldDef] = Field(default_factory=list)
    renames: dict[str, str] = Field(default_factory=dict)
    formulaOverrides: dict[str, str] = Field(default_factory=dict)
    manualFlags: dict[str, dict[str, bool]] = Field(default_factory=dict)


# ─── Service ──────────────────────────────────────────────────────

@dataclass
class FinancialsNsbuService:
    async def get_schema(
        self, code: str, db: AsyncSession, user: User,
    ) -> dict:
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        if not await has_effective_permission(
            db, user, "financials.view", company_id=co.id,
        ):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        schema = (co.extra or {}).get("nsbu_editor_schema", {}) if co.extra else {}

        rows = await db.execute(
            select(FinancialReport, FinancialLine)
            .join(FinancialLine, FinancialLine.report_id == FinancialReport.id)
            .where(
                FinancialReport.company_id == co.id,
                FinancialReport.standard == "NSBU",
                FinancialReport.is_detailed.is_(False),
                FinancialReport.quarter.is_(None),
            )
        )
        values: dict[str, dict[str, float]] = {}
        for fr, fl in rows.all():
            if fl.value is None:
                continue
            try:
                v = float(fl.value)
            except (TypeError, ValueError):
                continue
            values.setdefault(fl.line_code, {})[str(fr.year)] = v

        editor_token = await compute_financials_editor_token(
            db, company_id=co.id, standard="NSBU",
            consolidated=None, quarter=None,
            schema_updated_at=schema.get("updatedAt"),
        )

        return {
            "code": co.code,
            "values": values,
            "customFields": schema.get("customFields", []),
            "renames": schema.get("renames", {}),
            "formulaOverrides": schema.get("formulaOverrides", {}),
            "manualFlags": schema.get("manualFlags", {}),
            "updatedAt": schema.get("updatedAt"),
            "updatedBy": schema.get("updatedBy"),
            "_editor_token": editor_token,
        }

    async def save(
        self,
        code: str,
        payload: NsbuEditorSavePayload,
        db: AsyncSession,
        user: User,
        *,
        expected_token: Optional[str] = None,
    ) -> tuple[Optional[dict], Optional[dict]]:
        """Save NSBU-editor values + schema customization.

        Returns a 2-tuple `(result, queued)` (mirrors
        financials_reports.save_report):
          - `(result_dict, None)` — normal write-through; `result_dict` carries
            `_editor_token`, which the route lifts into the X-Editor-Token header;
          - `(None, queued_dict)` — moderation intercepted the change; the route
            turns it into HTTP 202 and MUST NOT read X-Editor-Token off it.
        """
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        if not await has_effective_permission(
            db, user, "financials.edit", company_id=co.id,
        ):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN,
                "Permission required: financials.edit",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )

        # Optimistic-lock: recompute the current token from the SAME slice the GET
        # emitted (reading the pre-write schema updatedAt) and 409 on mismatch,
        # BEFORE the moderation gate and any write. No-op when no token was sent.
        prev_schema = (co.extra or {}).get("nsbu_editor_schema", {}) if co.extra else {}
        current_token = await compute_financials_editor_token(
            db, company_id=co.id, standard="NSBU",
            consolidated=None, quarter=None,
            schema_updated_at=prev_schema.get("updatedAt"),
        )
        check_editor_token(
            scope_name=f"financials/nsbu/{code}",
            expected_token=expected_token, current_token=current_token,
        )

        # Moderation gate — AFTER permission + scope + editor-token checks and
        # BEFORE any DB mutation (mirrors financials_reports.save_report). Writes
        # through unchanged until an admin enables 'financials' moderation; then an
        # external author's save is queued as HTTP 202. `expected_token` also rides
        # in the payload so the apply path can re-check it against current state —
        # the schema blob (customFields/renames/…) is last-writer-wins.
        from app.services.moderation_service import gate_or_apply
        queued, sub = await gate_or_apply(
            db, user=user,
            module="financials", action="nsbu_editor_save",
            entity_id=co.code,
            entity_label=(
                f"НСБУ-редактор · {co.code} "
                f"{co.name_short or co.name_ru or ''}"
            ).strip(),
            company_id=co.id, sector_id=None, year=None,
            payload={
                "code": co.code,
                "expected_token": expected_token,
                **payload.model_dump(mode="json"),
            },
            editor_token=expected_token,
            diff_summary=f"Сохранение НСБУ-редактора · {co.code}",
        )
        if queued:
            return None, {
                "queued": True,
                "submission_id": str(sub.id),
                "status": sub.status,
                "message": "Изменение отправлено на модерацию",
            }

        # Live write-through. The token was already validated above, so the shared
        # core skips its apply-path re-check (expected_token=None).
        result = await _apply_nsbu_editor_core(
            db, company=co, payload=payload, user=user, expected_token=None,
        )
        return result, None

    async def get_history(
        self, code: str, db: AsyncSession, user: User, *, limit: int,
    ) -> dict:
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        if not await has_effective_permission(
            db, user, "financials.view", company_id=co.id,
        ):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        scope_ids = await allowed_company_ids(db, user)
        if scope_ids is not None and co.id not in scope_ids:
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "No access",
            )
        q = await db.execute(
            select(AuditLog)
            .where(
                AuditLog.action == "nsbu_editor.save",
                AuditLog.entity_type == "company",
                AuditLog.entity_id == str(co.id),
            )
            .order_by(AuditLog.created_at.desc())
            .limit(limit)
        )
        entries = list(q.scalars().all())
        return {
            "code": co.code,
            "company_name": co.name_short or co.name_ru,
            "total": len(entries),
            "entries": [
                {
                    "id": str(e.id),
                    "at": e.created_at.isoformat() if e.created_at else None,
                    "actor_email": e.actor_email,
                    "diff": e.diff or {},
                    "payload": e.payload or {},
                    "notes": e.notes,
                }
                for e in entries
            ],
        }

    async def download_template(
        self, code: str, years: str, db: AsyncSession, user: User,
    ) -> StreamingResponse:
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        if not await has_effective_permission(
            db, user, "financials.view", company_id=co.id,
        ):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        try:
            year_list = sorted({
                int(y.strip()) for y in years.split(",") if y.strip()
            })
        except ValueError:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                "Invalid years parameter",
            )
        if not year_list:
            year_list = [2021, 2022, 2023, 2024, 2025, 2026]

        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )

        wb = Workbook()
        ws_pl = wb.active
        ws_pl.title = "ОФР"
        ws_bs = wb.create_sheet("Баланс")

        header_font = Font(bold=True, size=11, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF7F77DD")
        auto_fill = PatternFill("solid", fgColor="FFFFFBF0")
        auto_font = Font(italic=True, color="FFD97706")
        border = Border(
            left=Side(style="thin", color="FFE2E8F0"),
            right=Side(style="thin", color="FFE2E8F0"),
            top=Side(style="thin", color="FFE2E8F0"),
            bottom=Side(style="thin", color="FFE2E8F0"),
        )
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def fill_sheet(ws, section):
            ws.cell(
                row=1, column=1,
                value=(
                    f"НСБУ {section} · {co.code} "
                    f"{co.name_short or co.name_ru or ''}"
                ),
            )
            ws.cell(row=1, column=1).font = Font(
                bold=True, size=13, color="FF1E2A4A"
            )
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=3 + len(year_list),
            )
            helper = (
                "Заполняйте числовые поля. Поля, помеченные «авто», "
                "пересчитываются автоматически и не обязательны. "
                "Числа в МЛРД СУМ (например 62,5)."
            )
            ws.cell(row=2, column=1, value=helper).font = Font(
                italic=True, size=9, color="FF94A3B8"
            )
            ws.merge_cells(
                start_row=2, start_column=1,
                end_row=2, end_column=3 + len(year_list),
            )
            ws.cell(row=4, column=1, value="Код")
            ws.cell(row=4, column=2, value="№ строки")
            ws.cell(row=4, column=3, value="Показатель")
            for i, yr in enumerate(year_list):
                ws.cell(row=4, column=4 + i, value=yr)
            for col in range(1, 4 + len(year_list)):
                c = ws.cell(row=4, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border
            row = 5
            for _field_id, (fid, label, nsbu_code, sect) in _NSBU_FIELD_LABELS.items():
                if sect != section:
                    continue
                is_auto = "(авто)" in label
                ws.cell(row=row, column=1, value=fid)
                ws.cell(row=row, column=2, value=nsbu_code)
                ws.cell(row=row, column=3, value=label)
                for col in range(1, 4 + len(year_list)):
                    c = ws.cell(row=row, column=col)
                    c.border = border
                    if col >= 4:
                        c.alignment = center
                    elif col == 3:
                        c.alignment = left
                    else:
                        c.alignment = center
                    if is_auto:
                        c.fill = auto_fill
                        if col == 3:
                            c.font = auto_font
                row += 1
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 45
            for i in range(len(year_list)):
                ws.column_dimensions[chr(ord("D") + i)].width = 14
            ws.freeze_panes = "D5"

        fill_sheet(ws_pl, "pnl")
        fill_sheet(ws_bs, "sofp")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"nsbu_template_{co.code}.xlsx"
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{fname}"',
            },
        )

    async def parse_excel(
        self,
        code: str,
        file: UploadFile,
        db: AsyncSession,
        user: User,
    ) -> dict:
        repo = FinancialsRepository(db)
        co = await repo.find_company_by_code(code)
        if not co:
            raise HTTPException(
                http_status.HTTP_404_NOT_FOUND,
                f"Company '{code}' not found",
            )
        if not await has_effective_permission(
            db, user, "financials.view", company_id=co.id,
        ):
            raise HTTPException(
                http_status.HTTP_403_FORBIDDEN, "Permission required",
            )
        contents = await file.read()
        if not contents:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST, "Empty file",
            )
        from openpyxl import load_workbook
        try:
            wb = load_workbook(
                io.BytesIO(contents), data_only=True, read_only=True,
            )
        except Exception as e:
            raise HTTPException(
                http_status.HTTP_400_BAD_REQUEST,
                f"Cannot parse XLSX: {e}",
            )

        values: dict[str, dict[int, float]] = {}
        parse_log: list[str] = []

        # Build a label → field_id mapping (case-insensitive)
        label_to_field: dict[str, str] = {}
        for _fid, (canonical, label, _code, _sect) in _NSBU_FIELD_LABELS.items():
            clean = label.replace("(авто)", "").strip().lower()
            label_to_field[clean] = canonical
            label_to_field[canonical.lower()] = canonical

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            header_row = None
            year_cols: dict[int, int] = {}
            for r in range(1, min(15, ws.max_row + 1)):
                row_cells = list(ws.iter_rows(
                    min_row=r, max_row=r, values_only=True,
                ))
                if not row_cells:
                    continue
                row_vals = row_cells[0]
                yc_local: dict[int, int] = {}
                for ci, val in enumerate(row_vals):
                    if isinstance(val, int | float) and 1990 < int(val) < 2100:
                        yc_local[ci] = int(val)
                    elif (
                        isinstance(val, str)
                        and val.strip().isdigit()
                        and 1990 < int(val.strip()) < 2100
                    ):
                        yc_local[ci] = int(val.strip())
                if yc_local:
                    header_row = r
                    year_cols = yc_local
                    break
            if not header_row or not year_cols:
                parse_log.append(
                    f"⚠ Лист «{sheet}»: не найдена строка с годами"
                )
                continue
            parse_log.append(
                f"Лист «{sheet}»: заголовок в строке {header_row}, "
                f"годы {sorted(year_cols.values())}"
            )

            rows_parsed = 0
            for r in range(header_row + 1, ws.max_row + 1):
                row_cells = list(ws.iter_rows(
                    min_row=r, max_row=r, values_only=True,
                ))
                if not row_cells:
                    continue
                row_vals = row_cells[0]
                field_id: Optional[str] = None
                for ci in range(min(3, len(row_vals))):
                    v = row_vals[ci]
                    if not isinstance(v, str):
                        continue
                    key = (
                        v.strip().lower().replace("(авто)", "").strip()
                    )
                    if key in label_to_field:
                        field_id = label_to_field[key]
                        break
                if not field_id:
                    continue
                for col_idx, year in year_cols.items():
                    if col_idx >= len(row_vals):
                        continue
                    cell_val = row_vals[col_idx]
                    if cell_val is None or cell_val == "":
                        continue
                    try:
                        num = (
                            float(cell_val) if not isinstance(cell_val, str)
                            else float(
                                cell_val.replace(",", ".").replace(" ", "")
                            )
                        )
                    except (TypeError, ValueError):
                        continue
                    if not (-1e12 < num < 1e12):
                        continue
                    values.setdefault(field_id, {})[year] = num
                rows_parsed += 1
            parse_log.append(f"  → распознано строк: {rows_parsed}")

        out_values = {
            fld: {str(y): v for y, v in ym.items()}
            for fld, ym in values.items()
        }
        return {
            "company_code": co.code,
            "company_name": co.name_short or co.name_ru,
            "values": out_values,
            "fields_count": len(out_values),
            "cells_count": sum(len(ym) for ym in out_values.values()),
            "log": parse_log,
            "filename": file.filename,
        }


# ─── Shared write core + moderation apply ─────────────────────────
# The LIVE save path (FinancialsNsbuService.save) and the moderation APPLY path
# (apply_submission, dispatched from moderation_apply/financials.py) call the SAME
# core, so the two can never drift.

async def _apply_nsbu_editor_core(
    db: AsyncSession,
    *,
    company,
    payload: NsbuEditorSavePayload,
    user: User,
    expected_token: Optional[str] = None,
) -> dict:
    """Perform the NSBU-editor writes (schema JSONB + financial_reports/lines) and
    commit. Attributes writes to `user` (the proposer on the apply path).

    When `expected_token` is truthy (apply path), re-check it against the current
    scope token BEFORE mutating: the schema store is a full-blob, last-writer-wins
    overwrite, so a stale apply would clobber edits made after the author submitted
    (mirrors save_report.expected_prev_checksum on its apply path). The live path
    validates the token before the gate and passes None here to skip the re-check.
    """
    co = company
    now_iso = datetime.now(UTC).isoformat()

    prev_schema = (co.extra or {}).get("nsbu_editor_schema", {}) if co.extra else {}
    if expected_token:
        current_token = await compute_financials_editor_token(
            db, company_id=co.id, standard="NSBU",
            consolidated=None, quarter=None,
            schema_updated_at=prev_schema.get("updatedAt"),
        )
        if expected_token != current_token:
            raise ValueError(
                "Данные НСБУ-редактора изменились после подачи заявки — "
                "применение затёрло бы новые правки. Отклоните заявку и "
                "попросите автора пересоздать её на актуальных данных.",
            )

    # 1. Persist customization to company.extra.nsbu_editor_schema
    extra = dict(co.extra or {})
    extra["nsbu_editor_schema"] = {
        "customFields": [cf.model_dump() for cf in payload.customFields],
        "renames": payload.renames,
        "formulaOverrides": payload.formulaOverrides,
        "manualFlags": payload.manualFlags,
        "updatedAt": now_iso,
        "updatedBy": user.email,
    }
    co.extra = extra

    # 2. Group values by (year, report_type)
    custom_section_by_id: dict[str, str] = {}
    for cf in payload.customFields:
        if cf.section in ("pnl", "sofp"):
            custom_section_by_id[cf.id] = cf.section

    changes_by_report: dict[
        tuple[int, str],
        list[tuple[str, Optional[float], str, Optional[str]]],
    ] = {}

    label_for_field: dict[str, str] = {
        f: f for f in (_NSBU_PL_FIELDS | _NSBU_BS_FIELDS)
    }
    for cf in payload.customFields:
        label_for_field[cf.id] = cf.label
    for fld, renamed in payload.renames.items():
        label_for_field[fld] = renamed

    # canonical mapping per field — for custom fields, allows
    # them to contribute to portfolio aggregations via parent_code.
    canonical_for_field: dict[str, Optional[str]] = {}
    for f in (_NSBU_PL_FIELDS | _NSBU_BS_FIELDS):
        canonical_for_field[f] = f
    for cf in payload.customFields:
        canonical_for_field[cf.id] = cf.canonical or None

    for field, year_map in payload.values.items():
        if field in _NSBU_PL_FIELDS:
            report_type = "PL"
        elif field in _NSBU_BS_FIELDS:
            report_type = "BS"
        elif field in custom_section_by_id:
            report_type = (
                "PL" if custom_section_by_id[field] == "pnl" else "BS"
            )
        else:
            continue
        for year_str, val in year_map.items():
            try:
                year = int(year_str)
            except (TypeError, ValueError):
                continue
            changes_by_report.setdefault((year, report_type), []).append((
                field, val,
                label_for_field.get(field, field),
                canonical_for_field.get(field),
            ))

    # 3. Upsert reports + lines
    reports_created = 0
    reports_updated = 0
    lines_upserted = 0
    lines_deleted = 0

    for (year, report_type), changes in changes_by_report.items():
        rep_q = await db.execute(
            select(FinancialReport).where(
                FinancialReport.company_id == co.id,
                FinancialReport.year == year,
                FinancialReport.quarter.is_(None),
                FinancialReport.standard == "NSBU",
                FinancialReport.report_type == report_type,
                FinancialReport.is_detailed == False,  # noqa: E712
            ).order_by(FinancialReport.updated_at.desc())
        )
        existing_reports = list(rep_q.scalars().all())
        report: Optional[FinancialReport] = None
        for r in existing_reports:
            if r.source == "nsbu-editor":
                report = r
                break
        if report is None and existing_reports:
            report = existing_reports[0]

        if report is None:
            report = FinancialReport(
                company_id=co.id,
                year=year, quarter=None,
                standard="NSBU", report_type=report_type,
                currency="UZS",
                unit_scale=1_000_000_000,
                source="nsbu-editor",
                is_audited=False, is_detailed=False,
                notes=f"Saved via NSBU editor by {user.email} on {now_iso}",
                extra={"editor_version": "p7.52"},
            )
            db.add(report)
            await db.flush()
            reports_created += 1
        else:
            reports_updated += 1
            report.source = report.source or "nsbu-editor"
            report.notes = (
                f"Last edit via NSBU editor by {user.email} on {now_iso}"
            )

        ln_q = await db.execute(
            select(FinancialLine).where(
                FinancialLine.report_id == report.id
            )
        )
        existing_lines = {
            ln.line_code: ln for ln in ln_q.scalars().all()
        }

        for field, val, label, canonical in changes:
            existing = existing_lines.get(field)
            if val is None:
                if existing is not None:
                    await db.delete(existing)
                    lines_deleted += 1
                continue
            decimal_val = Decimal(str(val))
            new_parent = canonical
            if existing is not None:
                existing.value = decimal_val
                existing.line_name = label
                existing.parent_code = new_parent
            else:
                db.add(FinancialLine(
                    report_id=report.id,
                    line_code=field,
                    parent_code=new_parent,
                    line_name=label,
                    value=decimal_val,
                    is_subtotal=False,
                    is_calculated=False,
                    sort_order=0,
                ))
            lines_upserted += 1

    # audit-log entry
    try:
        sample_fields = sorted(set(
            field for changes in changes_by_report.values()
            for field, _, _, _ in changes
        ))[:20]
        await append_audit_entry(
            db,
            actor_id=str(user.id) if user.id else None,
            actor_email=user.email,
            action="nsbu_editor.save",
            entity_type="company",
            entity_id=str(co.id),
            diff={
                "reports_created": reports_created,
                "reports_updated": reports_updated,
                "lines_upserted": lines_upserted,
                "lines_deleted": lines_deleted,
                "fields": sample_fields,
                "years": sorted({
                    y for (y, _) in changes_by_report.keys()
                }),
            },
            payload={
                "company_code": co.code,
                "customFields_count": len(payload.customFields),
                "renames_count": len(payload.renames),
                "formulaOverrides_count": len(payload.formulaOverrides),
            },
            notes=f"NSBU editor save · {co.code}",
        )
    except Exception:
        pass

    await db.commit()

    # Re-issue a fresh token (schema updatedAt just bumped to now_iso, lines
    # committed) so the same open editor can keep saving without a reload.
    new_token = await compute_financials_editor_token(
        db, company_id=co.id, standard="NSBU",
        consolidated=None, quarter=None,
        schema_updated_at=now_iso,
    )
    return {
        "ok": True,
        "saved_at": now_iso,
        "reports_created": reports_created,
        "reports_updated": reports_updated,
        "lines_upserted": lines_upserted,
        "lines_deleted": lines_deleted,
        "_editor_token": new_token,
    }


async def apply_submission(db: AsyncSession, *, sub, user: User) -> dict:
    """Apply an approved NSBU-editor save (moderation dispatch target).

    `moderation_apply/financials.py` routes action="nsbu_editor_save" here.
    Resolves the company by the code carried in `sub.proposed_value`, loads the
    PROPOSER for attribution (falls back to the approving moderator), validates
    the payload against NsbuEditorSavePayload, and calls the shared write core.
    Editor-token re-validation is carried by the payload's `expected_token` and
    performed inside the core (the schema blob is last-writer-wins).
    """
    pv = dict(sub.proposed_value or {})
    if not pv:
        raise ValueError("proposed_value is empty")

    code = pv.get("code") or sub.target_entity_id
    if not code:
        raise ValueError(
            "nsbu_editor_save requires a company code in proposed_value",
        )

    repo = FinancialsRepository(db)
    co = await repo.find_company_by_code(str(code))
    if co is None:
        raise ValueError(f"Company '{code}' no longer exists")

    proposer: Optional[User] = None
    if sub.proposer_user_id:
        proposer = (await db.execute(
            select(User).where(User.id == sub.proposer_user_id)
        )).scalar_one_or_none()
    actor = proposer or user

    expected_token = pv.get("expected_token")
    schema_payload = {
        k: v for k, v in pv.items() if k not in ("code", "expected_token")
    }
    try:
        payload = NsbuEditorSavePayload.model_validate(schema_payload)
    except Exception as e:
        raise ValueError(
            f"proposed_value does not match NsbuEditorSavePayload: {e}",
        ) from e

    result = await _apply_nsbu_editor_core(
        db, company=co, payload=payload, user=actor, expected_token=expected_token,
    )
    return {
        "action": "nsbu_editor_save",
        "code": co.code,
        "reports_created": result["reports_created"],
        "reports_updated": result["reports_updated"],
        "lines_upserted": result["lines_upserted"],
        "lines_deleted": result["lines_deleted"],
    }
