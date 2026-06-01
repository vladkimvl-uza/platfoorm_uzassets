"""FinModel v2 Excel importer — handoff Phase 1.8.

Heuristic NSBU/Excel parser:

  Column A  → row code ("010", "PL_270", "400" ...)
  Column B  → row name (referenced for matching, not used to validate)
  Columns C+ → year values (header row determines which column is which year)

The header row is auto-detected as the first row containing at least one
4-digit year (>= 2000, <= 2100) in cells C+. If a different column layout is
detected, importer returns a structured `mapping` proposal so the frontend
can let the user adjust before commit.

Returns a `PreviewResult` — never writes to DB directly. Frontend reviews +
calls a separate commit endpoint with the chosen `mapping`.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore


def _to_decimal(v) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    if isinstance(v, int | float):
        return Decimal(str(v))
    if isinstance(v, str):
        s = v.strip().replace(" ", "").replace(" ", "").replace(",", ".")
        if not s:
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    return None


def _is_year(v) -> Optional[int]:
    """Return int year if cell looks like a 4-digit year, else None."""
    if isinstance(v, int) and 2000 <= v <= 2100:
        return v
    if isinstance(v, float) and v.is_integer() and 2000 <= v <= 2100:
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if s.isdigit() and len(s) == 4:
            y = int(s)
            if 2000 <= y <= 2100:
                return y
    return None


def parse_excel(
    file_bytes: bytes,
    known_codes: set[str],
    sheet_name: Optional[str] = None,
) -> dict:
    """Read first usable sheet of an .xlsx, return preview structure:

    {
      "sheet": "Лист1",
      "header_row": 3,                         # 1-based row index of year header
      "year_columns": {"C": 2024, "D": 2025},  # detected mapping
      "rows": [
        {"code": "010", "name": "...", "matched": true, "values": {"2024": "100.00", "2025": "110.00"}},
        ...
      ],
      "unmatched_codes": ["XYZ_unknown"],      # codes from Excel that aren't in template
      "conflicts": [],                         # populated by caller, not here
      "warnings": ["Sheet contains 3 unknown rows", ...]
    }
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # Materialize first ~250 rows × first ~30 cols (NSBU forms never exceed this).
    grid: list[list] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=250, max_col=30, values_only=True)):
        grid.append(list(row))
    if not grid:
        return {"sheet": ws.title, "header_row": 0, "year_columns": {}, "rows": [], "unmatched_codes": [], "warnings": ["Лист пустой"]}

    # 1. Find header row: first row with >=1 cell at index >=2 that looks like a year
    header_row_idx: Optional[int] = None
    year_columns: dict[int, int] = {}  # 0-based col idx → year
    for i, row in enumerate(grid):
        years_in_row: dict[int, int] = {}
        for j in range(2, len(row)):
            y = _is_year(row[j])
            if y is not None:
                years_in_row[j] = y
        if years_in_row:
            header_row_idx = i
            year_columns = years_in_row
            break

    warnings: list[str] = []
    if header_row_idx is None:
        # Fallback: assume column C is single-year (current year) — let frontend re-map
        warnings.append("Не найдена строка с годами в шапке. Колонка C принята за текущий год по умолчанию.")
        header_row_idx = 0
        year_columns = {2: 0}

    # 2. Walk data rows below header
    rows_out: list[dict] = []
    unmatched: list[str] = []
    seen_codes: set[str] = set()
    for i in range(header_row_idx + 1, len(grid)):
        row = grid[i]
        if not row or not row[0]:
            continue
        code_raw = row[0]
        if isinstance(code_raw, int | float):
            # Excel often stores NSBU codes as numbers — pad to 3 digits
            code = str(int(code_raw)).zfill(3)
        else:
            code = str(code_raw).strip()
        if not code:
            continue
        if code in seen_codes:
            warnings.append(f"Дубликат кода {code} на строке {i + 1} — пропущен")
            continue
        seen_codes.add(code)
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        values: dict[str, str] = {}
        for col_idx, year in year_columns.items():
            if col_idx >= len(row):
                continue
            v = _to_decimal(row[col_idx])
            if v is not None:
                values[str(year)] = str(v)
        matched = code in known_codes
        if not matched:
            unmatched.append(code)
        rows_out.append({
            "code": code,
            "name": name,
            "matched": matched,
            "row_excel": i + 1,
            "values": values,
        })

    # Column letter map for nicer UI
    def col_letter(j: int) -> str:
        s = ""
        n = j + 1
        while n:
            n, rem = divmod(n - 1, 26)
            s = chr(65 + rem) + s
        return s

    year_columns_lettered = {col_letter(j): y for j, y in year_columns.items()}

    if unmatched:
        warnings.append(f"Не найдено в шаблоне NSBU: {len(unmatched)} строк")

    return {
        "sheet": ws.title,
        "header_row": header_row_idx + 1,
        "year_columns": year_columns_lettered,
        "rows": rows_out,
        "unmatched_codes": unmatched,
        "warnings": warnings,
    }


def build_commit_payload(
    preview: dict,
    selected_years: Optional[list[int]] = None,
    skip_unmatched: bool = True,
) -> list[tuple[int, str, str]]:
    """Convert preview rows into a flat list of (year, code, value_str) tuples
    ready for batch insert. Filters out unmatched + locked-year choices."""
    result: list[tuple[int, str, str]] = []
    years_filter = set(selected_years) if selected_years else None
    for r in preview.get("rows", []):
        if skip_unmatched and not r.get("matched"):
            continue
        code = r["code"]
        for year_str, value_str in r.get("values", {}).items():
            try:
                year = int(year_str)
            except ValueError:
                continue
            if years_filter is not None and year not in years_filter:
                continue
            result.append((year, code, value_str))
    return result
